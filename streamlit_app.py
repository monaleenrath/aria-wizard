"""
streamlit_app.py  —  ARIA Onboarding Wizard
============================================
8-step setup wizard for Board Room Narrator Agent.

    streamlit run streamlit_app.py

Features
--------
• Dark / Light theme toggle (top-right) — persists to ui_config.yaml
• Inline ℹ️ help tips on every technical field
• Floating ARIA help sidebar — FAQ + step-by-step guides
• Step 8 fully automated: GitHub account → repo → PAT → secrets → live

Step 7 renders the *actual* SVG card from svg_generator.py — pixel-identical
to what gets posted to Slack every morning.
"""

from __future__ import annotations

import base64
import hashlib
import json
import re
import sys
import uuid
from datetime import date, datetime as dt, timedelta
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import requests as _req
import streamlit as st
import yaml

# ── page config must be first streamlit call ────────────────────────────────
_ROOT = Path(__file__).parent
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "agent"))

_UI_CONFIG_PATH       = _ROOT / "ui_config.yaml"
_STREAMLIT_CONFIG_DIR = _ROOT / ".streamlit"
_STREAMLIT_TOML       = _STREAMLIT_CONFIG_DIR / "config.toml"

# ── read stored theme BEFORE set_page_config ────────────────────────────────
def _load_ui_config() -> dict:
    try:
        return yaml.safe_load(_UI_CONFIG_PATH.read_text()) or {}
    except Exception:
        return {}

def _save_ui_config(cfg: dict):
    try:
        _UI_CONFIG_PATH.write_text(yaml.dump(cfg, default_flow_style=False))
    except OSError:
        pass  # read-only FS on Streamlit Cloud

def _apply_theme(dark: bool):
    """Write .streamlit/config.toml — Streamlit auto-detects and reloads.
    Silently skips on read-only filesystems (e.g. Streamlit Community Cloud)."""
    try:
        _STREAMLIT_CONFIG_DIR.mkdir(exist_ok=True)
        if dark:
            toml = (
                "[theme]\n"
                'base = "dark"\n'
                'primaryColor = "#F59E0B"\n'
                'backgroundColor = "#0B1220"\n'
                'secondaryBackgroundColor = "#111827"\n'
                'textColor = "#F8FAFC"\n'
            )
        else:
            toml = (
                "[theme]\n"
                'base = "light"\n'
                'primaryColor = "#F59E0B"\n'
            )
        _STREAMLIT_TOML.write_text(toml)
    except OSError:
        pass  # read-only FS on Streamlit Cloud — theme toggle is a no-op

_ui_cfg  = _load_ui_config()
_is_dark = _ui_cfg.get("dark_mode", True)
# Ensure config.toml matches stored preference on first run
if not _STREAMLIT_TOML.exists():
    _apply_theme(_is_dark)

st.set_page_config(
    page_title="ARIA Setup Wizard",
    page_icon="⚡",
    layout="centered",
    initial_sidebar_state="collapsed",
)

TOTAL_STEPS = 8
STEP_LABELS = [
    "Welcome", "Upload Data", "Choose AI", "Discover KPIs",
    "Pick Role", "Preview Card", "Set Delivery", "Go Live",
]

# ════════════════════════════════════════════════════════════════════════════════
# CARD ACCENT THEMES
# ════════════════════════════════════════════════════════════════════════════════

# ── Card visual styles (4 distinct background palettes) ────────────────────────
CARD_STYLES = {
    "dark": {
        "label":   "🌑 Dark",
        "tagline": "Editorial dark — the original ARIA look.",
        "bg":      "#0B1220",
        "surface": "#111827",
        "accent_override": None,
        "swatch":  "#0B1220",
    },
    "navy": {
        "label":   "🌊 Navy",
        "tagline": "Classic navy — calm authority.",
        "bg":      "#1B3B6F",
        "surface": "#1E4A8A",
        "accent_override": None,
        "swatch":  "#1B3B6F",
    },
    "grey": {
        "label":   "🩶 Grey",
        "tagline": "Light slate — clean corporate feel.",
        "bg":      "#D8DEE9",
        "surface": "#C8D0E0",
        "accent_override": None,
        "swatch":  "#D8DEE9",
    },
    "beige": {
        "label":   "☀️ Beige",
        "tagline": "Warm light — stands out in dark feeds.",
        "bg":      "#F5F0E8",
        "surface": "#EDE8DF",
        "accent_override": None,
        "swatch":  "#F5F0E8",
    },
}

# ── Card templates (5 distinct layouts) ────────────────────────────────────────
CARD_TEMPLATES = {
    "editorial": {
        "label":    "📰 Editorial",
        "tagline":  "Three-act: KPIs · Drivers · Action",
        "desc":     "Hero number + sparkline trend, driver bar charts, action box. The original ARIA format.",
        "chart":    "Sparkline",
        "best_for": "CEO · CFO",
    },
    "scorecard": {
        "label":    "🏆 Scorecard",
        "tagline":  "KPI grid tiles + donut chart",
        "desc":     "All KPIs at a glance with RAG status dots. Donut chart shows contribution mix.",
        "chart":    "Donut",
        "best_for": "CFO · VP",
    },
    "story_arc": {
        "label":    "📖 Story Arc",
        "tagline":  "Narrative-first + treemap",
        "desc":     "Big pull-quote headline drives the story. Treemap shows proportional driver breakdown.",
        "chart":    "Treemap",
        "best_for": "CEO · Board",
    },
    "ops_dashboard": {
        "label":    "⚡ Ops Dashboard",
        "tagline":  "Traffic lights + heat map",
        "desc":     "Traffic-light KPI tiles for instant RAG status. Heat map surfaces dimension patterns.",
        "chart":    "Heat Map",
        "best_for": "COO · Operations",
    },
    "board_pack": {
        "label":    "🏛️ Board Pack",
        "tagline":  "Formal slide with donut",
        "desc":     "Clean board-presentation layout. Formal header, KPI row, donut, recommendation box.",
        "chart":    "Donut",
        "best_for": "Board · C-Suite",
    },
}

# ════════════════════════════════════════════════════════════════════════════════
# TIMEFRAME OPTIONS
# ════════════════════════════════════════════════════════════════════════════════

# Each entry defines a data window the user can pick.
# days=None means dynamic (YTD or all-time computed at runtime).
# sparkline_gran: "daily" | "weekly" | "monthly"  — how to bucket the trend line.
# comparison: label shown on the card delta chips.
TIMEFRAME_OPTIONS: list[dict] = [
    {"key": "1d",  "label": "Yesterday",        "short": "1D",  "days": 1,
     "comparison": "vs prior day",       "sparkline_gran": "daily",   "sparkpoints": 7,
     "desc": "Daily ops pulse. Ideal for executives who need last night's numbers first thing."},
    {"key": "wtd", "label": "Week to Date",     "short": "WTD", "days": None,
     "comparison": "vs prior week",      "sparkline_gran": "daily",   "sparkpoints": 7,
     "desc": "Current week performance from Monday to today. Great for weekly targets and field momentum."},
    {"key": "mtd", "label": "Month to Date",    "short": "MTD", "days": None,
     "comparison": "vs prior month",     "sparkline_gran": "daily",   "sparkpoints": 30,
     "desc": "Month-to-date snapshot. The standard window for P&L reviews and leadership reporting."},
    {"key": "qtd", "label": "Quarter to Date",  "short": "QTD", "days": None,
     "comparison": "vs prior quarter",   "sparkline_gran": "weekly",  "sparkpoints": 13,
     "desc": "Current quarter from Q-start to today. Surfaces QoQ performance shifts."},
    {"key": "ytd", "label": "Year to Date",     "short": "YTD", "days": None,
     "comparison": "vs same period LY",  "sparkline_gran": "monthly", "sparkpoints": 12,
     "desc": "Current year progress against the same period last year."},
    {"key": "alltime", "label": "All Time",     "short": "All", "days": None,
     "comparison": "vs prior period",    "sparkline_gran": "monthly", "sparkpoints": 24,
     "desc": "Full historical dataset. Best for baseline benchmarking and long-range trends."},
]

# Suggestion labels only — never auto-applied, shown as a badge on the option card.
# CEO and C-Suite → Yesterday (they want the freshest number in the room)
# Team Lead and Business Analyst → Last 12 Months (operational depth vs BI breadth)
ROLE_TIMEFRAME_DEFAULTS: dict[str, str] = {
    "CEO":                "1d",
    "CFO":                "1d",
    "COO":                "1d",
    "CTO":                "1d",
    "VP":                 "mtd",
    "Director":           "mtd",
    "Sales Head":         "wtd",
    "Senior Manager":     "wtd",
    "Manager":            "wtd",
    "Team Lead":          "ytd",
    "Business Analyst":   "ytd",
    "Operations Head":    "wtd",
}

def _tf_by_key(key: str) -> dict:
    """Return a TIMEFRAME_OPTIONS entry by key, defaulting to 30d."""
    return next((t for t in TIMEFRAME_OPTIONS if t["key"] == key),
                TIMEFRAME_OPTIONS[2])  # 30d

# ════════════════════════════════════════════════════════════════════════════════
# ROLE ROSTER
# ════════════════════════════════════════════════════════════════════════════════

ROLE_GROUPS: dict[str, dict[str, dict]] = {
    "🏛️ C-Suite": {
        "CEO": {
            "title": "Chief Executive Officer",
            "badge": "CEO  ·  EXECUTIVE BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Profit", "Orders", "Margin%"],
            "accent_color": "#F59E0B",
            "driver_focus": ["Category", "Region"],
            "tone": (
                "Strategic and big-picture. Lead with top-line revenue then profitability. "
                "One sharp strategic insight per section. Zero operational detail. "
                "If there is a tension between growth and margin, name it directly."
            ),
        },
        "CFO": {
            "title": "Chief Financial Officer",
            "badge": "CFO  ·  FINANCIAL BRIEFING",
            "primary_kpi": "Profit",
            "kpis": ["Profit", "Margin%", "Sales", "Orders"],
            "accent_color": "#10B981",
            "driver_focus": ["Category", "Region", "Segment"],
            "tone": (
                "Numbers-first. Margin health, cost discipline, profitability drivers. "
                "Flag any compression. One risk and one opportunity per briefing."
            ),
        },
        "COO": {
            "title": "Chief Operating Officer",
            "badge": "COO  ·  OPERATIONS BRIEFING",
            "primary_kpi": "Orders",
            "kpis": ["Orders", "Quantity", "AOV", "Sales"],
            "accent_color": "#2DD4BF",
            "driver_focus": ["Ship Mode", "Sub-Category", "Region"],
            "tone": (
                "Efficiency-focused. Order volume, fulfilment throughput, shipping performance. "
                "Surface bottlenecks. One fix with a clear owner and timeline."
            ),
        },
        "CTO": {
            "title": "Chief Technology Officer",
            "badge": "CTO  ·  TECHNOLOGY BRIEFING",
            "primary_kpi": "Orders",
            "kpis": ["Orders", "Sales", "Quantity", "Margin%"],
            "accent_color": "#8B5CF6",
            "driver_focus": ["Ship Mode", "Sub-Category", "Region"],
            "tone": (
                "Systems-thinking. Surface data quality issues and process breaks. "
                "One technical fix with measurable business impact."
            ),
        },
    },
    "🎯 Leadership": {
        "VP": {
            "title": "Vice President",
            "badge": "VP  ·  LEADERSHIP BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Profit", "Orders", "Quantity"],
            "accent_color": "#A78BFA",
            "driver_focus": ["Category", "Region", "Segment"],
            "tone": (
                "Balanced cross-functional. Cover both growth and efficiency. "
                "Surface tension between revenue and margin. "
                "Recommend one action spanning two functions."
            ),
        },
        "Director": {
            "title": "Director",
            "badge": "DIRECTOR  ·  LEADERSHIP BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Profit", "Orders", "Margin%"],
            "accent_color": "#F472B6",
            "driver_focus": ["Category", "Sub-Category", "Region"],
            "tone": (
                "Execution-focused. What is working, what is stalling. "
                "Bridge strategy to team-level action. One clear priority for the week."
            ),
        },
        "Sales Head": {
            "title": "Sales Head",
            "badge": "SALES  ·  COMMERCIAL BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Orders", "AOV", "Margin%"],
            "accent_color": "#60A5FA",
            "driver_focus": ["Sub-Category", "Region", "Segment"],
            "tone": (
                "Action-oriented. What is moving, what is stalling, where to push harder today. "
                "Name specific regions and segments. One clear move before end of day."
            ),
        },
        "Operations Head": {
            "title": "Operations Head",
            "badge": "OPS  ·  OPERATIONS BRIEFING",
            "primary_kpi": "Orders",
            "kpis": ["Orders", "Quantity", "AOV", "Sales"],
            "accent_color": "#2DD4BF",
            "driver_focus": ["Ship Mode", "Sub-Category", "Region"],
            "tone": (
                "Efficiency-focused. Order volume, fulfilment throughput, shipping performance. "
                "One fix with owner and timeline. Avoid revenue language — speak in units and orders."
            ),
        },
    },
    "⚙️ Management": {
        "Senior Manager": {
            "title": "Senior Manager",
            "badge": "SR MANAGER  ·  BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Orders", "Quantity", "AOV"],
            "accent_color": "#34D399",
            "driver_focus": ["Sub-Category", "Segment", "Region"],
            "tone": (
                "Team-level detail. What specific products or customers are moving the number. "
                "One actionable recommendation the team can act on this week."
            ),
        },
        "Manager": {
            "title": "Manager",
            "badge": "MANAGER  ·  BRIEFING",
            "primary_kpi": "Orders",
            "kpis": ["Orders", "Sales", "Quantity", "AOV"],
            "accent_color": "#22D3EE",
            "driver_focus": ["Sub-Category", "Segment", "Ship Mode"],
            "tone": (
                "Ground-level detail. Specific products, shipments, customer groups. "
                "Two actionable tasks the team can complete today."
            ),
        },
        "Team Lead": {
            "title": "Team Lead",
            "badge": "TEAM LEAD  ·  BRIEFING",
            "primary_kpi": "Orders",
            "kpis": ["Orders", "Quantity", "Sales", "AOV"],
            "accent_color": "#A3E635",
            "driver_focus": ["Sub-Category", "Ship Mode", "Segment"],
            "tone": (
                "Operational and granular. What is shipping, what is delayed, "
                "what needs attention today. No strategic language — just the facts."
            ),
        },
        "Business Analyst": {
            "title": "Business Analyst",
            "badge": "BA  ·  ANALYSIS BRIEFING",
            "primary_kpi": "Sales",
            "kpis": ["Sales", "Profit", "Margin%", "Orders"],
            "accent_color": "#818CF8",
            "driver_focus": ["Sub-Category", "Segment", "Region"],
            "tone": (
                "Data-driven and precise. Lead with the signal, then the root cause, then the so-what. "
                "Call out anomalies and pattern breaks. Frame every insight as a hypothesis with supporting data. "
                "One clear analytical recommendation with a measurable success metric."
            ),
        },
    },
}

# ════════════════════════════════════════════════════════════════════════════════
# HELP CONTENT  —  answers to every technical question a first-time user asks
# ════════════════════════════════════════════════════════════════════════════════

HELP_CONTENT = {
    "slack_channel_id": {
        "q": "How do I find my Slack Channel ID?",
        "a": (
            "**Method 1 (easiest):** \n"
            "1. Open Slack in a web browser (not the desktop app)\n"
            "2. Click on the channel you want ARIA to post to\n"
            "3. Look at the URL — it ends with something like `/C0B5U431C3U`\n"
            "4. That last part (starting with **C**) is your Channel ID\n\n"
            "**Method 2 (desktop app):**\n"
            "1. Right-click the channel name\n"
            "2. Select **Copy link**\n"
            "3. Paste it anywhere — the last segment after the final `/` is the ID"
        ),
    },
    "slack_bot_token": {
        "q": "How do I get a Slack Bot Token?",
        "a": (
            "1. Go to **[api.slack.com/apps](https://api.slack.com/apps)**\n"
            "2. Click **Create New App → From scratch**\n"
            "3. Name it **ARIA** and pick your workspace\n"
            "4. Go to **OAuth & Permissions**\n"
            "5. Under **Bot Token Scopes** add: `chat:write`, `files:write`, `channels:read`\n"
            "6. Click **Install to Workspace** at the top\n"
            "7. Copy the **Bot User OAuth Token** — it starts with `xoxb-`\n\n"
            "⚠️ Keep this token private — treat it like a password."
        ),
    },
    "teams_webhook": {
        "q": "How do I create a Teams Incoming Webhook?",
        "a": (
            "1. Open **Microsoft Teams**\n"
            "2. Go to the channel where you want ARIA to post\n"
            "3. Click the **⋯ (three dots)** next to the channel name\n"
            "4. Select **Connectors** (or **Manage channel → Connectors**)\n"
            "5. Search for **Incoming Webhook** and click **Add**\n"
            "6. Give it a name (e.g. **ARIA Daily Briefing**) and click **Create**\n"
            "7. Copy the webhook URL — it starts with `https://your-org.webhook.office.com/...`\n\n"
            "ℹ️ You need **Channel Manager** permissions to add connectors."
        ),
    },
    "gemini_key": {
        "q": "How do I get a free Gemini API key?",
        "a": (
            "1. Go to **[aistudio.google.com/app/apikey](https://aistudio.google.com/app/apikey)**\n"
            "2. Sign in with your Google account\n"
            "3. Click **Create API key**\n"
            "4. Copy the key and paste it below\n\n"
            "**Free tier limits:**\n"
            "- 1,500 requests/day\n"
            "- 15 requests/minute\n"
            "- No credit card required\n"
            "- Works from India and most countries\n\n"
            "ARIA uses one request per role per morning, so the free tier is more than enough."
        ),
    },
    "github_account": {
        "q": "Why do I need a GitHub account?",
        "a": (
            "GitHub hosts the ARIA agent code and runs it automatically every morning via "
            "**GitHub Actions** — a free scheduling service.\n\n"
            "Think of it as: your ARIA lives in GitHub, and GitHub wakes it up at 9 AM every day "
            "to fetch your data, generate the card, and post it to Slack.\n\n"
            "**GitHub Free plan** is completely free and sufficient for ARIA:\n"
            "- 2,000 Action minutes/month (ARIA uses ~2 min/day)\n"
            "- Unlimited public and private repositories\n"
            "- No credit card required"
        ),
    },
    "github_pat": {
        "q": "What is a Personal Access Token and how do I create one?",
        "a": (
            "A Personal Access Token (PAT) is like a password that lets ARIA's setup wizard "
            "write files and secrets to your GitHub repo automatically.\n\n"
            "**Quickest way — direct link:**\n"
            "👉 Go straight to **[github.com/settings/tokens](https://github.com/settings/tokens)**\n\n"
            "**Step by step (if the link doesn't work):**\n"
            "1. Click your **profile picture** (your avatar, top-right corner of GitHub)\n"
            "2. In the dropdown that appears, click **Settings**\n"
            "   ⚠️ This is YOUR ACCOUNT settings — NOT the repo settings tab\n"
            "3. In the left sidebar, scroll all the way to the **bottom**\n"
            "4. Click **Developer settings** (very last item)\n"
            "5. Click **Personal access tokens → Tokens (classic)**\n"
            "6. Click **Generate new token (classic)**\n"
            "7. Give it a name like **ARIA**, set expiration to **90 days**\n"
            "8. Tick these two boxes: ✅ **repo** (select all under it), ✅ **workflow**\n"
            "9. Scroll to the bottom → click **Generate token**\n"
            "10. Copy the token immediately — it starts with `ghp_`\n\n"
            "⚠️ You only see the token **once**. Copy it before closing the page."
        ),
    },
    "github_repo": {
        "q": "How do I find my repository URL?",
        "a": (
            "1. Go to **[github.com](https://github.com)** and sign in\n"
            "2. Click the repository you created for ARIA\n"
            "3. The URL in your browser is your repo URL, e.g.:\n"
            "   `https://github.com/yourname/board-room-narrator`\n\n"
            "If you haven't created a repo yet, click **New** (green button on your GitHub home page), "
            "name it **board-room-narrator**, set it to **Private**, and click **Create repository**."
        ),
    },
    "github_secrets": {
        "q": "What are GitHub Secrets and why do they matter?",
        "a": (
            "GitHub Secrets are encrypted variables that your ARIA agent reads at runtime — "
            "without exposing them in your code.\n\n"
            "Instead of putting your Slack token directly in a file (risky!), you store it as "
            "a Secret. GitHub injects it as an environment variable when the workflow runs.\n\n"
            "ARIA needs these secrets:\n"
            "- `GEMINI_API_KEY` — your AI key (if using Gemini)\n"
            "- `SLACK_BOT_TOKEN` — your Slack bot token\n"
            "- `SLACK_CHANNEL_ID` — your Slack channel ID\n"
            "- `TEAMS_WEBHOOK_URL` — your Teams webhook (if using Teams)\n\n"
            "ARIA's setup wizard sets all of these automatically when you provide your PAT."
        ),
    },
    "email_delivery": {
        "q": "How does ARIA email delivery work?",
        "a": (
            "ARIA sends the briefing card as an email attachment using Gmail.\n\n"
            "**What you need:**\n"
            "- A Gmail address to send *from* (can be a dedicated `aria.briefings@gmail.com`)\n"
            "- A Gmail **App Password** (not your regular password)\n\n"
            "**To create an App Password:**\n"
            "1. Go to your Google Account → **Security**\n"
            "2. Under 'How you sign in to Google', enable **2-Step Verification** if not already on\n"
            "3. Then go to **App Passwords** (search 'App Passwords' in your Google Account)\n"
            "4. Select app: **Mail**, device: **Other** → type 'ARIA'\n"
            "5. Google gives you a 16-character password — copy it\n\n"
            "ARIA stores this as a GitHub Secret — it never appears in your code."
        ),
    },
    "google_sheets": {
        "q": "How do I share my Google Sheet so ARIA can read it?",
        "a": (
            "1. Open your Google Sheet\n"
            "2. Click **Share** (top right)\n"
            "3. Click **Change to anyone with the link**\n"
            "4. Set the role to **Viewer** (not Editor)\n"
            "5. Click **Copy link** and paste it in ARIA\n\n"
            "ℹ️ ARIA only reads the sheet — it never writes to it. "
            "Viewer access is all it needs."
        ),
    },
}


# ════════════════════════════════════════════════════════════════════════════════
# USER ACCOUNTS — GOOGLE SHEETS BACKEND
# ════════════════════════════════════════════════════════════════════════════════

ADMIN_EMAIL    = "monaleenrath1401@gmail.com"
_ARIA_SHEET_ID = "1eEfyVh4VmFlmV6ZzqJkJ_ZTaogoKrE8O7_1pebFHW34"

_USER_HDRS  = ["email", "name", "password_hash", "created_at", "last_login", "is_admin"]
_ACT_HDRS   = ["timestamp", "user_email", "user_name", "action_type", "details"]
_SCHED_HDRS = ["schedule_id", "user_email", "user_name", "repo_url", "github_username",
               "role_name", "schedule_type", "delivery_channels", "timezone",
               "delivery_hour", "created_at", "status"]


def _get_gs():
    """Return authenticated gspread client, or None on failure.

    Supports three secret formats in Streamlit Cloud:
      1. [GOOGLE_CREDS_JSON] TOML section  ← recommended, no escaping issues
      2. GOOGLE_CREDS_JSON = '''{ raw json }'''  (literal TOML string)
      3. Local google_creds.json file (dev only)
    """
    try:
        import gspread, json
        SCOPES = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
            "https://www.googleapis.com/auth/spreadsheets",
        ]

        info = None

        # ── Option 1: TOML section [GOOGLE_CREDS_JSON] ──────────────────────
        # Streamlit parses this into a dict-like AttrDict — no JSON needed.
        try:
            secret = st.secrets.get("GOOGLE_CREDS_JSON")
            if secret is not None and hasattr(secret, "keys"):
                info = {k: str(v) for k, v in secret.items()}
        except Exception:
            pass

        # ── Option 2: raw JSON string ────────────────────────────────────────
        if info is None:
            raw = ""
            try:
                val = st.secrets.get("GOOGLE_CREDS_JSON", "")
                if val and not hasattr(val, "keys"):
                    raw = str(val)
            except Exception:
                pass
            if not raw:
                creds_path = _ROOT / "google_creds.json"
                if creds_path.exists():
                    raw = creds_path.read_text(encoding="utf-8")
            if raw:
                # Re-escape any control chars that TOML may have injected
                # into JSON string values (from basic """ strings).
                def _fix_ctrl(s: str) -> str:
                    result, in_str, i = [], False, 0
                    while i < len(s):
                        c = s[i]
                        if not in_str:
                            result.append(c)
                            if c == '"':
                                in_str = True
                        else:
                            if c == '\\':
                                result.append(c)
                                i += 1
                                if i < len(s):
                                    result.append(s[i])
                            elif c == '"':
                                result.append(c)
                                in_str = False
                            elif c == '\n':
                                result.append('\\n')
                            elif c == '\r':
                                result.append('\\r')
                            elif c == '\t':
                                result.append('\\t')
                            else:
                                result.append(c)
                        i += 1
                    return ''.join(result)
                info = json.loads(_fix_ctrl(raw))

        if info is None:
            st.session_state["_gs_last_error"] = (
                "GOOGLE_CREDS_JSON not found — add it to Streamlit Cloud secrets"
            )
            return None

        # gspread 6+ — service_account_from_dict (authorize() is deprecated)
        gc = gspread.service_account_from_dict(info, scopes=SCOPES)
        return gc
    except Exception as e:
        st.session_state["_gs_last_error"] = f"{type(e).__name__}: {e}"
        return None


def _gs_ensure_tab(sh, name: str, headers: list):
    """Return worksheet, creating it with headers if missing."""
    try:
        return sh.worksheet(name)
    except Exception:
        ws = sh.add_worksheet(title=name, rows=2000, cols=len(headers))
        ws.append_row(headers)
        return ws


@st.cache_data(ttl=20, show_spinner=False)
def _gs_read(tab: str) -> list[dict]:
    """Read all rows from a tab as list of dicts (cached 20 s)."""
    try:
        gc = _get_gs()
        if not gc:
            return []
        ws = gc.open_by_key(_ARIA_SHEET_ID).worksheet(tab)
        return ws.get_all_records()
    except Exception:
        return []


def _gs_append(tab: str, row: list, headers: list) -> bool:
    """Append a row, creating the tab if needed."""
    try:
        gc = _get_gs()
        if not gc:
            return False
        sh = gc.open_by_key(_ARIA_SHEET_ID)
        ws = _gs_ensure_tab(sh, tab, headers)
        ws.append_row(row, value_input_option="RAW")
        _gs_read.clear()          # bust cache
        return True
    except Exception as e:
        st.session_state["_gs_last_error"] = f"_gs_append({tab}): {type(e).__name__}: {e}"
        return False


def _gs_update_cell(tab: str, key_col: str, key_val: str,
                    upd_col: str, upd_val: str) -> bool:
    """Find a row by key_col == key_val and update upd_col."""
    try:
        gc = _get_gs()
        if not gc:
            return False
        ws  = gc.open_by_key(_ARIA_SHEET_ID).worksheet(tab)
        hdr = ws.row_values(1)
        ki  = hdr.index(key_col) + 1
        ui  = hdr.index(upd_col) + 1
        for r, row in enumerate(ws.get_all_values()[1:], start=2):
            if row[ki - 1] == key_val:
                ws.update_cell(r, ui, upd_val)
                _gs_read.clear()
                return True
        return False
    except Exception:
        return False


# ── Auth helpers ─────────────────────────────────────────────────────────── #

def _hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()


def auth_login(email: str, password: str) -> dict | None:
    """Return user dict on success, None on failure. Also updates last_login."""
    ph = _hash_pw(password)
    for u in _gs_read("ARIA_Users"):
        if u.get("email", "").lower() == email.lower() and u.get("password_hash") == ph:
            _gs_update_cell("ARIA_Users", "email", email.lower(),
                            "last_login", dt.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
            return u
    return None


def auth_register(email: str, name: str, password: str) -> tuple[bool, str]:
    """Register a new user. Returns (success, message)."""
    if not email or not name or not password:
        return False, "All fields are required."
    existing = _gs_read("ARIA_Users")
    if any(u.get("email", "").lower() == email.lower() for u in existing):
        return False, "An account with this email already exists."
    now      = dt.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    is_admin = 1 if email.lower() == ADMIN_EMAIL.lower() else 0
    ok = _gs_append("ARIA_Users",
                    [email.lower(), name, _hash_pw(password), now, now, is_admin],
                    _USER_HDRS)
    if ok:
        log_activity(email.lower(), name, "register", "New account created")
        return True, "Account created! You can now sign in."
    err = st.session_state.get("_gs_last_error", "")
    detail = f"\n\n`{err}`" if err else ""
    return False, f"Could not save account — please try again in a moment.{detail}"


def log_activity(user_email: str, user_name: str, action: str, details: str = ""):
    """Write one activity row (fire-and-forget)."""
    try:
        _gs_append("ARIA_Activity",
                   [dt.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                    user_email, user_name, action, details],
                   _ACT_HDRS)
    except Exception:
        pass


def log_schedule(user_email: str, user_name: str, repo_url: str, gh_user: str,
                 role_name: str, schedule_type: str, channels: str,
                 timezone: str, delivery_hour: int) -> str:
    """Log a new schedule row and return its schedule_id."""
    sid = str(uuid.uuid4())[:8].upper()
    _gs_append("ARIA_Schedules",
               [sid, user_email, user_name, repo_url, gh_user,
                role_name, schedule_type, channels, timezone,
                str(delivery_hour), dt.utcnow().strftime("%Y-%m-%d %H:%M:%S"), "active"],
               _SCHED_HDRS)
    return sid


def delete_schedule(schedule_id: str):
    _gs_update_cell("ARIA_Schedules", "schedule_id", schedule_id, "status", "deleted")
    log_activity(
        st.session_state.get("user", {}).get("email", ""),
        st.session_state.get("user", {}).get("name", ""),
        "delete_schedule", f"Deleted schedule {schedule_id}",
    )


def change_password(email: str, old_pw: str, new_pw: str) -> tuple[bool, str]:
    """Change password after verifying the old one."""
    if not auth_login(email, old_pw):
        return False, "Current password is incorrect."
    ok = _gs_update_cell("ARIA_Users", "email", email.lower(),
                          "password_hash", _hash_pw(new_pw))
    return (True, "Password updated.") if ok else (False, "Update failed — try again.")


# ════════════════════════════════════════════════════════════════════════════════
# HELP SYSTEM UI
# ════════════════════════════════════════════════════════════════════════════════

def help_tip(key: str, label: str = "ℹ️"):
    """Render a sleek info popover button next to a field."""
    content = HELP_CONTENT.get(key, {})
    if not content:
        return
    with st.popover(label, use_container_width=False):
        st.markdown(f"**{content['q']}**")
        st.markdown(content["a"])


def render_help_sidebar():
    """ARIA help assistant in the sidebar — FAQ + step guides."""
    with st.sidebar:
        st.markdown("### ⚡ ARIA Help")
        st.caption("Ask ARIA anything about the setup.")

        # FAQ quick links
        faq_options = [c["q"] for c in HELP_CONTENT.values()]
        chosen = st.selectbox("Frequently asked questions", ["— Select a question —"] + faq_options,
                              label_visibility="collapsed")
        if chosen != "— Select a question —":
            key = next(k for k, v in HELP_CONTENT.items() if v["q"] == chosen)
            st.markdown(f"**{HELP_CONTENT[key]['q']}**")
            st.markdown(HELP_CONTENT[key]["a"])

        st.divider()
        st.markdown("**Quick links**")
        st.markdown("- [Slack API apps](https://api.slack.com/apps)")
        st.markdown("- [Gemini free key](https://aistudio.google.com/app/apikey)")
        st.markdown("- [GitHub sign up](https://github.com/signup)")
        st.markdown("- [GitHub Actions docs](https://docs.github.com/en/actions)")
        st.markdown("- [Teams webhooks guide](https://learn.microsoft.com/en-us/microsoftteams/platform/webhooks-and-connectors/how-to/add-incoming-webhook)")

        st.divider()
        st.caption("Still stuck? Open an issue on the ARIA GitHub repo.")


# ════════════════════════════════════════════════════════════════════════════════
# AUTH SCREEN — login / register
# ════════════════════════════════════════════════════════════════════════════════

def screen_auth():
    """Full-page login / register screen shown when no user is in session."""
    # ── Email history via localStorage (persists across sessions in same browser) ── #
    # JS writes aria_email_history (JSON array) to localStorage on every login.
    # On page load, JS reads it and stuffs it into a hidden Streamlit text input
    # so Python can read it via session state.
    st.markdown("""
    <script>
    (function() {
        const KEY = 'aria_email_history';
        // Read existing history and push into hidden input
        function pushToStreamlit() {
            const raw = localStorage.getItem(KEY) || '[]';
            const input = window.parent.document.querySelector(
                'input[data-testid="stTextInput-aria_email_bridge"]'
            );
            if (input && input.value !== raw) {
                input.value = raw;
                input.dispatchEvent(new Event('input', {bubbles: true}));
            }
        }
        // Expose save function for use after login
        window._ariaEmailHistorySave = function(email) {
            const raw = localStorage.getItem(KEY) || '[]';
            let hist = [];
            try { hist = JSON.parse(raw); } catch(e) {}
            hist = hist.filter(e => e !== email);
            hist.unshift(email);
            if (hist.length > 5) hist = hist.slice(0, 5);
            localStorage.setItem(KEY, JSON.stringify(hist));
        };
        // Push on load + after short delay (Streamlit re-renders asynchronously)
        pushToStreamlit();
        setTimeout(pushToStreamlit, 800);
    })();
    </script>
    """, unsafe_allow_html=True)

    # Hidden bridge input — JS writes localStorage value here; Python reads it
    _ls_raw = st.text_input("__bridge__", key="aria_email_bridge",
                             label_visibility="collapsed", value="")
    _email_history: list[str] = []
    if _ls_raw:
        try:
            import json as _json
            _email_history = _json.loads(_ls_raw)
            if not isinstance(_email_history, list):
                _email_history = []
        except Exception:
            _email_history = []

    # "Remember me" still works as before (query param for cross-device)
    _qp_email = st.query_params.get("rem", "")
    if _qp_email and not st.session_state.get("_auth_rem_email"):
        st.session_state["_auth_rem_email"] = _qp_email
    _remembered_email = st.session_state.get("_auth_rem_email", "") or _qp_email
    # Merge remembered email into history if not already present
    if _remembered_email and _remembered_email not in _email_history:
        _email_history = [_remembered_email] + _email_history

    _aria_title_color = "#FFFFFF" if _is_dark else "#111827"
    st.markdown(f"""
    <style>
      /* Hide Streamlit's default top toolbar lines */
      header[data-testid="stHeader"] {{ display: none !important; }}
      #MainMenu {{ display: none !important; }}
      footer {{ display: none !important; }}
      .block-container {{ padding-top: 2rem !important; }}
    </style>
    <div style="text-align:center;padding:24px 0 12px">
      <div style="font-size:56px;margin-bottom:8px">⚡</div>
      <h1 style="font-size:52px;font-weight:900;letter-spacing:6px;margin:0;color:{_aria_title_color}">ARIA</h1>
      <p style="font-size:14px;color:#6B7280;margin:8px 0 0;letter-spacing:1px">
        Autonomous Report &amp; Insight AI Agent
      </p>
    </div>
    """, unsafe_allow_html=True)

    _, center, _ = st.columns([1, 2, 1])
    with center:
        tab_in, tab_up = st.tabs(["🔑 Sign In", "✨ Create Account"])

        # ── Sign In ──────────────────────────────────────────────────────── #
        with tab_in:
            # Seed email field from remembered value on first render only
            if "login_email" not in st.session_state:
                st.session_state.login_email = _remembered_email

            # Apply suggestion-button selection BEFORE widget renders (avoids
            # StreamlitAPIException from setting a widget key post-render)
            _prefill = st.session_state.pop("_email_prefill", None)
            if _prefill is not None:
                st.session_state.login_email = _prefill

            email_in = st.text_input(
                "Email address", key="login_email",
                placeholder="you@example.com",
            )

            # ── Email suggestions (history-based autocomplete) ────────────── #
            typed = (email_in or "").strip().lower()
            if _email_history:
                if typed:
                    # Filter history to entries that contain what's typed
                    matches = [
                        e for e in _email_history
                        if typed in e.lower() and e.lower() != typed
                    ]
                else:
                    # Nothing typed yet → show full history
                    matches = _email_history
                if matches:
                    st.caption("Recent accounts — click to fill:")
                    btn_cols = st.columns(min(len(matches), 3))
                    for idx, match in enumerate(matches[:3]):
                        with btn_cols[idx]:
                            if st.button(
                                match, key=f"email_hint_{match}",
                                use_container_width=True,
                            ):
                                # Stage the value; applied before widget renders next run
                                st.session_state._email_prefill = match
                                st.rerun()

            pw_in       = st.text_input("Password", type="password", key="login_pw")
            remember_me = st.checkbox(
                "Remember me", value=bool(_remembered_email), key="remember_me",
                help="Auto-fills your email on your next visit to this browser.",
            )
            if st.button("Sign In →", type="primary", use_container_width=True, key="btn_login"):
                if not email_in or not pw_in:
                    st.error("Please enter your email and password.")
                else:
                    with st.spinner("Signing in…"):
                        user = auth_login(email_in.strip(), pw_in)
                    if user:
                        st.session_state.user = user
                        log_activity(user["email"], user["name"], "login", "Signed in")
                        _e = email_in.strip()
                        # Save email to browser localStorage (persists across sessions)
                        st.markdown(
                            f"<script>window._ariaEmailHistorySave && "
                            f"window._ariaEmailHistorySave('{_e}');</script>",
                            unsafe_allow_html=True,
                        )
                        if remember_me:
                            st.session_state["_auth_rem_email"] = _e
                            st.query_params["rem"] = _e
                        else:
                            st.session_state.pop("_auth_rem_email", None)
                            st.query_params.pop("rem", None)
                        st.rerun()
                    else:
                        st.error("Incorrect email or password.")

        # ── Create Account ────────────────────────────────────────────────── #
        with tab_up:
            name_up  = st.text_input("Your full name", key="reg_name",
                                      placeholder="Mona Rath")
            email_up = st.text_input("Email address", key="reg_email",
                                      placeholder="you@example.com")
            pw_up    = st.text_input("Password (min 6 chars)", type="password",
                                      key="reg_pw")
            pw_up2   = st.text_input("Confirm password", type="password",
                                      key="reg_pw2")
            if st.button("Create Account →", type="primary",
                         use_container_width=True, key="btn_register"):
                if pw_up != pw_up2:
                    st.error("Passwords don't match.")
                elif len(pw_up) < 6:
                    st.error("Password must be at least 6 characters.")
                else:
                    with st.spinner("Creating account…"):
                        ok, msg = auth_register(email_up.strip(), name_up.strip(), pw_up)
                    if ok:
                        st.success(msg + " Switch to the Sign In tab.")
                    else:
                        st.error(msg)

        st.divider()


# ════════════════════════════════════════════════════════════════════════════════
# ACCOUNT SIDEBAR — left panel shown when user is logged in
# ════════════════════════════════════════════════════════════════════════════════

def render_account_sidebar():
    """Left sidebar: user avatar, account links, sign-out."""
    user       = st.session_state.get("user", {})
    name       = user.get("name", "User")
    email      = user.get("email", "")
    is_admin   = bool(user.get("is_admin", 0))
    first      = name.split()[0] if name else "User"
    initials   = "".join(w[0].upper() for w in name.split()[:2]) or "U"
    accent_col = "#10B981"

    with st.sidebar:
        # Avatar + name
        _admin_badge = (
            '<div style="display:inline-block;background:#F59E0B20;color:#F59E0B;'
            'font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;'
            'margin-top:6px">ADMIN</div>'
            if is_admin else ""
        )
        st.markdown(
            f'<div style="text-align:center;padding:10px 0 8px">'
            f'<div style="width:52px;height:52px;border-radius:50%;'
            f'background:{accent_col};display:inline-flex;'
            f'align-items:center;justify-content:center;'
            f'color:#fff;font-size:22px;font-weight:800">{initials}</div>'
            f'<div style="font-weight:700;font-size:15px;margin-top:10px">{first}</div>'
            f'<div style="font-size:11px;color:#9CA3AF;margin-top:2px">{email}</div>'
            f'{_admin_badge}'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.divider()

        if st.button("📊 My Dashboard", use_container_width=True, key="sb_dashboard"):
            st.session_state.aria_screen = "dashboard"
            st.rerun()
        if st.button("⚡ Run Wizard", use_container_width=True, key="sb_wizard"):
            # Clear wizard state, preserve user
            _u = st.session_state.user
            for k in list(st.session_state.keys()):
                if k not in ("user",):
                    st.session_state.pop(k, None)
            st.session_state.user = _u
            st.session_state.aria_screen = "wizard"
            st.session_state.step = 1
            st.rerun()
        if is_admin:
            if st.button("🛡️ Admin Panel", use_container_width=True, key="sb_admin"):
                st.session_state.aria_screen = "admin"
                st.rerun()

        st.divider()

        # Theme toggle in sidebar
        ui_cfg  = _load_ui_config()
        is_dark = ui_cfg.get("dark_mode", True)
        icon    = "☀️ Light mode" if is_dark else "🌙 Dark mode"
        if st.button(icon, key="sb_theme", use_container_width=True):
            ui_cfg["dark_mode"] = not is_dark
            _save_ui_config(ui_cfg)
            _apply_theme(not is_dark)
            st.rerun()

        if st.button("🚪 Sign Out", use_container_width=True,
                     type="secondary", key="sb_signout"):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

        st.divider()

        # ── Help section (collapsible) ────────────────────────────────────── #
        with st.expander("⚡ ARIA Help", expanded=False):
            faq_options = [c["q"] for c in HELP_CONTENT.values()]
            chosen = st.selectbox(
                "Frequently asked questions",
                ["— Select a question —"] + faq_options,
                key="help_faq_sel_sb",
                label_visibility="collapsed",
            )
            if chosen != "— Select a question —":
                key = next(k for k, v in HELP_CONTENT.items() if v["q"] == chosen)
                st.markdown(f"**{HELP_CONTENT[key]['q']}**")
                st.markdown(HELP_CONTENT[key]["a"])
            st.divider()
            st.markdown(
                "**Quick links**\n"
                "- [Slack API apps](https://api.slack.com/apps)\n"
                "- [Gemini free key](https://aistudio.google.com/app/apikey)\n"
                "- [GitHub sign up](https://github.com/signup)\n"
                "- [GitHub Actions docs](https://docs.github.com/en/actions)\n"
                "- [Teams webhooks](https://learn.microsoft.com/en-us/microsoftteams/platform/webhooks-and-connectors/how-to/add-incoming-webhook)"
            )


# ════════════════════════════════════════════════════════════════════════════════
# HELP BUTTON — floating right-side popover (replaces left-sidebar help)
# ════════════════════════════════════════════════════════════════════════════════

def render_help_button():
    """Render a ❓ help popover anchored to the top-right of the main area."""
    _, hcol = st.columns([9, 1])
    with hcol:
        with st.popover("❓", use_container_width=False):
            st.markdown("### ⚡ ARIA Help")
            faq_options = [c["q"] for c in HELP_CONTENT.values()]
            chosen = st.selectbox("Frequently asked questions",
                                  ["— Select a question —"] + faq_options,
                                  key="help_faq_sel",
                                  label_visibility="collapsed")
            if chosen != "— Select a question —":
                key = next(k for k, v in HELP_CONTENT.items() if v["q"] == chosen)
                st.markdown(f"**{HELP_CONTENT[key]['q']}**")
                st.markdown(HELP_CONTENT[key]["a"])
            st.divider()
            st.markdown("**Quick links**")
            st.markdown(
                "- [Slack API apps](https://api.slack.com/apps)\n"
                "- [Gemini free key](https://aistudio.google.com/app/apikey)\n"
                "- [GitHub sign up](https://github.com/signup)\n"
                "- [GitHub Actions docs](https://docs.github.com/en/actions)\n"
                "- [Teams webhooks](https://learn.microsoft.com/en-us/microsoftteams/platform/webhooks-and-connectors/how-to/add-incoming-webhook)"
            )


def render_theme_toggle():
    """Theme toggle button — top right of main area."""
    ui_cfg  = _load_ui_config()
    is_dark = ui_cfg.get("dark_mode", True)
    icon    = "☀️ Light" if is_dark else "🌙 Dark"

    # Align to top-right via columns trick
    _, tcol = st.columns([6, 1])
    if tcol.button(icon, key="theme_toggle", help="Switch between dark and light mode"):
        new_dark = not is_dark
        ui_cfg["dark_mode"] = new_dark
        _save_ui_config(ui_cfg)
        _apply_theme(new_dark)
        st.rerun()


# ════════════════════════════════════════════════════════════════════════════════
# DOMAIN DETECTION + KPI CATALOG
# ════════════════════════════════════════════════════════════════════════════════

# ── Domain signal keywords ─────────────────────────────────────────────────── #
_DOMAIN_SIGNALS: dict[str, list[str]] = {
    "qsr":          ["wait_time","service_time","drive_thru","drivethrough","prep_time",
                     "order_time","ticket_time","table_turn","speed_of_service","sos",
                     "window_time","queue_time","dine_in","takeaway","kds","throughput_time"],
    "satisfaction": ["nps","csat","satisfaction","cleanliness","accuracy","behaviour",
                     "behavior","complaint","feedback","friendliness","survey","review",
                     "net_promoter","food_quality","staff_score","hygiene","ambiance"],
    "banking":      ["loan","deposit","account","balance","interest","credit","debit",
                     "mortgage","overdraft","npl","emi","repayment","disbursement","casa",
                     "nii","fee_income","branch","atm","current_account"],
    "insurance":    ["claim","premium","policy","coverage","deductible","underwriting",
                     "loss_ratio","indemnity","reinsurance","renewal","gwp"],
    "pharma":       ["prescription","drug","molecule","clinical","trial","dosage",
                     "patient","adverse_event","efficacy","formulary","rx","sku_pharma"],
    "hr":           ["employee","headcount","attrition","salary","hire","tenure",
                     "leave","absenteeism","engagement","payroll","workforce","ctc",
                     "resignation","joining","onboard"],
    "marketing":    ["impression","click","conversion","ctr","cpm","cpc","reach",
                     "engagement","campaign","ad_spend","media_spend","lead","mql","sql",
                     "funnel","acquisition","attribution","roas"],
    "operations":   ["utilization","defect","downtime","yield","cycle_time","capacity",
                     "oee","throughput","scrap","rework","production","machine"],
    "logistics":    ["shipment","delivery","freight","carrier","tracking","warehouse",
                     "fulfillment","otd","transit","dispatch","consignment","parcel"],
    "hospitality":  ["room","occupancy","adr","revpar","booking","checkin","checkout",
                     "stay","guest","reservation","hotel","property"],
    "aviation":     ["flight","passenger","load_factor","departure","arrival","route",
                     "origin","destination","aircraft","cabin","delay","on_time","rpk",
                     "ask","cancelled","boarding","fleet","sector","airfare","fare"],
    "retail":       ["sale","revenue","product","sku","category","inventory","order",
                     "customer","discount","return","margin","sub_category","segment"],
}

_DOMAIN_META: dict[str, tuple[str, str]] = {
    "qsr":          ("🍔", "QSR / Food Service"),
    "satisfaction": ("⭐", "Customer Satisfaction"),
    "banking":      ("🏦", "Banking & Finance"),
    "insurance":    ("🛡️",  "Insurance"),
    "pharma":       ("💊", "Pharma & Healthcare"),
    "hr":           ("👥", "Human Resources"),
    "marketing":    ("📣", "Marketing & Growth"),
    "operations":   ("⚙️",  "Operations / Manufacturing"),
    "logistics":    ("🚚", "Logistics & Supply Chain"),
    "hospitality":  ("🏨", "Hospitality"),
    "aviation":     ("✈️",  "Airlines & Aviation"),
    "retail":       ("🛒", "Retail & E-Commerce"),
    "generic":      ("📊", "Business Analytics"),
}

def detect_domain(df: pd.DataFrame) -> str:
    """Score every column name against domain keyword lists; return winning domain."""
    all_cols = " ".join(
        c.lower().replace(" ", "_").replace("-", "_") for c in df.columns
    )
    scores = {
        d: sum(1 for kw in kws if kw in all_cols)
        for d, kws in _DOMAIN_SIGNALS.items()
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "generic"


# ── Direct column → KPI patterns (all domains) ────────────────────────────── #
_DIRECT_PATTERNS: list[dict] = [
    # ── Universal ─────────────────────────────────────────────────────────── #
    {"pattern": r"(sale|revenue|gmv|turnover|gross.?revenue|net.?sales|income(?!_tax))",
     "name": "Revenue",         "agg": "sum",     "format": "currency",
     "description": "Total revenue / sales generated in the period."},
    {"pattern": r"(^profit$|net.?profit|net.?income|ebit(?!da)|ebitda|earnings|margin.?amount)",
     "name": "Profit",          "agg": "sum",     "format": "currency",
     "description": "Net profit or gross margin amount."},
    {"pattern": r"(order.?id|order.?no|transaction.?id|invoice.?id|ticket.?id|receipt.?id)",
     "name": "Transactions",    "agg": "nunique", "format": "integer",
     "description": "Unique orders / transactions processed."},
    {"pattern": r"(^qty$|quantity|units?.?sold|^volume$|pieces?|items?.?sold)",
     "name": "Units Sold",      "agg": "sum",     "format": "integer",
     "description": "Total units / items sold or shipped."},
    {"pattern": r"(customer.?id|customer.?no|client.?id|member.?id|user.?id|guest.?id)",
     "name": "Customers",       "agg": "nunique", "format": "integer",
     "description": "Unique customers / clients served."},
    {"pattern": r"(discount|rebate|coupon.?amount)",
     "name": "Discounts",       "agg": "sum",     "format": "currency",
     "description": "Total discount value applied."},
    {"pattern": r"(^cost$|cogs|cost.?of.?goods|^expense$|total.?cost)",
     "name": "Cost",            "agg": "sum",     "format": "currency",
     "description": "Total cost or cost of goods sold."},
    {"pattern": r"(return.?qty|return.?id|refund.?id|returned|^returns$)",
     "name": "Returns",         "agg": "sum",     "format": "integer",
     "description": "Total units returned or refunded."},
    # ── QSR / Food Service ─────────────────────────────────────────────────── #
    {"pattern": r"(wait.?time|queue.?time|lobby.?wait|customer.?wait)",
     "name": "Avg Wait Time",   "agg": "mean",    "format": "minutes",
     "description": "Average customer wait time (minutes). Benchmark: < 3 min."},
    {"pattern": r"(service.?time|speed.?of.?service|sos.?time|total.?service)",
     "name": "Avg Service Time","agg": "mean",    "format": "minutes",
     "description": "Average end-to-end service time per order."},
    {"pattern": r"(drive.?thru|drive.?through|dt.?time|window.?time|drivethrough)",
     "name": "Drive-Thru Speed","agg": "mean",    "format": "minutes",
     "description": "Average drive-thru completion time. Industry target: < 4 min."},
    {"pattern": r"(prep.?time|kitchen.?time|cook.?time|kds.?time|fulfil.?time)",
     "name": "Avg Prep Time",   "agg": "mean",    "format": "minutes",
     "description": "Average kitchen / prep time per order."},
    {"pattern": r"(order.?accuracy|accuracy.?rate|correct.?order|accuracy.?pct)",
     "name": "Order Accuracy %","agg": "mean",    "format": "percent",
     "description": "% of orders fulfilled correctly without errors."},
    {"pattern": r"(table.?turn|turnover.?rate|seat.?utiliz|covers.?per)",
     "name": "Table Turnover",  "agg": "mean",    "format": "decimal",
     "description": "Average table turns per service period."},
    # ── Customer Satisfaction ─────────────────────────────────────────────── #
    {"pattern": r"(nps|net.?promoter|promoter.?score)",
     "name": "NPS Score",       "agg": "mean",    "format": "score",
     "description": "Net Promoter Score (-100 to +100). > 50 is excellent."},
    {"pattern": r"(csat|satisfaction.?score|sat.?score|customer.?sat(?!isfaction.?id))",
     "name": "CSAT Score",      "agg": "mean",    "format": "score",
     "description": "Customer Satisfaction Score. Typically 1–5 or 1–10."},
    {"pattern": r"(^rating$|star.?rating|review.?score|overall.?score|avg.?score)",
     "name": "Avg Rating",      "agg": "mean",    "format": "score",
     "description": "Average customer rating across all interactions."},
    {"pattern": r"(cleanliness|clean.?score|hygiene.?score|sanitation)",
     "name": "Cleanliness Score","agg": "mean",   "format": "score",
     "description": "Average cleanliness / hygiene audit score."},
    {"pattern": r"(cashier.?behav|staff.?behav|friendliness|staff.?score|employee.?rating|crew.?score)",
     "name": "Staff Score",     "agg": "mean",    "format": "score",
     "description": "Average staff behaviour / friendliness score."},
    {"pattern": r"(complaint|complaint.?count|issue.?count|escalation.?count)",
     "name": "Complaints",      "agg": "sum",     "format": "integer",
     "description": "Total customer complaints / escalations raised."},
    {"pattern": r"(food.?quality|quality.?score|product.?quality|item.?quality)",
     "name": "Quality Score",   "agg": "mean",    "format": "score",
     "description": "Average food / product quality score."},
    {"pattern": r"(ambiance|atmosphere.?score|environment.?score)",
     "name": "Ambiance Score",  "agg": "mean",    "format": "score",
     "description": "Average ambiance / environment score."},
    # ── Banking & Finance ─────────────────────────────────────────────────── #
    {"pattern": r"(loan.?amount|loan.?disburs|credit.?disburs|^advance$)",
     "name": "Loans Disbursed", "agg": "sum",     "format": "currency",
     "description": "Total loan / credit amount disbursed."},
    {"pattern": r"(^deposit$|total.?deposit|saving.?balance|casa.?balance)",
     "name": "Deposits",        "agg": "sum",     "format": "currency",
     "description": "Total deposits collected."},
    {"pattern": r"(account.?id|account.?no|^acct|current.?account.?id)",
     "name": "Accounts",        "agg": "nunique", "format": "integer",
     "description": "Unique accounts opened / active."},
    {"pattern": r"(npl|non.?perform|default.?amount|bad.?debt)",
     "name": "NPL Amount",      "agg": "sum",     "format": "currency",
     "description": "Non-performing / defaulted loan amount."},
    {"pattern": r"(interest.?income|net.?interest|^nii$)",
     "name": "Interest Income",  "agg": "sum",    "format": "currency",
     "description": "Net interest income earned."},
    {"pattern": r"(fee.?income|fee.?revenue|non.?interest.?income)",
     "name": "Fee Income",      "agg": "sum",     "format": "currency",
     "description": "Non-interest / fee-based income."},
    # ── Insurance ─────────────────────────────────────────────────────────── #
    {"pattern": r"(^premium$|gwp|gross.?written|written.?premium)",
     "name": "Gross Premium",   "agg": "sum",     "format": "currency",
     "description": "Total gross written premium collected."},
    {"pattern": r"(claim.?amount|claims.?paid|loss.?amount|incurred.?loss)",
     "name": "Claims Paid",     "agg": "sum",     "format": "currency",
     "description": "Total claims amount paid out."},
    {"pattern": r"(policy.?id|policy.?no|policy.?count|^policies$)",
     "name": "Policies",        "agg": "nunique", "format": "integer",
     "description": "Total active policies."},
    {"pattern": r"(renewal.?rate|retention.?rate|persistency)",
     "name": "Renewal Rate %",  "agg": "mean",    "format": "percent",
     "description": "Policy renewal / retention rate."},
    # ── Pharma & Healthcare ───────────────────────────────────────────────── #
    {"pattern": r"(prescription.?count|rx.?count|script.?count|^prescriptions$)",
     "name": "Prescriptions",   "agg": "sum",     "format": "integer",
     "description": "Total prescriptions written / dispensed."},
    {"pattern": r"(patient.?id|patient.?count|^patients$)",
     "name": "Patients",        "agg": "nunique", "format": "integer",
     "description": "Unique patients treated."},
    {"pattern": r"(adverse.?event|side.?effect|ae.?count|^adr$)",
     "name": "Adverse Events",  "agg": "sum",     "format": "integer",
     "description": "Reported adverse events or side effects."},
    {"pattern": r"(fill.?rate|dispensing.?rate|compliance.?rate|pdc)",
     "name": "Fill Rate %",     "agg": "mean",    "format": "percent",
     "description": "Prescription fill / patient compliance rate."},
    # ── HR ────────────────────────────────────────────────────────────────── #
    {"pattern": r"(employee.?id|emp.?id|staff.?id|^headcount$|workforce.?id)",
     "name": "Headcount",       "agg": "nunique", "format": "integer",
     "description": "Total employee headcount."},
    {"pattern": r"(new.?hire|joining.?id|onboard.?id|hire.?date.?id)",
     "name": "New Hires",       "agg": "sum",     "format": "integer",
     "description": "Number of new employees onboarded."},
    {"pattern": r"(attrition|resign|termination.?count|exit.?count)",
     "name": "Attritions",      "agg": "sum",     "format": "integer",
     "description": "Employee exits / voluntary resignations."},
    {"pattern": r"(^salary$|compensation|payroll.?amount|ctc|gross.?pay)",
     "name": "Total Payroll",   "agg": "sum",     "format": "currency",
     "description": "Total payroll / compensation spend."},
    {"pattern": r"(tenure.?years|years.?of.?service|seniority.?months)",
     "name": "Avg Tenure (yrs)","agg": "mean",    "format": "decimal",
     "description": "Average employee tenure in years."},
    {"pattern": r"(leave.?days|absent.?days|sick.?days|absenteeism.?count)",
     "name": "Leave Days",      "agg": "sum",     "format": "integer",
     "description": "Total leave / absenteeism days taken."},
    # ── Marketing & Growth ─────────────────────────────────────────────────── #
    {"pattern": r"(impression|ad.?view|page.?view|view.?count)",
     "name": "Impressions",     "agg": "sum",     "format": "integer",
     "description": "Total ad / content impressions served."},
    {"pattern": r"(^clicks?$|link.?click|cta.?click|ad.?click)",
     "name": "Clicks",          "agg": "sum",     "format": "integer",
     "description": "Total link / ad clicks."},
    {"pattern": r"(^conversion|purchase.?event|sign.?up.?count|signup.?count|lead.?convert)",
     "name": "Conversions",     "agg": "sum",     "format": "integer",
     "description": "Total conversions / sign-ups achieved."},
    {"pattern": r"(ad.?spend|media.?spend|marketing.?spend|campaign.?cost|total.?spend)",
     "name": "Ad Spend",        "agg": "sum",     "format": "currency",
     "description": "Total marketing / advertising spend."},
    {"pattern": r"(^leads?$|lead.?id|^mql$|^sql$|prospect.?id)",
     "name": "Leads",           "agg": "nunique", "format": "integer",
     "description": "Total marketing / sales qualified leads generated."},
    {"pattern": r"(engagement.?count|like|share|comment|reaction)",
     "name": "Engagements",     "agg": "sum",     "format": "integer",
     "description": "Total social / content engagements."},
    # ── Operations / Manufacturing ─────────────────────────────────────────── #
    {"pattern": r"(unit.?produc|output.?count|pieces.?produc|finished.?goods)",
     "name": "Units Produced",  "agg": "sum",     "format": "integer",
     "description": "Total units produced / finished goods output."},
    {"pattern": r"(^defect$|reject.?count|scrap.?qty|defective.?unit|rework.?count)",
     "name": "Defects",         "agg": "sum",     "format": "integer",
     "description": "Total defective / rejected units."},
    {"pattern": r"(downtime.?hrs|machine.?idle|breakdown.?hrs|unplanned.?stop)",
     "name": "Downtime (hrs)",  "agg": "sum",     "format": "decimal",
     "description": "Total machine / line downtime in hours."},
    {"pattern": r"(cycle.?time|takt.?time|process.?duration|avg.?prod.?time)",
     "name": "Avg Cycle Time",  "agg": "mean",    "format": "minutes",
     "description": "Average production cycle time per unit."},
    {"pattern": r"(utiliz.?pct|capacity.?used|^oee$|availability.?pct)",
     "name": "Utilization %",   "agg": "mean",    "format": "percent",
     "description": "Average asset / line utilisation rate."},
    # ── Logistics & Supply Chain ──────────────────────────────────────────── #
    {"pattern": r"(shipment.?id|parcel.?id|delivery.?id|consignment.?id)",
     "name": "Shipments",       "agg": "nunique", "format": "integer",
     "description": "Total shipments dispatched."},
    {"pattern": r"(on.?time.?flag|otd.?flag|on.?time.?delivery|delivered.?on.?time)",
     "name": "On-Time Deliveries","agg": "sum",   "format": "integer",
     "description": "Total shipments delivered on or before promised date."},
    {"pattern": r"(freight.?cost|shipping.?cost|delivery.?cost|carrier.?charge)",
     "name": "Freight Cost",    "agg": "sum",     "format": "currency",
     "description": "Total freight / shipping cost incurred."},
    {"pattern": r"(transit.?days|delivery.?days|lead.?time.?days|avg.?transit)",
     "name": "Avg Transit Time","agg": "mean",    "format": "decimal",
     "description": "Average transit / lead time in days."},
    # ── Airlines & Aviation ───────────────────────────────────────────────── #
    {"pattern": r"(^passenger|total.?pax|pax.?count|^pax$|economy.?pax|business.?pax)",
     "name": "Passengers",        "agg": "sum",     "format": "integer",
     "description": "Total passengers carried across all flights."},
    {"pattern": r"(^revenue.*eur|total.?revenue|flight.?revenue|fare.?revenue)",
     "name": "Revenue (EUR)",     "agg": "sum",     "format": "currency",
     "description": "Total flight revenue generated in the period."},
    {"pattern": r"(profit.*eur|^profit$|net.?profit|flight.?profit)",
     "name": "Profit (EUR)",      "agg": "sum",     "format": "currency",
     "description": "Net profit after all operating costs."},
    {"pattern": r"(total.?cost.*eur|^total.?cost$|operating.?cost|fuel.*crew)",
     "name": "Total Cost (EUR)",  "agg": "sum",     "format": "currency",
     "description": "Total operating costs including fuel, crew and fees."},
    {"pattern": r"(load.?factor|seat.?factor|cabin.?factor|lf.?pct)",
     "name": "Load Factor %",     "agg": "mean",    "format": "percent",
     "description": "Avg % of seats filled — key airline efficiency metric."},
    {"pattern": r"(on.?time.?flag|otd.?flag|on_time_flag|on.?time.?perf)",
     "name": "On-Time Rate",      "agg": "mean",    "format": "percent",
     "description": "% of flights departing / arriving on schedule."},
    {"pattern": r"(^delay|delay.?min|avg.?delay|departure.?delay|arrival.?delay)",
     "name": "Avg Delay (min)",   "agg": "mean",    "format": "decimal",
     "description": "Average flight delay in minutes across all services."},
    {"pattern": r"(^cancelled$|cancellation.?flag|flight.?cancel|cancel.?count)",
     "name": "Cancellations",     "agg": "sum",     "format": "integer",
     "description": "Total flights cancelled in the period."},
    {"pattern": r"(^rpk$|revenue.?passenger.?km|revenue.?pax.?km)",
     "name": "RPK",               "agg": "sum",     "format": "integer",
     "description": "Revenue Passenger Kilometres — traffic volume measure."},
    {"pattern": r"(^ask$|available.?seat.?km|avail.?seat.?km)",
     "name": "ASK",               "agg": "sum",     "format": "integer",
     "description": "Available Seat Kilometres — total capacity offered."},
    {"pattern": r"(fuel.?cost.*eur|^fuel.?cost$|fuel.?expense)",
     "name": "Fuel Cost (EUR)",   "agg": "sum",     "format": "currency",
     "description": "Total fuel expenditure — largest airline cost driver."},
    {"pattern": r"(^flights?$|flight.?count|num.?flights|total.?flight|sector.?count)",
     "name": "Flights Operated",  "agg": "sum",     "format": "integer",
     "description": "Total flight sectors operated in the period."},
    # ── Hospitality ───────────────────────────────────────────────────────── #
    {"pattern": r"(room.?night|rooms.?occupied|occupied.?room|room.?sold)",
     "name": "Room Nights",     "agg": "sum",     "format": "integer",
     "description": "Total room-nights sold."},
    {"pattern": r"(booking.?id|reservation.?id|checkin.?id|stay.?id)",
     "name": "Bookings",        "agg": "nunique", "format": "integer",
     "description": "Total bookings / reservations confirmed."},
    {"pattern": r"(^adr$|avg.?daily.?rate|room.?rate|rack.?rate)",
     "name": "ADR",             "agg": "mean",    "format": "currency",
     "description": "Average Daily Rate per occupied room."},
    {"pattern": r"(revpar|revenue.?per.?avail)",
     "name": "RevPAR",          "agg": "mean",    "format": "currency",
     "description": "Revenue Per Available Room — core hotel yield metric."},
    {"pattern": r"(length.?of.?stay|stay.?nights|nights.?stayed|los)",
     "name": "Avg Stay (nights)","agg": "mean",   "format": "decimal",
     "description": "Average guest stay duration in nights."},
]

# ── Derived / computed KPI patterns ───────────────────────────────────────── #
# Added automatically when both num_pattern and den_pattern match a column.
_DERIVED_PATTERNS: list[dict] = [
    # ── Universal ratios ──────────────────────────────────────────────────── #
    {"num_pattern": r"(^profit$|net.?profit|net.?income|ebit(?!da)|ebitda|earnings|margin.?amount)",
     "den_pattern": r"(sale|revenue|gmv|turnover|net.?sales|income(?!_tax))",
     "name": "Profit Margin %", "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Net profit as % of revenue. Tracks pricing power and cost efficiency."},
    {"num_pattern": r"(sale|revenue|gmv|net.?sales)",
     "den_pattern": r"(order.?id|transaction.?id|invoice.?id|ticket.?id|receipt.?id)",
     "den_agg": "nunique",
     "name": "Avg Order Value", "scale": 1, "format": "currency",
     "formula_tmpl": "SUM({num}) ÷ COUNT DISTINCT({den})",
     "description": "Average revenue per transaction. Rising AOV = better basket size."},
    {"num_pattern": r"(sale|revenue|gmv|net.?sales)",
     "den_pattern": r"(customer.?id|client.?id|member.?id|user.?id|guest.?id)",
     "den_agg": "nunique",
     "name": "Revenue / Customer", "scale": 1, "format": "currency",
     "formula_tmpl": "SUM({num}) ÷ COUNT DISTINCT({den})",
     "description": "Average revenue per unique customer — tracks customer value."},
    {"num_pattern": r"(discount|rebate)",
     "den_pattern": r"(sale|revenue|gmv|net.?sales)",
     "name": "Discount Rate %",  "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Discounts as % of gross sales. High rates compress margins."},
    {"num_pattern": r"(^cost$|cogs|cost.?of.?goods|^expense$)",
     "den_pattern": r"(sale|revenue|gmv|net.?sales)",
     "name": "Cost-to-Revenue %","scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Operating cost as % of revenue. Lower = more efficient."},
    {"num_pattern": r"(return.?qty|^returns$|refund.?qty|returned)",
     "den_pattern": r"(^qty$|quantity|units?.?sold|pieces?)",
     "name": "Return Rate %",    "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "% of units sold that were returned. Signals fulfilment or quality issues."},
    # ── Banking ───────────────────────────────────────────────────────────── #
    {"num_pattern": r"(npl|non.?perform|default.?amount|bad.?debt)",
     "den_pattern": r"(loan.?amount|loan.?disburs|credit.?disburs|^advance$)",
     "name": "NPL Ratio %",      "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Non-performing loans as % of total book. Key credit quality indicator."},
    {"num_pattern": r"(loan.?amount|loan.?disburs|^advance$)",
     "den_pattern": r"(^deposit$|total.?deposit|saving.?balance)",
     "name": "Loan-to-Deposit %","scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Loans as % of deposits. Measures funding stability and liquidity."},
    # ── Insurance ─────────────────────────────────────────────────────────── #
    {"num_pattern": r"(claim.?amount|claims.?paid|loss.?amount|incurred.?loss)",
     "den_pattern": r"(^premium$|gwp|gross.?written|written.?premium)",
     "name": "Loss Ratio %",     "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Claims paid as % of premium. Core insurance profitability metric."},
    # ── Marketing ─────────────────────────────────────────────────────────── #
    {"num_pattern": r"(^clicks?$|link.?click|ad.?click)",
     "den_pattern": r"(impression|ad.?view|view.?count)",
     "name": "CTR %",            "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Click-Through Rate — % of impressions that resulted in a click."},
    {"num_pattern": r"(^conversion|sign.?up.?count|lead.?convert)",
     "den_pattern": r"(^clicks?$|link.?click|ad.?click)",
     "name": "Conversion Rate %","scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "% of clicks that converted. Measures funnel efficiency."},
    {"num_pattern": r"(sale|revenue|gmv)",
     "den_pattern": r"(ad.?spend|media.?spend|marketing.?spend)",
     "name": "ROAS",             "scale": 1, "format": "decimal",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Return on Ad Spend — revenue generated per $1 of marketing spend."},
    {"num_pattern": r"(ad.?spend|media.?spend|marketing.?spend)",
     "den_pattern": r"(^conversion|sign.?up.?count|purchase.?event)",
     "name": "Cost per Acquisition","scale": 1, "format": "currency",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Average cost to acquire one customer or conversion."},
    # ── Operations ────────────────────────────────────────────────────────── #
    {"num_pattern": r"(^defect$|reject.?count|scrap.?qty|defective.?unit)",
     "den_pattern": r"(unit.?produc|output.?count|finished.?goods)",
     "name": "Defect Rate %",    "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "% of units produced with defects. Target < 1%."},
    # ── Logistics ─────────────────────────────────────────────────────────── #
    {"num_pattern": r"(on.?time.?flag|otd.?flag|delivered.?on.?time)",
     "den_pattern": r"(shipment.?id|parcel.?id|delivery.?id|consignment.?id)",
     "den_agg": "nunique",
     "name": "On-Time Delivery %","scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ COUNT DISTINCT({den})",
     "description": "% of shipments delivered on time. Industry target > 95%."},
    {"num_pattern": r"(freight.?cost|shipping.?cost|delivery.?cost)",
     "den_pattern": r"(shipment.?id|parcel.?id|delivery.?id)",
     "den_agg": "nunique",
     "name": "Cost per Shipment", "scale": 1, "format": "currency",
     "formula_tmpl": "SUM({num}) ÷ COUNT DISTINCT({den})",
     "description": "Average freight cost per shipment dispatched."},
    # ── HR ────────────────────────────────────────────────────────────────── #
    {"num_pattern": r"(attrition|resign|termination.?count|exit.?count)",
     "den_pattern": r"(employee.?id|emp.?id|staff.?id|^headcount$|workforce.?id)",
     "den_agg": "nunique",
     "name": "Attrition Rate %", "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ COUNT DISTINCT({den})",
     "description": "% of workforce that left in the period. High attrition signals engagement risk."},
    # ── Aviation ─────────────────────────────────────────────────────────── #
    {"num_pattern": r"(profit.*eur|^profit$|net.?profit|flight.?profit)",
     "den_pattern": r"(^revenue.*eur|total.?revenue|flight.?revenue)",
     "name": "Profit Margin %",   "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Net profit as % of revenue — core airline financial health metric."},
    {"num_pattern": r"(^revenue.*eur|total.?revenue|flight.?revenue)",
     "den_pattern": r"(^passenger|total.?pax|pax.?count|^pax$)",
     "name": "Revenue / Pax",     "scale": 1, "format": "currency",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "Average revenue per passenger — measures yield per seat sold."},
    {"num_pattern": r"(^rpk$|revenue.?passenger.?km)",
     "den_pattern": r"(^ask$|available.?seat.?km)",
     "name": "Load Factor (RPK/ASK)", "scale": 1, "format": "percent",
     "formula_tmpl": "SUM({num}) ÷ SUM({den})",
     "description": "RPK÷ASK — efficiency of capacity utilisation."},
]


# ════════════════════════════════════════════════════════════════════════════════
# DATA UTILITIES
# ════════════════════════════════════════════════════════════════════════════════

def detect_kpis(df: pd.DataFrame) -> list:
    """
    Domain-aware KPI detection.
    1. Match _DIRECT_PATTERNS against column names → direct KPIs
    2. Match _DERIVED_PATTERNS against detected column pairs → derived KPIs
    3. Fallback: any unmatched numeric column gets a generic SUM KPI
    """
    numeric_cols  = set(df.select_dtypes(include=[np.number]).columns.tolist())
    suggestions: list = []
    used_names:  set  = set()
    col_map: dict     = {}   # normalised_name → original col name

    for col in df.columns:
        norm = col.lower().replace(" ", "_").replace("-", "_").replace(".", "_")
        col_map[norm] = col

    def _find_col(pattern: str) -> str | None:
        """Return first original col whose normalised name matches pattern."""
        for norm, orig in col_map.items():
            if re.search(pattern, norm):
                return orig
        return None

    # ── 1. Direct KPIs ────────────────────────────────────────────────────── #
    for h in _DIRECT_PATTERNS:
        col = _find_col(h["pattern"])
        if col is None:
            continue
        agg = h["agg"]
        if agg in ("sum", "mean", "median", "max", "min") and col not in numeric_cols:
            continue
        if h["name"] in used_names:
            continue
        used_names.add(h["name"])
        if agg == "nunique":
            formula = f"COUNT DISTINCT({col})"
        elif agg == "mean":
            formula = f"AVG({col})"
        elif agg == "sum":
            formula = f"SUM({col})"
        else:
            formula = f"{agg.upper()}({col})"
        suggestions.append({
            "column": col, "suggested_name": h["name"], "user_name": h["name"],
            "formula": formula, "agg": agg, "format": h["format"],
            "description": h["description"], "kpi_type": "direct", "enabled": True,
        })

    # ── 2. Derived KPIs ───────────────────────────────────────────────────── #
    for d in _DERIVED_PATTERNS:
        if d["name"] in used_names:
            continue
        num_col = _find_col(d["num_pattern"])
        den_col = _find_col(d["den_pattern"])
        if num_col is None or den_col is None:
            continue
        if num_col not in numeric_cols:
            continue
        den_agg = d.get("den_agg", "sum")
        if den_agg == "sum" and den_col not in numeric_cols:
            continue
        used_names.add(d["name"])
        formula = d["formula_tmpl"].format(num=num_col, den=den_col)
        suggestions.append({
            "column": "DERIVED",
            "num_col": num_col,
            "den_col": den_col,
            "den_agg": den_agg,
            "scale": d.get("scale", 1),
            "suggested_name": d["name"], "user_name": d["name"],
            "formula": formula, "agg": "ratio", "format": d["format"],
            "description": d["description"], "kpi_type": "derived", "enabled": True,
        })

    # ── 3. Generic numeric fallback ───────────────────────────────────────── #
    # If fewer than 3 KPIs detected, surface any remaining numeric columns
    if len(suggestions) < 3:
        for col in df.columns:
            if col not in numeric_cols:
                continue
            if any(s["column"] == col for s in suggestions):
                continue
            label = col.replace("_", " ").replace("-", " ").title()
            if label in used_names:
                continue
            used_names.add(label)
            suggestions.append({
                "column": col, "suggested_name": label, "user_name": label,
                "formula": f"SUM({col})", "agg": "sum", "format": "decimal",
                "description": f"Sum of {col}.", "kpi_type": "direct", "enabled": True,
            })
            if len(suggestions) >= 6:
                break

    return suggestions


# ════════════════════════════════════════════════════════════════════════════════
# LLM-BASED KPI DETECTION  (Gemini 2.5 Flash → free LLM fallback → error)
# ════════════════════════════════════════════════════════════════════════════════

def _kpi_cache_key(df: pd.DataFrame) -> tuple[str, str]:
    """Return (domain, cache_key) where cache_key = domain + column fingerprint."""
    import hashlib
    domain   = detect_domain(df)
    col_hash = hashlib.md5("|".join(sorted(df.columns.tolist())).encode()).hexdigest()[:10]
    return domain, f"{domain}_{col_hash}"


def _read_kpi_cache() -> dict:
    cache_path = _ROOT / "aria_kpi_cache.json"
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _write_kpi_cache(key: str, domain: str, domain_label: str,
                     domain_emoji: str, kpis: list,
                     role_kpi_map: dict | None = None) -> None:
    from datetime import datetime as _dt
    cache      = _read_kpi_cache()
    entry      = cache.get(key, {})   # preserve any existing fields
    entry.update({
        "domain":        domain,
        "domain_label":  domain_label,
        "domain_emoji":  domain_emoji,
        "generated_at":  _dt.now().isoformat(),
        "kpis":          kpis,
    })
    if role_kpi_map is not None:
        entry["role_kpi_map"] = role_kpi_map
    cache[key] = entry
    try:
        (_ROOT / "aria_kpi_cache.json").write_text(
            json.dumps(cache, indent=2), encoding="utf-8"
        )
    except Exception:
        pass


def _validate_kpis_against_df(kpis_raw: list, df: pd.DataFrame) -> list:
    """Drop any KPI whose required columns don't exist in df."""
    valid_cols = set(df.columns.tolist())
    out = []
    seen_names = set()
    for k in kpis_raw:
        if k.get("agg") == "ratio":
            # Ratio KPIs use num_col + den_col — column field may be "DERIVED" or absent
            if k.get("num_col") not in valid_cols or k.get("den_col") not in valid_cols:
                continue
            # Set column to num_col so downstream code has something to reference
            k["column"] = k["num_col"]
            k["kpi_type"] = "derived"
        else:
            col = k.get("column", "")
            if col == "DERIVED" or col not in valid_cols:
                continue
        # Deduplicate by name
        name = k.get("name", k.get("column", "KPI"))
        if name in seen_names:
            continue
        seen_names.add(name)
        k.setdefault("suggested_name", name)
        k.setdefault("user_name",      name)
        k.setdefault("enabled",        True)
        k.setdefault("kpi_type",       "direct")
        k.setdefault("scale",          1)
        # Sanitize: scale=100 + format="percent" is a double-multiply bug.
        # _fmt(val, "percent") uses Python's :.1% which already ×100.
        # Always store ratio KPIs as decimals (scale=1) for percent format.
        if k.get("scale") == 100 and k.get("format") == "percent":
            k["scale"] = 1
        out.append(k)
    return out


def _build_kpi_prompt(df: pd.DataFrame) -> str:
    col_types  = {c: str(df[c].dtype) for c in df.columns}
    sample     = df.head(3).to_dict(orient="records")
    # Compute basic stats to help Gemini understand numeric ranges
    num_cols   = df.select_dtypes(include="number").columns.tolist()
    col_stats  = {
        c: {"min": round(float(df[c].min()), 2),
            "max": round(float(df[c].max()), 2),
            "mean": round(float(df[c].mean()), 2)}
        for c in num_cols[:12]   # cap to keep prompt size reasonable
    }
    return f"""You are a Chief Analytics Officer with deep domain expertise across industries.

Your task: analyse this dataset and design a COMPREHENSIVE KPI framework for executive reporting.
Think like a board member — what metrics reveal risk, opportunity, efficiency, and growth?

DATASET COLUMNS AND TYPES:
{json.dumps(col_types, indent=2)}

NUMERIC COLUMN STATISTICS (min / max / mean):
{json.dumps(col_stats, indent=2)}

SAMPLE DATA (first 3 rows):
{json.dumps(sample, indent=2, default=str)}

STEP 1 — Identify the business domain precisely (e.g. "Airlines & Aviation", "Quick Service Restaurant", "Retail & E-Commerce", "Logistics & Supply Chain", "Healthcare", "Banking & Finance").

STEP 2 — Design 10 to 15 KPIs. Go BEYOND simple totals. Include:

  A. SIMPLE DIRECT KPIs — totals, averages, unique counts (2-3 max)
     e.g. Total Revenue, Total Passengers, Unique Routes

  B. RATIO / RATE KPIs — percentage or rate derived from two columns (4-6 KPIs)
     e.g. Profit Margin % = Profit / Revenue
          On-Time Rate % = On-Time Flights / Total Flights
          Load Factor % = Passengers / Capacity
          Avg Revenue per Passenger = Revenue / Passengers
          Return Rate % = Returns / Orders
          Conversion Rate % = Conversions / Visits

  C. EFFICIENCY / PRODUCTIVITY KPIs — how much output per unit of input (2-3 KPIs)
     e.g. Revenue per Flight, Revenue per Employee, Sales per Store
          Avg Delay Minutes = mean(Delay column)
          Fuel Efficiency = Revenue / Fuel Cost
          Cost per Passenger = Cost / Passengers

  D. QUALITY / RISK KPIs — anomaly, failure, or exception metrics (2-3 KPIs)
     e.g. Delayed Flights Count = sum of a boolean/flag delay column if it exists
          Cancellation Rate % = Cancelled / Total
          Complaint Rate, Defect Rate, Churn Rate

RULES:
1. ONLY use column names that EXACTLY EXIST in the dataset — never invent column names
2. For ratio KPIs: both num_col AND den_col must exist in the dataset
3. Do NOT create two KPIs that measure the same thing in different units or currencies
   (e.g. if Revenue and Revenue EUR have identical values, include only one)
4. Be domain-specific — name KPIs the way an industry expert would name them
5. Keep "description" under 10 words. Keep "formula" under 8 words.
6. For "agg": use "ratio" for any KPI derived from two columns (A/B), "mean" for averages, "sum" for totals, "nunique" for unique counts

Return ONLY a valid JSON object — no markdown, no prose, no explanation:
{{
  "domain": "snake_case_id",
  "domain_label": "Human Readable Label",
  "domain_emoji": "single emoji",
  "kpis": [
    {{
      "name": "KPI Display Name",
      "column": "exact_column_name or DERIVED if ratio",
      "agg": "sum|mean|nunique|ratio",
      "format": "currency|integer|percent|number",
      "formula": "e.g. SUM(Revenue) or Profit/Revenue*100",
      "description": "what this measures in ≤10 words",
      "kpi_type": "direct|derived",
      "num_col": "numerator column name (ratio KPIs only)",
      "den_col": "denominator column name (ratio KPIs only)",
      "den_agg": "sum|nunique (ratio KPIs only)",
      "scale": 1
    }}
  ]
}}

For ratio KPIs that are percentages, always set scale=1 and format="percent".
Python's :.1% format already multiplies by 100, so store the raw ratio as a decimal (0–1 range).
For ratio KPIs that are plain rates (e.g. revenue per unit), set scale=1 and format="currency" or "number".
"""


def _robust_json_parse(raw: str) -> dict:
    """
    Robustly extract and parse JSON from an LLM response that may contain
    markdown fences, trailing commas, single quotes, or extra prose.
    """
    # 1. Strip markdown fences
    if "```" in raw:
        parts = raw.split("```")
        for part in parts:
            candidate = part.lstrip("json").strip()
            if candidate.startswith("{"):
                raw = candidate
                break

    # 2. Find the outermost { ... } block
    start = raw.find("{")
    end   = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        raw = raw[start:end + 1]

    # 3. Try direct parse first
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass

    # 4. Fix trailing commas  e.g.  [1, 2, 3,]  or  {"a":1,}
    import re as _re
    fixed = _re.sub(r",\s*([}\]])", r"\1", raw)
    try:
        return json.loads(fixed)
    except json.JSONDecodeError:
        pass

    # 5. Replace Python literals
    fixed2 = fixed.replace("True", "true").replace("False", "false").replace("None", "null")
    return json.loads(fixed2)   # let this raise if still broken


def _pick_gemini_model(api_key: str) -> str | None:
    """
    Query the models list endpoint to find the best available flash model
    that supports generateContent. Returns model id string or None.
    """
    import requests as _rq
    # Free-tier models first — never pick a paid model when a free one is available
    PREFERRED = [
        "gemini-1.5-flash",        # free tier, 15 RPM
        "gemini-1.5-flash-8b",     # free tier, smaller/faster
        "gemini-1.5-flash-001",    # free tier pinned version
        "gemini-1.5-flash-002",    # free tier pinned version
        "gemini-2.0-flash-lite",   # free tier (new)
        "gemini-1.5-pro",          # free tier (limited)
        "gemini-1.5-pro-001",
        "gemini-2.0-flash",        # requires billing — last resort
        "gemini-pro",
    ]
    try:
        r = _rq.get(
            f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}",
            timeout=10,
        )
        if r.status_code != 200:
            return None
        available = {
            m["name"].split("/")[-1]
            for m in r.json().get("models", [])
            if "generateContent" in m.get("supportedGenerationMethods", [])
        }
        for model in PREFERRED:
            if model in available:
                return model
        # Last resort: return first available generateContent model
        if available:
            return next(iter(available))
    except Exception:
        pass
    return None


def _call_gemini_for_kpis(prompt: str) -> dict | None:
    """Call Gemini via REST API — auto-detects the best available model."""
    api_key = st.session_state.get("gemini_key") or st.session_state.get("gemini_key_input")
    if not api_key:
        st.session_state["_kpi_llm_error"] = "No Gemini API key found. Please go back to Step 3 and enter your key."
        return None
    try:
        import requests as _rq

        # Auto-pick best model available for this key
        model = _pick_gemini_model(api_key)
        if not model:
            st.session_state["_kpi_llm_error"] = "Could not list Gemini models — check your API key is valid."
            return None

        url = (
            f"https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model}:generateContent?key={api_key}"
        )
        # gemini-pro doesn't support responseMimeType — use it only as last resort
        gen_cfg: dict = {"temperature": 0.1, "maxOutputTokens": 4096}
        if model != "gemini-pro":
            gen_cfg["responseMimeType"] = "application/json"

        body = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": gen_cfg,
        }
        r = _rq.post(url, json=body, timeout=60)
        if r.status_code != 200:
            err = r.json().get("error", {}).get("message", r.text[:300])
            st.session_state["_kpi_llm_error"] = f"Gemini ({model}) error: {r.status_code} — {err}"
            return None
        text = r.json()["candidates"][0]["content"]["parts"][0]["text"]
        result = _robust_json_parse(text)
        st.session_state.pop("_kpi_llm_error", None)
        return result
    except Exception as e:
        st.session_state["_kpi_llm_error"] = f"Gemini error: {type(e).__name__}: {str(e)[:200]}"
        return None


def _call_openrouter_for_kpis(prompt: str) -> dict | None:
    """Fallback: call a free OpenRouter model for KPI detection."""
    try:
        import requests as _req
        headers = {
            "Authorization": "Bearer sk-or-v1-free",   # OpenRouter free tier
            "Content-Type":  "application/json",
            "HTTP-Referer":  "https://aria-wizard.streamlit.app",
            "X-Title":       "ARIA Wizard",
        }
        body = {
            "model":    "mistralai/mistral-7b-instruct:free",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens":  3000,
        }
        r = _req.post("https://openrouter.ai/api/v1/chat/completions",
                      json=body, headers=headers, timeout=30)
        if r.status_code != 200:
            return None
        raw = r.json()["choices"][0]["message"]["content"].strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.rsplit("```", 1)[0].strip()
        return json.loads(raw)
    except Exception:
        return None


def detect_kpis_llm(df: pd.DataFrame) -> list:
    """
    LLM-powered KPI discovery.
    1. Check aria_kpi_cache.json by (domain + column fingerprint) — 30-day TTL
    2. Call Gemini 2.5 Flash
    3. Fallback: call free OpenRouter model
    4. Validate all returned KPIs have columns that exist in df
    Returns list of KPI dicts compatible with the existing wizard UI.
    """
    from datetime import datetime as _dt, timedelta as _td

    domain, cache_key = _kpi_cache_key(df)

    # ── 1. Cache check ────────────────────────────────────────────────────── #
    cache = _read_kpi_cache()
    if cache_key in cache:
        entry = cache[cache_key]
        try:
            age = _dt.now() - _dt.fromisoformat(entry["generated_at"])
            if age < _td(days=30):
                # Restore domain metadata + role map to session state
                st.session_state.detected_domain       = entry["domain"]
                st.session_state.detected_domain_label = entry.get("domain_label", entry["domain"])
                st.session_state.detected_domain_emoji = entry.get("domain_emoji", "📊")
                if "role_kpi_map" in entry:
                    st.session_state.role_kpi_map = entry["role_kpi_map"]
                # Re-run dedup in case cache was saved before this fix was applied
                return _deduplicate_kpis(entry["kpis"], df)
        except Exception:
            pass

    # ── 2. Build prompt ───────────────────────────────────────────────────── #
    prompt = _build_kpi_prompt(df)

    # ── 3. Try Gemini ─────────────────────────────────────────────────────── #
    result = _call_gemini_for_kpis(prompt)

    # ── 4. Fallback to free LLM ───────────────────────────────────────────── #
    if result is None:
        result = _call_openrouter_for_kpis(prompt)

    # ── 5. If both LLMs failed, fall back to built-in pattern engine ─────────── #
    if result is None:
        st.session_state["_kpi_used_fallback"] = True
        fallback = detect_kpis(df)
        fallback = _deduplicate_kpis(fallback, df)   # remove duplicates even in fallback
        # Still set a basic domain from the regex detector
        domain_fb = detect_domain(df)
        d_emoji_fb, d_label_fb = _DOMAIN_META.get(domain_fb, ("📊", "Business Analytics"))
        st.session_state.detected_domain       = domain_fb
        st.session_state.detected_domain_label = d_label_fb
        st.session_state.detected_domain_emoji = d_emoji_fb
        return fallback

    # ── 6. Validate columns ───────────────────────────────────────────────── #
    kpis = _validate_kpis_against_df(result.get("kpis", []), df)

    # ── 6b. Deduplicate (name-similarity + value-identity) ───────────────── #
    kpis = _deduplicate_kpis(kpis, df)

    # ── 7. Store domain metadata ──────────────────────────────────────────── #
    d_label = result.get("domain_label", domain)
    d_emoji = result.get("domain_emoji", "📊")
    st.session_state.detected_domain       = result.get("domain", domain)
    st.session_state.detected_domain_label = d_label
    st.session_state.detected_domain_emoji = d_emoji

    # ── 8. Write cache ────────────────────────────────────────────────────── #
    _write_kpi_cache(cache_key, result.get("domain", domain), d_label, d_emoji, kpis)

    return kpis


# ════════════════════════════════════════════════════════════════════════════════
# KPI DEDUPLICATION  (runs right after detection, before user sees Step 3)
# ════════════════════════════════════════════════════════════════════════════════

def _deduplicate_kpis(kpis: list, df: "pd.DataFrame | None" = None) -> list:
    """
    Auto-disable KPIs that are semantically identical to another KPI in the list.

    Pass 1 — Name similarity:
      Normalise each name by stripping currency/unit suffixes (EUR, USD, %, min …).
      Within each group that shares the same normalised name, keep the shortest /
      bracket-free name; disable the rest with a '_duplicate_of' annotation.

    Pass 2 — Value identity (when df is available):
      For any two *enabled* direct KPIs whose source columns are identical, or whose
      column values have Pearson r > 0.999, disable the longer-named one.
    """
    import re
    from collections import defaultdict

    _UNIT_RE = re.compile(
        r'[\s\(\[]*('
        r'eur|usd|cad|gbp|inr|jpy|aud|chf|sgd|'
        r'%|pct|percent|rate|ratio|'
        r'min|mins|minutes|hrs|hours|days|'
        r'km|mi|miles|'
        r'k|m|bn|'
        r'yoy|mom|wow|dod|ytd|mtd|qtd'
        r')[\s\)\]]*$',
        re.IGNORECASE,
    )

    def _norm(name: str) -> str:
        n = name.strip().lower()
        n = _UNIT_RE.sub('', n).strip()
        n = re.sub(r'[^a-z0-9 ]', ' ', n)
        return re.sub(r'\s+', ' ', n).strip()

    result = [dict(k) for k in kpis]

    # ── Pass 1: name-based dedup ─────────────────────────────────────────── #
    groups: dict[str, list[int]] = defaultdict(list)
    for i, k in enumerate(result):
        groups[_norm(k['user_name'])].append(i)

    for norm, indices in groups.items():
        if len(indices) < 2:
            continue
        # Prefer: no bracket suffix first, then shortest name
        def _score(idx):
            name = result[idx]['user_name']
            return (bool(re.search(r'[\(\[]', name)), len(name))
        indices_sorted = sorted(indices, key=_score)
        keep_name = result[indices_sorted[0]]['user_name']
        for idx in indices_sorted[1:]:
            if result[idx].get('enabled', True):           # don't clobber already-disabled
                result[idx]['enabled']      = False
                result[idx]['_duplicate_of'] = keep_name

    # ── Pass 2: value-based dedup (only when df is available) ───────────── #
    if df is not None:
        active = [
            (i, k) for i, k in enumerate(result)
            if k.get('enabled', True)
            and k.get('kpi_type') == 'direct'
            and k.get('column') in df.columns
        ]
        for ai in range(len(active)):
            i, ki = active[ai]
            if not result[i].get('enabled', True):
                continue
            col_i = ki['column']
            try:
                s_i = pd.to_numeric(df[col_i], errors='coerce').dropna()
            except Exception:
                continue
            for aj in range(ai + 1, len(active)):
                j, kj = active[aj]
                if not result[j].get('enabled', True):
                    continue
                col_j = kj['column']
                # Same underlying column → always a duplicate
                if col_i == col_j:
                    result[j]['enabled']      = False
                    result[j]['_duplicate_of'] = ki['user_name']
                    continue
                try:
                    s_j = pd.to_numeric(df[col_j], errors='coerce').dropna()
                    common = s_i.index.intersection(s_j.index)
                    if len(common) < 10:
                        continue
                    corr = s_i.loc[common].corr(s_j.loc[common])
                    if corr is not None and abs(corr) > 0.999:
                        # Near-identical values — disable the longer-named one
                        n_i, n_j = ki['user_name'], kj['user_name']
                        if len(n_i) <= len(n_j):
                            result[j]['enabled']      = False
                            result[j]['_duplicate_of'] = n_i
                        else:
                            result[i]['enabled']      = False
                            result[i]['_duplicate_of'] = n_j
                except Exception:
                    continue

    return result


# ════════════════════════════════════════════════════════════════════════════════
# ROLE → KPI ASSIGNMENT  (Gemini second pass → tier-rule fallback)
# ════════════════════════════════════════════════════════════════════════════════

# Role metadata for the assignment prompt
_ROLE_TIERS = {
    "CEO":              ("c_suite",    "Strategic, board-level. Revenue growth, margin health, risk."),
    "CFO":              ("c_suite",    "Financial precision. Profit, margin, cost discipline."),
    "COO":              ("operations", "Operational efficiency. Order volume, fulfilment, throughput."),
    "CTO":              ("c_suite",    "Systems and data quality. Process reliability, anomalies."),
    "VP":               ("leadership", "Cross-functional. Growth and efficiency balance."),
    "Director":         ("leadership", "Execution focus. Weekly priorities and team actions."),
    "Sales Head":       ("commercial", "Revenue momentum. Deal activity, pipeline, regions."),
    "Operations Head":  ("operations", "Throughput, on-time delivery, shipping performance."),
    "Senior Manager":   ("management", "Team-level product and segment detail."),
    "Manager":          ("management", "Ground-level daily activity and task completion."),
    "Team Lead":        ("management", "Granular operational, today-focused, specific SKUs."),
    "Business Analyst": ("management", "Data-driven. Anomaly detection, trend breaks, root cause."),
}

# Tier → keyword preferences for the rule-based fallback
_TIER_KEYWORDS: dict[str, list[str]] = {
    "c_suite":    ["revenue", "profit", "margin", "income", "return", "ebit", "growth",
                   "sales", "load factor", "yield"],
    "leadership": ["revenue", "profit", "orders", "load", "volume", "passengers",
                   "bookings", "sales", "margin"],
    "commercial": ["revenue", "sales", "orders", "aov", "conversion", "bookings",
                   "passengers", "ticket", "yield"],
    "operations": ["time", "delay", "on-time", "on_time", "throughput", "quantity",
                   "utilization", "performance", "service", "fulfilment", "defect",
                   "downtime", "speed", "efficiency", "rpk", "ask"],
    "management": ["quantity", "orders", "transactions", "customers", "units",
                   "on-time", "delay", "service time", "passengers", "bookings"],
}


def _assign_kpis_to_roles_fallback(kpis: list) -> dict:
    """
    Rule-based tier assignment — no API call needed.
    Scores each KPI against tier keyword lists, picks top 4 per role.
    """
    kpi_names = [k["user_name"] for k in kpis if k.get("enabled", True)]
    if not kpi_names:
        return {}

    def _score(kpi_name: str, keywords: list[str]) -> int:
        n = kpi_name.lower()
        return sum(1 for kw in keywords if kw in n)

    result: dict = {}
    for role, (tier, _) in _ROLE_TIERS.items():
        keywords = _TIER_KEYWORDS.get(tier, [])
        scored   = sorted(kpi_names,
                          key=lambda n: _score(n, keywords),
                          reverse=True)
        # Take top scored KPIs — no fixed cap, min 3, max all available
        assigned = scored[:max(3, len(kpi_names))]
        result[role] = assigned
    return result


def _assign_kpis_to_roles_gemini(kpis: list, domain_label: str) -> dict | None:
    """
    Second Gemini call: assign 4 KPIs per role.
    Returns dict {role_name: [kpi_name, ...]} or None on failure.
    """
    api_key = st.session_state.get("gemini_key") or st.session_state.get("gemini_key_input")
    if not api_key:
        return None

    enabled_kpis = [k for k in kpis if k.get("enabled", True)]
    kpi_summary  = "\n".join(
        f'- "{k["user_name"]}": {k.get("description", "")}' for k in enabled_kpis
    )
    role_lines   = "\n".join(
        f'- {role}: ({tier}) {desc}' for role, (tier, desc) in _ROLE_TIERS.items()
    )

    prompt = f"""You are an expert in data analytics and executive reporting for the {domain_label} industry.

The following KPIs have been detected in the dataset:
{kpi_summary}

Assign the most relevant KPIs to each business role below. Consider each role's tier, priorities and what they would track daily.

Roles:
{role_lines}

Rules:
1. Use ONLY KPI names from the list above — exact spelling
2. Assign between 3 and 6 KPIs per role based on genuine relevance — do NOT pad with irrelevant KPIs just to reach a number
3. First KPI in each list = the primary / hero KPI for that role
4. Different roles should have different primary KPIs where possible
5. AVOID DUPLICATE METRICS: if two KPIs measure the same thing in different currencies or units (e.g. "Revenue" and "Revenue (EUR)", "Sales" and "Net Sales", "Profit" and "Profit (EUR)"), assign ONLY ONE of them per role — pick the most business-relevant one. Never assign both to the same role.
6. Prefer KPIs that give genuinely different business insight for each role

Return ONLY valid JSON — no markdown, no explanation:
{{
  "CEO": ["kpi1", "kpi2", "kpi3"],
  "CFO": ["kpi1", "kpi2", "kpi3", "kpi4"],
  "COO": ["kpi1", "kpi2", "kpi3"],
  "CTO": ["kpi1", "kpi2", "kpi3"],
  "VP": ["kpi1", "kpi2", "kpi3", "kpi4"],
  "Director": ["kpi1", "kpi2", "kpi3"],
  "Sales Head": ["kpi1", "kpi2", "kpi3", "kpi4"],
  "Operations Head": ["kpi1", "kpi2", "kpi3", "kpi4"],
  "Senior Manager": ["kpi1", "kpi2", "kpi3"],
  "Manager": ["kpi1", "kpi2", "kpi3"],
  "Team Lead": ["kpi1", "kpi2", "kpi3"],
  "Business Analyst": ["kpi1", "kpi2", "kpi3", "kpi4", "kpi5"]
}}"""

    try:
        import requests as _rq
        model = _pick_gemini_model(api_key)
        if not model:
            return None
        url = (
            f"https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model}:generateContent?key={api_key}"
        )
        gen_cfg: dict = {"temperature": 0.1, "maxOutputTokens": 2048}
        if model != "gemini-pro":
            gen_cfg["responseMimeType"] = "application/json"
        body = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": gen_cfg,
        }
        r = _rq.post(url, json=body, timeout=60)
        if r.status_code != 200:
            return None
        text = r.json()["candidates"][0]["content"]["parts"][0]["text"]
        mapping = _robust_json_parse(text)
        # Validate: only keep KPI names that exist in our detected list
        valid = {k["user_name"] for k in enabled_kpis}
        for role in mapping:
            mapping[role] = [n for n in mapping[role] if n in valid]
        return mapping
    except Exception:
        return None


def assign_kpis_to_roles(kpis: list, domain_label: str, cache_key: str) -> dict:
    """
    Orchestrator:
    1. Return cached role_kpi_map if present in aria_kpi_cache.json
    2. Try Gemini assignment
    3. Fall back to rule-based tier assignment
    4. Write result back to cache (same entry as KPI list)
    """
    # ── 1. Check cache ────────────────────────────────────────────────────── #
    cache = _read_kpi_cache()
    if cache_key in cache and "role_kpi_map" in cache[cache_key]:
        return cache[cache_key]["role_kpi_map"]

    # ── 2. Try Gemini ─────────────────────────────────────────────────────── #
    mapping = _assign_kpis_to_roles_gemini(kpis, domain_label)

    # ── 3. Fallback ───────────────────────────────────────────────────────── #
    if not mapping:
        mapping = _assign_kpis_to_roles_fallback(kpis)

    # ── 4. Write to cache (alongside existing KPI entry) ─────────────────── #
    if cache_key in cache:
        entry = cache[cache_key]
        _write_kpi_cache(
            cache_key,
            entry.get("domain", ""),
            entry.get("domain_label", ""),
            entry.get("domain_emoji", "📊"),
            entry.get("kpis", kpis),
            role_kpi_map=mapping,
        )

    return mapping


def detect_date_col(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        if "datetime" in str(df[col].dtype):
            return col
    for col in df.columns:
        if any(k in col.lower() for k in ["date", "day", "time", "period", "month"]):
            try:
                pd.to_datetime(df[col].dropna().iloc[:5])
                return col
            except Exception:
                pass
    return None


def _fmt_kv(v: float, fmt: str) -> str:
    if fmt == "currency":  return f"${v:,.0f}"
    if fmt == "integer":   return f"{int(v):,}"
    if fmt == "percent":   return f"{v:.1f}%"
    if fmt == "decimal":   return f"{v:,.2f}"
    if fmt == "minutes":   return f"{v:.1f} min"
    if fmt == "score":     return f"{v:.1f}"
    return f"{v:,.2f}"


def _pct(a: float, b: float):
    return None if not b else (a - b) / abs(b)


def compute_preview_metrics_v2(
    df: pd.DataFrame,
    kpis_cfg: list,
    date_col: str,
    timeframe_key: str = "1d",
) -> dict:
    """
    Replaces compute_preview_metrics() — uses the EXACT same agent engines
    (metrics_engine + driver_analysis) so the wizard preview is guaranteed
    to match the GitHub Actions generated card.

    Returns a dict with the same keys as the old function so all callers
    (build_live_narrative, generate_svg_preview) work unchanged.
    """
    try:
        from agent.metrics_engine import compute_metrics
        from agent.driver_analysis import analyze_drivers, drivers_to_dict
    except ImportError:
        try:
            from metrics_engine import compute_metrics
            from driver_analysis import analyze_drivers, drivers_to_dict
        except ImportError:
            # Hard fallback to old implementation if agent modules unavailable
            return compute_preview_metrics(df, kpis_cfg, date_col, timeframe_key)

    import datetime as _dt

    # ── Build a config dict identical to what _build_configs() produces ───── #
    # Detect dimension columns dynamically from the dataframe
    kpi_cols  = {k.get("column") for k in kpis_cfg} | \
                {k.get("num_col") for k in kpis_cfg if k.get("num_col")} | \
                {k.get("den_col") for k in kpis_cfg if k.get("den_col")}
    date_like = {"date", "time", "year", "month", "quarter", "week", "day"}
    dim_cols  = [
        c for c in df.columns
        if c not in kpi_cols and c != date_col
        and df[c].dtype == object
        and not any(d in c.lower() for d in date_like)
        and df[c].nunique() < 50
    ][:6] or ["Category", "Region", "Segment"]

    config = {
        "data": {
            "date_column":                   date_col,
            "timezone":                      "America/Toronto",
            "fallback_to_max_date_if_missing": True,
        },
        "metrics": {
            "timeframe": timeframe_key,
            "kpis": [
                {
                    "name":   k["user_name"],
                    "column": k.get("column", k["user_name"]),
                    "agg":    k.get("agg", "sum"),
                    "format": k.get("format", "number"),
                    **({
                        "num_col":  k["num_col"],
                        "den_col":  k["den_col"],
                        "den_agg":  k.get("den_agg", "sum"),
                        "scale":    k.get("scale", 1),
                    } if k.get("agg") == "ratio" else {}),
                }
                for k in kpis_cfg if k.get("enabled", True)
            ],
            "anomaly_zscore_threshold": 2.0,
            "anomaly_lookback_days":    90,
        },
        "drivers": {
            "dimensions": dim_cols,
            "top_n": 3,
        },
        "delivery": {"channels": [], "file": {"output_dir": "output"}},
    }

    # ── Run the actual agent engines ─────────────────────────────────────── #
    try:
        snapshot = compute_metrics(df, config)
    except Exception:
        return compute_preview_metrics(df, kpis_cfg, date_col, timeframe_key)

    ref = _dt.date.fromisoformat(snapshot.reference_date)

    try:
        drivers_yoy = analyze_drivers(df, config, ref, compare_to="yoy")
        drivers_raw = drivers_to_dict(drivers_yoy)
    except Exception:
        drivers_raw = {}

    # ── Sparkline — 30-day daily series for the primary KPI ──────────────── #
    trend_series: list = []
    try:
        _first_kpi  = next((k for k in kpis_cfg if k.get("enabled", True)), None)
        _spark_col  = (_first_kpi.get("column") or _first_kpi.get("num_col", "")) \
                      if _first_kpi else ""
        if _spark_col and _spark_col in df.columns:
            _df_spark = df.copy()
            _df_spark["_date"] = pd.to_datetime(_df_spark[date_col]).dt.date
            _thirty_ago = ref - timedelta(days=29)
            _df_spark = _df_spark[
                (_df_spark["_date"] >= _thirty_ago) &
                (_df_spark["_date"] <= ref)
            ]
            trend_series = (
                _df_spark.groupby("_date")[_spark_col]
                .sum().sort_index().tolist()
            )
    except Exception:
        trend_series = []

    # ── Flatten drivers to a list for backward-compat with callers ────────── #
    # Callers (build_live_narrative, generate_svg_preview) expect either:
    #   - flat list [{dimension, member, delta, ...}]  (old format)
    #   - or dict {kpi_name: [...]}                    (agent format)
    # We return the dict format — callers updated to handle both.
    prim_kpi = (
        snapshot.kpis[next(iter(snapshot.kpis), "")].name
        if snapshot.kpis else "KPI"
    )

    payload = snapshot.to_dict()   # already has reference_date, window_start,
                                   # timeframe, kpis (with value_fmt, mom_pct etc.)
    payload["drivers"]      = drivers_raw   # dict {kpi_name: [DriverItem dicts]}
    payload["trend_series"] = trend_series
    payload["timeframe_key"] = timeframe_key

    # ── Target / Achievement KPIs (optional) ──────────────────────────────── #
    # If the user uploaded a target/plan file in Step 2, compute:
    #   - Achievement % = Actual / Target × 100  (per KPI)
    #   - Gap to Target  = Actual − Target        (per KPI)
    # and inject them as additional KPIs into the payload.
    df_target = st.session_state.get("df_target")
    if df_target is not None:
        try:
            _tf = df_target.copy()
            _tf["_date"] = pd.to_datetime(_tf[date_col], errors="coerce").dt.date
            _ref  = snapshot.reference_date  # string
            import datetime as _dtt
            _ref_date  = _dtt.date.fromisoformat(_ref)
            _win_start = _dtt.date.fromisoformat(snapshot.window_start) \
                         if snapshot.window_start else _ref_date
            _tgt_win = _tf[(_tf["_date"] >= _win_start) & (_tf["_date"] <= _ref_date)]

            _achievement_kpis = {}
            for _k in kpis_cfg:
                if not _k.get("enabled", True):
                    continue
                _name   = _k["user_name"]
                _col    = _k.get("column", "")
                _agg    = _k.get("agg", "sum")
                _fmt    = _k.get("format", "number")
                if _agg == "ratio" or _col == "DERIVED" or _col not in _tgt_win.columns:
                    continue
                # Compute target value using same aggregation
                if _agg == "sum":
                    _tval = float(_tgt_win[_col].sum())
                elif _agg == "mean":
                    _tval = float(_tgt_win[_col].mean()) if len(_tgt_win) else 0.0
                elif _agg == "nunique":
                    _tval = float(_tgt_win[_col].nunique())
                else:
                    continue
                if _tval == 0:
                    continue
                # Actual value from snapshot
                _aval = payload["kpis"].get(_name, {}).get("value", 0.0)
                if not isinstance(_aval, (int, float)):
                    continue
                _achiev = (_aval / _tval) * 100
                _gap    = _aval - _tval

                def _fmt_v(v, fmt):
                    if fmt == "currency":
                        av = abs(v)
                        if av >= 1_000_000: return f"${v/1_000_000:.1f}M"
                        if av >= 1_000:     return f"${v/1_000:.1f}K"
                        return f"${v:,.0f}"
                    if fmt == "percent": return f"{v:.1f}%"
                    if fmt == "integer": return f"{int(v):,}"
                    return f"{v:,.1f}"

                _achievement_kpis[f"{_name} Achievement %"] = {
                    "name":      f"{_name} Achievement %",
                    "value":     round(_achiev, 2),
                    "value_fmt": f"{_achiev:.1f}%",
                    "mom_pct":   None, "yoy_pct": None,
                    "dod_pct":   None, "wow_pct": None,
                    "direction": "up" if _achiev >= 100 else "down",
                    "format":    "percent",
                }
                _achievement_kpis[f"{_name} Gap to Target"] = {
                    "name":      f"{_name} Gap to Target",
                    "value":     round(_gap, 2),
                    "value_fmt": _fmt_v(_gap, _fmt),
                    "mom_pct":   None, "yoy_pct": None,
                    "dod_pct":   None, "wow_pct": None,
                    "direction": "up" if _gap >= 0 else "down",
                    "format":    _fmt,
                }
            payload["kpis"].update(_achievement_kpis)
            payload["has_target"] = True
        except Exception:
            pass   # target computation failed silently — card still works

    return payload


def compute_preview_metrics(
    df: pd.DataFrame,
    kpis_cfg: list,
    date_col: str,
    timeframe_key: str = "30d",
) -> dict:
    """DEPRECATED — use compute_preview_metrics_v2() instead.
    Kept as fallback only."""
    """
    Compute KPI metrics for the selected timeframe window.

    Current window  = last <window_days> of data ending at ref date.
    Previous window = the equal-length period immediately before that.
    Change %        = (current - previous) / |previous|.

    Backward-compatible: populates mom_pct / yoy_pct on the KPI dict so the
    SVG card renderer (which reads those fields) continues to work unchanged.
    """
    df = df.copy()
    try:
        df["_date"] = pd.to_datetime(df[date_col]).dt.date
    except Exception:
        return {}

    ref  = df["_date"].max()
    dmin = df["_date"].min()
    tf   = _tf_by_key(timeframe_key)

    # ── Determine window boundaries ─────────────────────────────────────── #
    if timeframe_key == "ytd":
        start_curr  = date(ref.year, 1, 1)
        window_days = max((ref - start_curr).days + 1, 1)
    elif timeframe_key == "alltime":
        start_curr  = dmin
        window_days = max((ref - dmin).days + 1, 1)
    elif timeframe_key == "wtd":
        # Monday of the week containing ref
        start_curr  = ref - timedelta(days=ref.weekday())
        window_days = max((ref - start_curr).days + 1, 1)
    elif timeframe_key == "mtd":
        start_curr  = date(ref.year, ref.month, 1)
        window_days = max((ref - start_curr).days + 1, 1)
    elif timeframe_key == "qtd":
        _q_start_month = ((ref.month - 1) // 3) * 3 + 1
        start_curr  = date(ref.year, _q_start_month, 1)
        window_days = max((ref - start_curr).days + 1, 1)
    else:
        window_days = tf["days"] or 30   # safe fallback
        start_curr  = ref - timedelta(window_days - 1)

    start_prev = start_curr - timedelta(window_days)
    end_prev   = start_curr - timedelta(1)

    curr_df = df[(df["_date"] >= start_curr) & (df["_date"] <= ref)]
    prev_df = df[(df["_date"] >= start_prev) & (df["_date"] <= end_prev)]

    # ── KPI aggregator ───────────────────────────────────────────────────── #
    def agg_val(sub, k):
        agg  = k.get("agg", "sum")
        ktyp = k.get("kpi_type", "direct")
        if ktyp == "derived" or agg == "ratio":
            num_col = k.get("num_col", "")
            den_col = k.get("den_col", "")
            den_agg = k.get("den_agg", "sum")
            if not num_col or not den_col or sub.empty:
                return 0.0
            if num_col not in sub.columns or den_col not in sub.columns:
                return 0.0
            num = float(sub[num_col].sum())
            den = (float(sub[den_col].nunique())
                   if den_agg == "nunique"
                   else float(sub[den_col].sum()))
            if not den:
                return 0.0
            return (num / den) * k.get("scale", 1)
        col = k.get("column", "")
        if col == "DERIVED" or col not in sub.columns or sub.empty:
            return 0.0
        s = sub[col]
        if agg == "sum":     return float(s.sum())
        if agg == "nunique": return float(s.nunique())
        if agg == "mean":    return float(s.mean()) if len(s) else 0.0
        if agg == "median":  return float(s.median()) if len(s) else 0.0
        if agg == "max":     return float(s.max()) if len(s) else 0.0
        if agg == "min":     return float(s.min()) if len(s) else 0.0
        return 0.0

    # ── KPI values ───────────────────────────────────────────────────────── #
    comparison_label = tf["comparison"]
    kpis: dict = {}
    for k in kpis_cfg:
        if not k.get("enabled", True):
            continue
        name   = k["user_name"]
        curr   = agg_val(curr_df, k)
        prev   = agg_val(prev_df, k)
        chg    = _pct(curr, prev)
        # Populate the field the SVG card reads based on timeframe key
        _is_short = timeframe_key in ("1d", "wtd", "mtd") or window_days <= 90
        _is_long  = timeframe_key in ("ytd", "alltime", "qtd") or window_days >= 90
        mom_v  = chg if _is_short else None
        yoy_v  = chg if _is_long  else None
        kpis[name] = {
            "value":           curr,
            "value_fmt":       _fmt_kv(curr, k["format"]),
            "change_pct":      chg,
            "change_label":    comparison_label,
            "timeframe_label": tf["label"],
            "mom_pct":         mom_v,
            "yoy_pct":         yoy_v,
            "dod_pct":         chg if timeframe_key == "1d"  else None,
            "wow_pct":         chg if timeframe_key == "wtd" else None,
        }

    # ── Sparkline — granularity matches timeframe ─────────────────────────── #
    primary_cfg = next((k for k in kpis_cfg if k.get("enabled")), None)
    trend_series: list[float] = []
    if primary_cfg:
        gran   = tf["sparkline_gran"]
        npts   = tf["sparkpoints"]

        if gran == "daily":
            # one point per day in the current window (capped at npts)
            days_in_window = (ref - start_curr).days + 1
            step = max(1, days_in_window // npts)
            for i in range(days_in_window - 1, -1, -step):
                d_df = df[df["_date"] == ref - timedelta(i)]
                trend_series.append(agg_val(d_df, primary_cfg))
                if len(trend_series) >= npts:
                    break
            trend_series.reverse()

        elif gran == "weekly":
            # one point per 7-day bucket, last npts weeks
            for w in range(npts - 1, -1, -1):
                w_start = start_curr + timedelta(weeks=w * window_days // (npts * 7))
                w_end   = w_start + timedelta(6)
                w_df    = df[(df["_date"] >= w_start) & (df["_date"] <= min(w_end, ref))]
                trend_series.append(agg_val(w_df, primary_cfg))
            # simpler: iterate through 7-day chunks within current window
            trend_series = []
            chunk = max(1, window_days // npts)
            for i in range(npts):
                d0 = start_curr + timedelta(i * chunk)
                d1 = d0 + timedelta(chunk - 1)
                w_df = df[(df["_date"] >= d0) & (df["_date"] <= min(d1, ref))]
                trend_series.append(agg_val(w_df, primary_cfg))

        else:  # monthly
            # one point per calendar month bucket
            months: list[tuple] = []
            cur = start_curr.replace(day=1)
            while cur <= ref:
                import calendar
                last_day = calendar.monthrange(cur.year, cur.month)[1]
                m_end = cur.replace(day=last_day)
                months.append((cur, min(m_end, ref)))
                if cur.month == 12:
                    cur = cur.replace(year=cur.year + 1, month=1)
                else:
                    cur = cur.replace(month=cur.month + 1)
            # keep last npts months
            for (m0, m1) in months[-npts:]:
                m_df = df[(df["_date"] >= m0) & (df["_date"] <= m1)]
                trend_series.append(agg_val(m_df, primary_cfg))

    # ── Driver analysis ──────────────────────────────────────────────────── #
    drivers: list[dict] = []
    _dim_priority = ["Category", "Sub-Category", "Region", "Segment", "Ship Mode",
                     "Department", "Brand", "Store", "Channel", "Product_Type",
                     "Location", "Country", "City", "State", "Zone", "Division",
                     "Team", "Owner", "Source", "Medium", "Campaign", "Platform"]
    _cat_cols = [c for c in df.select_dtypes(include=["object", "category"]).columns
                 if 2 <= df[c].nunique() <= 30]
    dim_col = (next((c for c in _dim_priority if c in df.columns), None)
               or (next(iter(_cat_cols), None)))

    if dim_col and primary_cfg:
        pcol  = primary_cfg.get("column", "")
        pagg  = primary_cfg.get("agg", "sum")
        if primary_cfg.get("kpi_type") == "derived":
            pcol = primary_cfg.get("num_col", pcol)
            pagg = "sum"

        def grp(sub):
            if sub.empty or pcol not in sub.columns:
                return pd.Series(dtype=float)
            if pagg == "sum":     return sub.groupby(dim_col)[pcol].sum()
            if pagg == "nunique": return sub.groupby(dim_col)[pcol].nunique()
            if pagg == "mean":    return sub.groupby(dim_col)[pcol].mean()
            return sub.groupby(dim_col)[pcol].sum()

        curr_g = grp(curr_df)
        prev_g = grp(prev_df)
        for member in curr_g.index:
            cv    = float(curr_g[member])
            pv    = float(prev_g.get(member, 0))
            delta = cv - pv
            if delta != 0:
                drivers.append({
                    "dimension": dim_col, "member": str(member),
                    "delta": delta, "delta_pct": (delta / pv) if pv else None,
                })
        drivers.sort(key=lambda d: abs(d["delta"]), reverse=True)

    return {
        "reference_date":    str(ref),
        "window_start":      str(start_curr),
        "timeframe_key":     timeframe_key,
        "timeframe_label":   tf["label"],
        "comparison_label":  comparison_label,
        "kpis":              kpis,
        "trend_series":      trend_series,
        "drivers":           drivers,
    }


def build_stub_narrative(metrics: dict, role_cfg: dict) -> tuple[dict, object]:
    try:
        from narrative_generator import _generate_stub
        prim    = role_cfg.get("primary_kpi", "Sales")
        drivers = metrics.get("drivers", [])
        payload = {
            "reference_date": metrics.get("reference_date", str(date.today())),
            "kpis":           metrics.get("kpis", {}),
            "anomalies":      [],
            "drivers":        {prim: drivers, "Sales": drivers},
        }
        result = _generate_stub(payload, {}, role_cfg)
        result.reference_date = metrics.get("reference_date", str(date.today()))
        return result.to_dict() | {"reference_date": result.reference_date}, result
    except Exception:
        kpis = metrics.get("kpis", {})
        prim = role_cfg.get("primary_kpi", "Sales")
        pval = kpis.get(prim, {}).get("value_fmt", "—")
        ref  = metrics.get("reference_date", str(date.today()))
        rows = "\n".join(
            f"| {n} | {d.get('value_fmt','—')} | — | — | — | — |"
            for n, d in list(kpis.items())[:4]
        )
        d = {
            "reference_date":      ref,
            "headline":            f"{pval} on {ref} — {prim} performance briefing",
            "exec_summary":        f"{prim} closed at {pval} on {ref}. Broad-based performance.",
            "kpi_table_md":        "| Metric | Value | DoD | WoW | MoM | YoY |\n|---|---|---|---|---|---|\n" + rows,
            "anomaly":             "No anomalies on the 90-day window.",
            "recommended_action":  f"Review {prim} by dimension. Owner: Analytics. By end of week.",
            "speaker_notes":       f"Prepared for {role_cfg.get('title','Leadership')}.",
            "drivers_md":          "- Drivers unavailable for this preview.",
            "model":               "fallback",
            "role":                role_cfg.get("title", "Leadership"),
        }
        obj = SimpleNamespace(**d)
        return d, obj


def build_live_narrative(metrics: dict, role_cfg: dict) -> tuple[dict, object]:
    """Call real Gemini to generate the Step-6 preview narrative — exact
    parity with what agent/main.py produces for the Slack-delivered card.

    Fallback chain (silent, never raises):
      1. No GEMINI_API_KEY in session_state or env → stub
      2. agent.narrative_generator not importable     → stub
      3. Gemini call raises (rate limit, network)     → stub (built into
         generate_narrative itself — defensive layer kept just in case)

    Returns (narr_dict, narr_obj) — identical shape to build_stub_narrative
    so the caller in step_preview_card doesn't need to change.
    """
    import os as _os

    api_key = (st.session_state.get("gemini_key", "")
               or _os.environ.get("GEMINI_API_KEY", "")).strip()
    if not api_key:
        # No key configured — graceful degrade to stub (banner is shown by caller)
        return build_stub_narrative(metrics, role_cfg)

    # Try to import the real generator; fall back to stub if not on sys.path
    try:
        from narrative_generator import generate_narrative
    except ImportError:
        try:
            from agent.narrative_generator import generate_narrative
        except ImportError:
            return build_stub_narrative(metrics, role_cfg)

    prim    = role_cfg.get("primary_kpi", "Sales")
    _raw_drivers = metrics.get("drivers", {})
    # Support both formats:
    #   new: dict {kpi_name: [DriverItem dicts]}  — from compute_preview_metrics_v2
    #   old: flat list [{dimension, member, delta}] — from compute_preview_metrics
    if isinstance(_raw_drivers, dict):
        drivers_dict = _raw_drivers
        # Ensure primary KPI key exists
        if prim not in drivers_dict and drivers_dict:
            drivers_dict[prim] = next(iter(drivers_dict.values()), [])
    else:
        drivers_dict = {prim: _raw_drivers}
    payload = {
        "reference_date":  metrics.get("reference_date", str(date.today())),
        "window_start":    metrics.get("window_start", ""),
        "timeframe":       metrics.get("timeframe", metrics.get("timeframe_key", "1d")),
        "kpis":            metrics.get("kpis", {}),
        "anomalies":       metrics.get("anomalies", []),
        "drivers":         drivers_dict,
        "daily_sales_30d": metrics.get("trend_series", []),
    }
    cfg = {
        "llm": {
            "provider":          "gemini",
            "model":             "gemini-2.5-flash",
            "temperature":       0.4,
            "max_output_tokens": 8192,
        }
    }

    # generate_narrative reads GEMINI_API_KEY from the environment.
    # Set it for this call, restore the prior value (or unset) when done so
    # we don't pollute the Streamlit process env.
    _prev_key = _os.environ.get("GEMINI_API_KEY")
    _os.environ["GEMINI_API_KEY"] = api_key
    try:
        result = generate_narrative(payload, cfg, role_cfg)
        result.reference_date = metrics.get("reference_date", str(date.today()))
        try:
            d = result.to_dict()
        except Exception:
            # Defensive — if result is some other shape, gather public attrs
            d = {k: v for k, v in vars(result).items() if not k.startswith("_")}
        d["reference_date"] = result.reference_date
        return d, result
    except Exception:
        # Any unexpected failure — stub keeps the wizard alive
        return build_stub_narrative(metrics, role_cfg)
    finally:
        if _prev_key is None:
            _os.environ.pop("GEMINI_API_KEY", None)
        else:
            _os.environ["GEMINI_API_KEY"] = _prev_key


# ════════════════════════════════════════════════════════════════════════════════
# HTML CARD PREVIEW  (replaces SVG preview)
# ════════════════════════════════════════════════════════════════════════════════

def generate_svg_preview(narrative_obj, metrics: dict, role_cfg: dict,
                         style_key: str, template_key: str = "editorial") -> str | None:
    """
    Now returns a self-contained HTML string rendered via st.components.v1.html().
    Name kept as generate_svg_preview() so all call sites work unchanged.
    """
    try:
        from agent.html_generator import generate_html_card
    except ImportError:
        try:
            from html_generator import generate_html_card
        except ImportError:
            return None

    all_kpis = metrics.get("kpis", {})

    # Filter to role-assigned KPIs only
    role_kpi_names    = role_cfg.get("kpis", [])
    role_kpis_filtered = {k: v for k, v in all_kpis.items() if k in role_kpi_names}
    if not role_kpis_filtered:
        role_kpis_filtered = dict(list(all_kpis.items())[:6])

    prim_kpi = role_cfg.get("primary_kpi") or next(iter(role_kpis_filtered), "KPI")
    if prim_kpi not in role_kpis_filtered and role_kpis_filtered:
        prim_kpi = next(iter(role_kpis_filtered))

    # Drivers — support dict and flat-list formats
    _raw_drivers = metrics.get("drivers", {})
    if isinstance(_raw_drivers, dict):
        drivers_dict = _raw_drivers
        if prim_kpi not in drivers_dict and drivers_dict:
            drivers_dict[prim_kpi] = next(iter(drivers_dict.values()), [])
    else:
        drivers_dict = {prim_kpi: _raw_drivers}

    payload = {
        "reference_date":  metrics.get("reference_date", str(date.today())),
        "window_start":    metrics.get("window_start", ""),
        "timeframe":       metrics.get("timeframe", metrics.get("timeframe_key", "1d")),
        "kpis":            role_kpis_filtered,
        "drivers":         drivers_dict,
        "daily_sales_30d": metrics.get("trend_series", []),
    }

    mod_role = dict(role_cfg)
    mod_role["card_template"] = template_key
    mod_role["card_style"]    = style_key
    mod_role["primary_kpi"]   = prim_kpi
    mod_role["kpis"]          = list(role_kpis_filtered.keys())

    try:
        html = generate_html_card(narrative_obj, payload, {}, mod_role)
    except Exception:
        return None

    # Inject fluid CSS so the card fills the Streamlit iframe width instead of
    # overflowing at 900px.  In the PNG path Chrome's viewport IS 900px, so
    # width:100% still renders at exactly 900px — no change to Slack output.
    # Cards are designed at 900px. Force body to that width so all templates
    # (including editorial's fixed flex columns) render at their natural size.
    fluid_css = """<style>
html { overflow-x: hidden; }
body { margin: 0; padding: 0; width: 900px; transform-origin: top left; }
</style>"""
    html = html.replace('</head>', fluid_css + '\n</head>')

    # Scale-to-fit + auto-height script:
    # • Body is set to 900px (the card's design width) so all internal layouts
    #   render correctly — editorial 3-column, scorecard grid, etc.
    # • If the iframe viewport is narrower than 900px, we scale the whole body
    #   down proportionally. No cropping, no reflow for any template.
    # • The scaled height is reported to Streamlit via postMessage so the
    #   iframe snaps to exactly the right height with no empty space.
    fit_and_resize_script = """
<script>
(function fitAndResize() {
    var DESIGN_W = 900;

    function apply() {
        var vw = window.innerWidth || document.documentElement.clientWidth;
        // Only scale down; never enlarge
        var scale = (vw > 0 && vw < DESIGN_W) ? vw / DESIGN_W : 1;

        if (scale < 1) {
            document.body.style.transform = 'scale(' + scale + ')';
        } else {
            document.body.style.transform = '';
        }

        // Measure the body's natural (pre-scale) height, then scale it
        var naturalH = Math.max(
            document.body.scrollHeight,
            document.body.offsetHeight
        );
        var scaledH = Math.ceil(naturalH * scale);

        window.parent.postMessage(
            { type: 'streamlit:setFrameHeight', height: scaledH + 16 },
            '*'
        );
    }

    // Fire after Chart.js finishes (two passes for safety)
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', function() {
            setTimeout(apply, 400);
            setTimeout(apply, 1200);
        });
    } else {
        setTimeout(apply, 400);
        setTimeout(apply, 1200);
    }
    window.addEventListener('resize', apply);
})();
</script>"""
    html = html.replace('</body>', fit_and_resize_script + '\n</body>')
    return html


# ════════════════════════════════════════════════════════════════════════════════
# GITHUB AUTOMATION HELPERS
# ════════════════════════════════════════════════════════════════════════════════

def _gh_api(method: str, path: str, token: str, data=None):
    headers = {
        "Authorization": f"token {token}",
        "Accept":        "application/vnd.github.v3+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    url = f"https://api.github.com{path}"
    return getattr(_req, method)(url, headers=headers, json=data, timeout=12)


def _gh_encrypt_secret(public_key_str: str, secret_value: str):
    """Encrypt a secret for GitHub Actions using libsodium crypto_box_seal.
    Tries PyNaCl first; falls back to ctypes/libsodium directly for Python 3.14+
    environments where PyNaCl's cffi bindings may not yet be compatible.
    """
    pk_raw  = base64.b64decode(public_key_str)
    # Guard: convert AttrDict / any non-str to plain string
    if not isinstance(secret_value, str):
        try:
            import json as _j
            secret_value = _j.dumps({k: str(v) for k, v in secret_value.items()}) \
                           if hasattr(secret_value, "keys") else str(secret_value)
        except Exception:
            secret_value = str(secret_value)
    msg     = secret_value.encode("utf-8")

    # ── Attempt 1: PyNaCl (clean path) ────────────────────────────────────────
    try:
        from nacl.public import PublicKey, SealedBox
        pk  = PublicKey(pk_raw)
        box = SealedBox(pk)
        enc = box.encrypt(msg)
        return base64.b64encode(bytes(enc)).decode("utf-8")
    except Exception:
        pass  # fall through to ctypes path

    # ── Attempt 2: libsodium via ctypes (works when PyNaCl cffi breaks on 3.14) ─
    try:
        import ctypes, ctypes.util
        _lib_name = ctypes.util.find_library("sodium")
        # common paths on Debian/Ubuntu (Streamlit Cloud)
        for candidate in [_lib_name, "libsodium.so", "libsodium.so.23", "libsodium.so.26"]:
            if not candidate:
                continue
            try:
                _lib = ctypes.CDLL(candidate)
                break
            except OSError:
                continue
        else:
            raise RuntimeError("libsodium shared library not found")

        SEAL_BYTES = 48  # crypto_box_SEALBYTES
        cipher_buf = ctypes.create_string_buffer(len(msg) + SEAL_BYTES)
        ret = _lib.crypto_box_seal(
            cipher_buf,
            ctypes.c_char_p(msg),
            ctypes.c_ulonglong(len(msg)),
            ctypes.c_char_p(pk_raw),
        )
        if ret != 0:
            raise RuntimeError(f"crypto_box_seal returned {ret}")
        return base64.b64encode(bytes(cipher_buf)).decode("utf-8")
    except Exception as e:
        st.session_state["_nacl_err"] = f"{type(e).__name__}: {e}"
        return None


def _gh_set_secret(owner: str, repo: str, token: str, name: str, value: str) -> tuple[bool, str]:
    r = _gh_api("get", f"/repos/{owner}/{repo}/actions/secrets/public-key", token)
    if r.status_code != 200:
        return False, f"Could not fetch repo public key (HTTP {r.status_code})"
    pk_data = r.json()
    st.session_state.pop("_nacl_err", None)
    enc = _gh_encrypt_secret(pk_data["key"], value)
    if enc is None:
        nacl_err = st.session_state.pop("_nacl_err", "")
        if nacl_err:
            return False, f"Encryption error ({nacl_err}) — add this secret manually in GitHub → Settings → Secrets → Actions"
        return False, "PyNaCl unavailable — add this secret manually in GitHub → Settings → Secrets → Actions"
    r2 = _gh_api("put", f"/repos/{owner}/{repo}/actions/secrets/{name}", token,
                  data={"encrypted_value": enc, "key_id": pk_data["key_id"]})
    if r2.status_code in (201, 204):
        return True, "OK"
    return False, f"HTTP {r2.status_code}: {r2.text[:120]}"


def _gh_push_file(owner: str, repo: str, token: str,
                   path: str, content: str, message: str) -> tuple[bool, str]:
    """Create or update a file in the repo via GitHub Contents API."""
    # Check if file exists (need its SHA to update)
    r_get = _gh_api("get", f"/repos/{owner}/{repo}/contents/{path}", token)
    sha = r_get.json().get("sha") if r_get.status_code == 200 else None
    body: dict = {
        "message": message,
        "content": base64.b64encode(content.encode()).decode(),
    }
    if sha:
        body["sha"] = sha
    r = _gh_api("put", f"/repos/{owner}/{repo}/contents/{path}", token, data=body)
    if r.status_code in (200, 201):
        return True, "OK"
    return False, f"HTTP {r.status_code}: {r.text[:200]}"


def _parse_repo_url(url: str) -> tuple[str, str] | None:
    """Extract owner and repo name from a GitHub URL."""
    m = re.search(r"github\.com[/:]([^/]+)/([^/.\s]+)", url.strip().rstrip("/"))
    if m:
        return m.group(1), m.group(2).removesuffix(".git")
    return None


def _generate_workflow_yaml(del_hour: int, tz: str, provider: str, channels: list,
                             one_time: bool = False) -> str:
    """Generate the GitHub Actions workflow YAML."""
    import pytz
    try:
        tz_obj  = pytz.timezone(tz)
        # Convert local hour to UTC
        from datetime import datetime as _dt
        local = _dt.now(tz_obj).replace(hour=del_hour, minute=0, second=0, microsecond=0)
        utc_h = local.astimezone(pytz.utc).hour
    except Exception:
        utc_h = max(0, del_hour - 5)  # rough IST→UTC fallback

    data_h = (utc_h - 1) % 24

    # Build schedule block — omit cron lines for one-time jobs
    if one_time:
        schedule_block = "  # One-time job — trigger manually via workflow_dispatch"
    else:
        schedule_block = (
            f"  schedule:\n"
            f"    - cron: '0 {data_h} * * *'   # Data refresh  ({del_hour-1:02d}:00 {tz})\n"
            f"    - cron: '0 {utc_h} * * *'   # Narrative + delivery ({del_hour:02d}:00 {tz})"
        )

    secrets_env = "          GEMINI_API_KEY: ${{ secrets.GEMINI_API_KEY }}\n" if provider == "gemini" else ""
    if "slack" in channels:
        secrets_env += (
            "          SLACK_BOT_TOKEN: ${{ secrets.SLACK_BOT_TOKEN }}\n"
            "          SLACK_CHANNEL_ID: ${{ secrets.SLACK_CHANNEL_ID }}\n"
        )
    if "teams" in channels:
        secrets_env += "          TEAMS_WEBHOOK_URL: ${{ secrets.TEAMS_WEBHOOK_URL }}\n"
    if "email" in channels:
        secrets_env += (
            "          EMAIL_SENDER: ${{ secrets.EMAIL_SENDER }}\n"
            "          EMAIL_APP_PASSWORD: ${{ secrets.EMAIL_APP_PASSWORD }}\n"
            "          EMAIL_RECIPIENT: ${{ secrets.EMAIL_RECIPIENT }}\n"
        )

    return f"""name: ARIA Daily Briefing

on:
{schedule_block}
  workflow_dispatch:              # Manual trigger

jobs:
  aria-daily:
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4

      - name: Set up Python
        uses: actions/setup-python@v5
        with:
          python-version: '3.11'

      - name: Install dependencies
        run: pip install -r requirements.txt

      - name: Run ARIA
        env:
          GOOGLE_CREDS_JSON: ${{ secrets.GOOGLE_CREDS_JSON }}
{secrets_env.rstrip()}
        run: python agent/main.py

      - name: Upload narrative reports
        if: always()
        uses: actions/upload-artifact@v4
        with:
          name: aria-reports
          path: output/
          retention-days: 30
"""


# ════════════════════════════════════════════════════════════════════════════════
# PROGRESS BAR
# ════════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════════
# SCREEN: USER DASHBOARD
# ════════════════════════════════════════════════════════════════════════════════

def screen_dashboard():
    user     = st.session_state.get("user", {})
    name     = user.get("name", "User")
    email    = user.get("email", "")
    first    = name.split()[0]

    st.markdown(
        f'<h2 style="margin-bottom:4px">👋 Hi, {first}!</h2>'
        f'<p style="color:#9CA3AF;margin-top:0">Here\'s your ARIA activity at a glance.</p>',
        unsafe_allow_html=True,
    )

    # ── Stats row ─────────────────────────────────────────────────────────── #
    schedules = [s for s in _gs_read("ARIA_Schedules")
                 if s.get("user_email", "").lower() == email.lower()
                 and s.get("status") == "active"]
    activity  = [a for a in _gs_read("ARIA_Activity")
                 if a.get("user_email", "").lower() == email.lower()]

    total_runs   = sum(1 for a in activity if a.get("action_type") in
                       ("wizard_completed", "one_time_card"))
    last_ts      = max((a.get("timestamp","") for a in activity), default="—")
    active_scheds = len(schedules)

    s1, s2, s3 = st.columns(3)
    for col, icon, val, label in [
        (s1, "📅", str(active_scheds), "Active Schedules"),
        (s2, "🚀", str(total_runs),    "Cards Generated"),
        (s3, "🕐", last_ts[:10] if last_ts != "—" else "—", "Last Activity"),
    ]:
        col.markdown(
            f'<div style="background:#111827;border:1px solid #374151;border-radius:12px;'
            f'padding:20px;text-align:center">'
            f'<div style="font-size:28px">{icon}</div>'
            f'<div style="font-size:26px;font-weight:800;margin:6px 0">{val}</div>'
            f'<div style="font-size:12px;color:#9CA3AF">{label}</div>'
            f'</div>', unsafe_allow_html=True,
        )

    st.divider()

    # ── Tabs: Schedules / Activity / Settings ─────────────────────────────── #
    t_sched, t_act, t_settings = st.tabs(["📅 My Schedules", "📜 Activity History", "⚙️ Settings"])

    # ── Schedules tab ─────────────────────────────────────────────────────── #
    with t_sched:
        if not schedules:
            st.info("No active schedules yet. Run the wizard to create your first ARIA briefing!", icon="📅")
        for s in schedules:
            with st.container():
                c1, c2, c3 = st.columns([4, 3, 1])
                c1.markdown(
                    f"**{s.get('role_name','?')}** — {s.get('schedule_type','daily').title()}\n\n"
                    f"📬 {s.get('delivery_channels','—')}  ·  "
                    f"⏰ {s.get('delivery_hour','?')}:00 {s.get('timezone','')}"
                )
                repo_url = s.get("repo_url", "")
                if repo_url:
                    c2.markdown(f"[Open repo ↗]({repo_url})")
                c2.caption(f"Created {str(s.get('created_at',''))[:10]}")
                sid = s.get("schedule_id", "")
                if sid and c3.button("🗑️", key=f"del_{sid}", help="Delete this schedule"):
                    delete_schedule(sid)
                    st.success(f"Schedule {sid} deleted.")
                    _gs_read.clear()
                    st.rerun()
                st.divider()

    # ── Activity tab ──────────────────────────────────────────────────────── #
    with t_act:
        if not activity:
            st.info("No activity recorded yet.", icon="📜")
        else:
            _ICONS = {
                "register": "✨", "login": "🔑", "wizard_completed": "🚀",
                "one_time_card": "📄", "delete_schedule": "🗑️", "preview_card": "🎨",
            }
            for a in sorted(activity, key=lambda x: x.get("timestamp",""), reverse=True)[:50]:
                icon = _ICONS.get(a.get("action_type",""), "📌")
                ts   = str(a.get("timestamp",""))[:16]
                act  = a.get("action_type","").replace("_", " ").title()
                det  = a.get("details","")
                st.markdown(f"{icon} **{act}** — {det}  \n`{ts} UTC`")

    # ── Settings tab ──────────────────────────────────────────────────────── #
    with t_settings:
        st.subheader("Change Password")
        old_pw  = st.text_input("Current password", type="password", key="chpw_old")
        new_pw  = st.text_input("New password (min 6 chars)", type="password", key="chpw_new")
        new_pw2 = st.text_input("Confirm new password", type="password", key="chpw_new2")
        if st.button("Update Password", key="btn_chpw"):
            if new_pw != new_pw2:
                st.error("New passwords don't match.")
            elif len(new_pw) < 6:
                st.error("Password must be at least 6 characters.")
            else:
                ok, msg = change_password(email, old_pw, new_pw)
                st.success(msg) if ok else st.error(msg)

        st.divider()
        st.subheader("Account Info")
        st.markdown(f"**Name:** {name}  \n**Email:** {email}  \n**Member since:** {str(user.get('created_at',''))[:10]}")


# ════════════════════════════════════════════════════════════════════════════════
# SCREEN: ADMIN PANEL
# ════════════════════════════════════════════════════════════════════════════════

def screen_admin():
    st.markdown("## 🛡️ Admin Panel")
    st.caption("Visible only to you — all user and activity data across ARIA.")

    all_users  = _gs_read("ARIA_Users")
    all_acts   = _gs_read("ARIA_Activity")
    all_scheds = _gs_read("ARIA_Schedules")

    # ── Top stats ─────────────────────────────────────────────────────────── #
    a1, a2, a3, a4 = st.columns(4)
    for col, icon, val, label in [
        (a1, "👥", str(len(all_users)),   "Total Users"),
        (a2, "📜", str(len(all_acts)),    "Total Events"),
        (a3, "📅", str(sum(1 for s in all_scheds if s.get("status") == "active")),
                                          "Active Schedules"),
        (a4, "🚀", str(sum(1 for a in all_acts
                           if a.get("action_type") == "wizard_completed")),
                                          "Completed Setups"),
    ]:
        col.markdown(
            f'<div style="background:#1F2937;border:1px solid #374151;border-radius:10px;'
            f'padding:16px;text-align:center">'
            f'<div style="font-size:24px">{icon}</div>'
            f'<div style="font-size:22px;font-weight:800;margin:4px 0">{val}</div>'
            f'<div style="font-size:11px;color:#9CA3AF">{label}</div>'
            f'</div>', unsafe_allow_html=True,
        )

    st.divider()

    ta, tb, tc = st.tabs(["👥 Users", "📜 Recent Activity", "📅 All Schedules"])

    with ta:
        if not all_users:
            st.info("No users yet.")
        else:
            _cols = ["email", "name", "created_at", "last_login", "is_admin"]
            import pandas as _pd2
            df_u = _pd2.DataFrame(all_users)[_cols] if all_users else _pd2.DataFrame(columns=_cols)
            df_u["is_admin"] = df_u["is_admin"].apply(lambda x: "✅ Admin" if str(x) == "1" else "—")
            df_u.columns = ["Email", "Name", "Joined", "Last Login", "Role"]
            st.dataframe(df_u, use_container_width=True, hide_index=True)

    with tb:
        recent = sorted(all_acts, key=lambda x: x.get("timestamp",""), reverse=True)[:100]
        if not recent:
            st.info("No activity recorded.")
        else:
            import pandas as _pd3
            df_a = _pd3.DataFrame(recent)[["timestamp","user_name","user_email","action_type","details"]]
            df_a.columns = ["Time (UTC)","User","Email","Action","Details"]
            st.dataframe(df_a, use_container_width=True, hide_index=True)

    with tc:
        active_s = [s for s in all_scheds if s.get("status") == "active"]
        if not active_s:
            st.info("No active schedules.")
        else:
            import pandas as _pd4
            _sc = ["user_name","user_email","role_name","schedule_type",
                   "delivery_channels","timezone","delivery_hour","created_at","repo_url"]
            df_s = _pd4.DataFrame(active_s)[[c for c in _sc if c in _pd4.DataFrame(active_s).columns]]
            st.dataframe(df_s, use_container_width=True, hide_index=True)


def render_progress():
    step  = st.session_state.get("step", 1)
    # Active step: white text on green circle in dark mode, dark text on green circle in light mode
    _active_tc      = "#ffffff" if _is_dark else "#111827"
    _inactive_tc    = "#6B7280"
    _inactive_lc    = "#D1D5DB" if not _is_dark else "#374151"
    items = ""
    for i, label in enumerate(STEP_LABELS, 1):
        if i < step:
            c, bg, tc, sym = "#10B981", "#10B98120", "#10B981", "✓"
        elif i == step:
            c, bg, tc, sym = "#10B981", "#10B981", _active_tc, str(i)
        else:
            c, bg, tc, sym = (_inactive_lc, "transparent", _inactive_tc, str(i))
        items += (
            f'<div style="display:flex;flex-direction:column;align-items:center;min-width:64px">'
            f'<div style="width:36px;height:36px;border-radius:50%;background:{bg};border:2.5px solid {c};'
            f'display:flex;align-items:center;justify-content:center;color:{tc};font-size:13px;'
            f'font-weight:700;margin-bottom:5px">{sym}</div>'
            f'<div style="font-size:11px;color:{tc};text-align:center;line-height:1.3;max-width:60px">{label}</div>'
            f'</div>'
        )
        if i < TOTAL_STEPS:
            lc = "#10B981" if i < step else _inactive_lc
            items += f'<div style="flex:1;height:2px;background:{lc};margin:18px 2px 0"></div>'
    st.markdown(
        f'<div style="display:flex;align-items:flex-start;padding:16px 0 24px;overflow-x:auto">{items}</div>',
        unsafe_allow_html=True,
    )


def nav_buttons(back=True, next_label="Next →", next_disabled=False):
    cols = st.columns([1, 2, 2])
    if back and st.session_state.step > 1:
        if cols[0].button("← Back", use_container_width=True):
            st.session_state.step -= 1
            st.rerun()
    if cols[2].button(next_label, type="primary", disabled=next_disabled, use_container_width=True):
        st.session_state.step += 1
        st.rerun()


def _resolve_role() -> dict:
    role_name = st.session_state.get("role_name", "CEO")
    for grp in ROLE_GROUPS.values():
        if role_name in grp:
            return grp[role_name]
    return ROLE_GROUPS["🏛️ C-Suite"]["CEO"]


def _clear_preview():
    for k in ("narrative", "narrative_obj", "metrics"):
        st.session_state.pop(k, None)


def _clear_kpi_cache():
    """Wipe KPI discovery state so Step 4 re-runs LLM detection on new data."""
    for k in ("kpis", "detected_domain", "detected_domain_label",
              "detected_domain_emoji", "date_col", "role_kpi_map",
              "_kpi_llm_error", "_kpi_used_fallback"):
        st.session_state.pop(k, None)
    # Also delete the on-disk cache so stale domain results don't persist
    try:
        cache_path = _ROOT / "aria_kpi_cache.json"
        if cache_path.exists():
            cache_path.unlink()
    except Exception:
        pass


# ════════════════════════════════════════════════════════════════════════════════
# STEP 1 — WELCOME
# ════════════════════════════════════════════════════════════════════════════════

def step_welcome():
    user   = st.session_state.get("user", {})
    name   = user.get("name", "")
    first  = name.split()[0] if name else ""
    email  = user.get("email", "")

    # ── Personalised hero banner ──────────────────────────────────────────── #
    if first:
        # Returning user — show greeting + quick stats
        activity   = [a for a in _gs_read("ARIA_Activity")
                      if a.get("user_email","").lower() == email.lower()]
        total_cards = sum(1 for a in activity if a.get("action_type") in
                          ("wizard_completed","one_time_card"))
        last_ts    = max((a.get("timestamp","") for a in activity), default="")
        last_label = f"Last active {last_ts[:10]}" if last_ts else "Welcome back!"

        initials = "".join(w[0].upper() for w in name.split()[:2]) or "U"
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:20px;'
            f'background:linear-gradient(135deg,#10B98118 0%,#0B122000 100%);'
            f'border:1px solid #10B98130;border-radius:16px;padding:24px 28px;margin-bottom:24px">'
            f'<div style="width:64px;height:64px;border-radius:50%;background:#10B981;'
            f'display:flex;align-items:center;justify-content:center;'
            f'color:#fff;font-size:24px;font-weight:800;flex-shrink:0">{initials}</div>'
            f'<div>'
            f'<div style="font-size:24px;font-weight:800">👋 Hi, {first}!</div>'
            f'<div style="color:#9CA3AF;font-size:13px;margin-top:4px">{last_label}'
            f'{"  ·  " + str(total_cards) + " card" + ("s" if total_cards != 1 else "") + " generated" if total_cards else ""}'
            f'</div>'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown("""
        <div style="text-align:center;padding:32px 0 24px">
          <div style="font-size:56px;margin-bottom:8px">⚡</div>
          <h1 style="font-size:32px;font-weight:800;margin:0">Meet ARIA</h1>
          <p style="font-size:16px;color:#9CA3AF;margin:8px 0 0">
            Autonomous Report &amp; Insight AI Agent
          </p>
        </div>
        """, unsafe_allow_html=True)

    # ── Feature highlights ────────────────────────────────────────────────── #
    _feat_bg     = "#111827" if _is_dark else "#F3F4F6"
    _feat_border = "#374151" if _is_dark else "#D1D5DB"
    _feat_title  = "#F9FAFB" if _is_dark else "#111827"
    _feat_body   = "#9CA3AF" if _is_dark else "#4B5563"
    f1, f2, f3 = st.columns(3)
    for col, icon, title, body in [
        (f1, "📊", "Any Data",
         "Upload Excel, CSV, or connect a Google Sheet. ARIA auto-detects your KPIs."),
        (f2, "🎨", "Role-Personalised",
         "CEO, CFO, Sales Head, Business Analyst — each gets a different card, tone, and focus."),
        (f3, "⚡", "Zero Effort",
         "Set up once. ARIA posts to Slack, Teams, or email every morning — no manual work."),
    ]:
        col.markdown(
            f'<div style="background:{_feat_bg};border:1px solid {_feat_border};border-radius:12px;'
            f'padding:20px;height:100%">'
            f'<div style="font-size:28px;margin-bottom:8px">{icon}</div>'
            f'<div style="font-weight:700;margin-bottom:6px;color:{_feat_title}">{title}</div>'
            f'<div style="font-size:12px;color:{_feat_body};line-height:1.6">{body}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
ARIA transforms your business data into **AI-powered insight cards** — hyper-personalised for every role, delivered to the right person, every morning, automatically. Each card includes:
a **hero KPI** with MoM / YoY deltas, a **30-day sparkline**, **driver bar charts**, a **recommended action** with owner and timeline, and **speaker notes** crafted for your specific role and priorities.

This wizard takes **5 minutes.** At the end you'll have a live preview card built from your real data — with ARIA running on autopilot.
    """)

    st.info("📊 Works with any business dataset — sales, operations, finance, marketing, and more.", icon="ℹ️")

    _, col, _ = st.columns([1, 2, 1])
    btn_label = "Continue →" if first else "Let's build your ARIA →"
    if col.button(btn_label, type="primary", use_container_width=True):
        st.session_state.step = 2
        st.rerun()


# ════════════════════════════════════════════════════════════════════════════════
# STEP 2 — UPLOAD DATA
# ════════════════════════════════════════════════════════════════════════════════

def step_upload_data():
    st.header("📁 Upload Your Data")
    st.caption("Excel (.xlsx / .xls) or CSV. Needs a date column and at least one numeric column.")

    # ── Data guidance callout ─────────────────────────────────────────────── #
    with st.expander("📋 What to upload — and what to keep private", expanded=False):
        col_why, col_do, col_dont = st.columns(3)
        col_why.markdown(
            "**🤔 Why upload data?**\n\n"
            "ARIA reads your numbers to detect trends, compute KPIs (Sales, Profit, "
            "Orders, Margin), spot anomalies, and write a personalised briefing card "
            "tailored to your role. Without data, ARIA works with a sample — which "
            "is fine for a demo, but not for your real morning briefing."
        )
        col_do.markdown(
            "**✅ Good data to upload**\n\n"
            "- Sales or revenue exports (e.g. from Salesforce, SAP, Excel)\n"
            "- Order or transaction logs\n"
            "- Ops / fulfilment reports\n"
            "- Any tabular file with a **date column** + at least one **numeric column**\n"
            "- 100 – 500K rows works well\n"
            "- Column names in plain English are ideal"
        )
        col_dont.markdown(
            "**🚫 Do NOT upload**\n\n"
            "- Customer names, emails, phone numbers, addresses (PII)\n"
            "- Confidential client contracts or pricing agreements\n"
            "- HR, salary, or headcount data\n"
            "- Passwords, API keys, or credentials\n"
            "- Legally restricted or NDA-covered data\n\n"
            "💡 Use anonymised or aggregated exports — "
            "ARIA only needs the numbers, not the names."
        )
        st.caption(
            "⚠️ Your file is processed locally in this session and is not stored "
            "by ARIA or Streamlit after you close the browser."
        )

    tab1, tab2 = st.tabs(["📤 Upload File", "🔗 Google Sheets URL"])

    with tab1:
        uploaded = st.file_uploader("Drop your file here", type=["xlsx", "xls", "csv"])
        if uploaded:
            try:
                # Store raw bytes for Step 8 GitHub push (binary-safe)
                _raw_bytes = uploaded.getvalue()
                if uploaded.name.endswith(".csv"):
                    df = pd.read_csv(uploaded)
                    _sheet_name = "Sheet1"
                else:
                    import openpyxl as _oxl
                    _wb = _oxl.load_workbook(
                        __import__("io").BytesIO(_raw_bytes), read_only=True)
                    _sheet_name = _wb.sheetnames[0] if _wb.sheetnames else "Sheet1"
                    _wb.close()
                    df = pd.read_excel(
                        __import__("io").BytesIO(_raw_bytes),
                        sheet_name=_sheet_name)
                st.session_state.df                  = df
                st.session_state.data_source         = "file"
                st.session_state.data_name           = uploaded.name
                st.session_state.uploaded_file_bytes = _raw_bytes
                st.session_state.excel_sheet_name    = _sheet_name
                _clear_preview()
                _clear_kpi_cache()
                st.success(f"✅ Loaded {len(df):,} rows × {len(df.columns)} columns from **{uploaded.name}**")
                st.dataframe(df.head(5), use_container_width=True)
            except Exception as e:
                st.error(f"Could not read file: {e}")

    with tab2:
        gc1, gc2 = st.columns([6, 1])
        url = gc1.text_input(
            "Google Sheets URL",
            placeholder="https://docs.google.com/spreadsheets/d/...",
            label_visibility="collapsed",
        )
        with gc2:
            help_tip("google_sheets")
        if st.button("Load from Sheets") and url:
            with st.spinner("Fetching sheet…"):
                try:
                    csv_url = url.split("/edit")[0] + "/export?format=csv&gid=0"
                    df = pd.read_csv(csv_url)
                    st.session_state.df               = df
                    st.session_state.data_source      = "google_sheets"
                    st.session_state.google_sheet_url = url
                    _clear_preview()
                    _clear_kpi_cache()
                    st.success(f"✅ Loaded {len(df):,} rows from Google Sheets")
                    st.dataframe(df.head(5), use_container_width=True)
                except Exception as e:
                    st.error(f"Could not load: {e}")

    # ── Target / Forecast / Plan data (optional) ─────────────────────────── #
    st.divider()
    st.subheader("🎯 Target / Forecast / Plan Data  *(optional)*")
    st.caption(
        "Upload your annual plan, forecast, or budget file. "
        "ARIA will automatically compute **Achievement %** and **Gap to Target** "
        "for each KPI and add them to your card. "
        "Skip this if you don't have target data — the wizard continues normally."
    )

    _has_target = st.session_state.get("df_target") is not None
    _tgt_col1, _tgt_col2 = st.columns([3, 1])

    with _tgt_col1:
        target_uploaded = st.file_uploader(
            "Drop your target / plan file here (same format as your data file)",
            type=["xlsx", "xls", "csv"],
            key="target_file_uploader",
        )

    with _tgt_col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if _has_target:
            if st.button("🗑️ Remove target data", use_container_width=True):
                st.session_state.pop("df_target", None)
                st.session_state.pop("target_data_name", None)
                _clear_preview()
                st.rerun()

    if target_uploaded:
        try:
            _tb = target_uploaded.getvalue()
            if target_uploaded.name.endswith(".csv"):
                df_target = pd.read_csv(__import__("io").BytesIO(_tb))
            else:
                df_target = pd.read_excel(__import__("io").BytesIO(_tb))
            st.session_state["df_target"]      = df_target
            st.session_state["target_data_name"] = target_uploaded.name
            _clear_preview()
            st.success(
                f"✅ Target data loaded: **{target_uploaded.name}** "
                f"— {len(df_target):,} rows × {len(df_target.columns)} columns. "
                "ARIA will compute Achievement % and Gap to Target for each KPI."
            )
            st.dataframe(df_target.head(3), use_container_width=True)
        except Exception as e:
            st.error(f"Could not read target file: {e}")
    elif _has_target:
        _tn = st.session_state.get("target_data_name", "Target file")
        _td = st.session_state["df_target"]
        st.success(
            f"✅ Target data active: **{_tn}** — {len(_td):,} rows. "
            "Achievement % and Gap KPIs will appear on your card."
        )

    st.divider()
    st.caption("No file? Use the live Superstore sample (refreshes daily).")
    if st.button("Use Superstore sample data"):
        _sample_url = "https://docs.google.com/spreadsheets/d/1eEfyVh4VmFlmV6ZzqJkJ_ZTaogoKrE8O7_1pebFHW34/edit?gid=0"
        with st.spinner("Loading live Superstore data…"):
            try:
                _csv_url = _sample_url.split("/edit")[0] + "/export?format=csv&gid=0"
                df = pd.read_csv(_csv_url)
                st.session_state.df               = df
                st.session_state.data_source      = "google_sheets"
                st.session_state.google_sheet_url = _sample_url
                st.session_state.data_name        = "Superstore Live (Google Sheets)"
                _clear_kpi_cache()
                _clear_preview()
                st.success(f"✅ Loaded {len(df):,} rows from Superstore Live")
                st.rerun()
            except Exception as e:
                st.error(f"Could not load Superstore sample: {e}")

    nav_buttons(back=True, next_label="Next: Discover KPIs →",
                next_disabled="df" not in st.session_state)


# ════════════════════════════════════════════════════════════════════════════════
# STEP 3 — DISCOVER KPIs
# ════════════════════════════════════════════════════════════════════════════════

def step_discover_kpis():
    st.header("🔬 Discover KPIs")
    st.caption("ARIA analysed your dataset and generated domain-specific KPIs. Review, rename, or disable each one.")

    df: pd.DataFrame = st.session_state.get("df")
    if df is None:
        st.warning("No data loaded — go back to Step 2.")
        nav_buttons(back=True, next_label="Next →", next_disabled=True)
        return

    if "date_col" not in st.session_state:
        st.session_state.date_col = detect_date_col(df)

    # ── LLM KPI detection (runs once per dataset, cached) ────────────────── #
    if "kpis" not in st.session_state:
        provider = st.session_state.get("ai_provider", "stub")
        gemini_key_present = bool(
            st.session_state.get("gemini_key") or
            st.session_state.get("gemini_key_input")
        )
        if provider == "gemini" and gemini_key_present:
            with st.spinner("🧠 ARIA is analysing your dataset with Gemini 2.5 Flash…"):
                st.session_state.kpis = detect_kpis_llm(df)
        elif provider == "stub":
            # Stub selected — try Gemini anyway if key exists, else free LLM
            with st.spinner("🧠 ARIA is analysing your dataset…"):
                st.session_state.kpis = detect_kpis_llm(df)
        else:
            with st.spinner("🧠 ARIA is analysing your dataset…"):
                st.session_state.kpis = detect_kpis_llm(df)

    # ── Role KPI assignment (runs once after KPI detection, cached) ──────── #
    if "role_kpi_map" not in st.session_state and st.session_state.get("kpis"):
        _, cache_key = _kpi_cache_key(df)
        d_label_for_assign = st.session_state.get(
            "detected_domain_label",
            st.session_state.get("detected_domain", "Business")
        )
        with st.spinner("🎯 Assigning KPIs to roles…"):
            st.session_state.role_kpi_map = assign_kpis_to_roles(
                st.session_state.kpis, d_label_for_assign, cache_key
            )

    # ── Domain badge ─────────────────────────────────────────────────────── #
    d_emoji = st.session_state.get("detected_domain_emoji", "📊")
    d_label = st.session_state.get("detected_domain_label",
              st.session_state.get("detected_domain", "Business Analytics"))
    st.markdown(
        f'<div style="display:inline-flex;align-items:center;gap:8px;'
        f'background:#1F2937;border:1px solid #374151;border-radius:20px;'
        f'padding:6px 16px;margin-bottom:12px">'
        f'<span style="font-size:18px">{d_emoji}</span>'
        f'<span style="font-size:13px;font-weight:600;color:#D1D5DB">Detected domain: '
        f'<span style="color:#60A5FA">{d_label}</span></span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if "date_col" not in st.session_state:
        st.session_state.date_col = detect_date_col(df)
    st.session_state.date_col = st.selectbox(
        "📅 Date / time column",
        options=df.columns.tolist(),
        index=df.columns.tolist().index(st.session_state.date_col)
              if st.session_state.date_col in df.columns else 0,
        help="ARIA uses this for period-over-period comparisons. "
             "You will choose your data window in the Preview step.",
    )

    st.divider()

    # ── Show LLM error / fallback notice if applicable ───────────────────── #
    llm_err = st.session_state.get("_kpi_llm_error")
    used_fallback = st.session_state.get("_kpi_used_fallback")
    if used_fallback and not llm_err:
        # Silent fallback — no error, just a subtle note
        pass
    elif llm_err:
        # Gemini failed — show collapsible detail, not a scary red banner
        with st.expander("ℹ️ Gemini unavailable — using built-in analytics engine", expanded=False):
            st.caption(f"Technical detail: {llm_err}")
            st.caption("KPIs below were detected using ARIA's built-in domain pattern engine. Results are accurate for your dataset.")
            if st.button("🔄 Retry with Gemini", key="retry_kpi_btn"):
                _clear_kpi_cache()
                st.session_state.pop("kpis", None)
                st.session_state.pop("role_kpi_map", None)
                st.session_state.pop("_kpi_llm_error", None)
                st.session_state.pop("_kpi_used_fallback", None)
                st.rerun()

    kpis = st.session_state.get("kpis", [])
    if not kpis:
        st.warning("⚠️ ARIA could not generate KPIs. This can happen if both Gemini and the fallback LLM are unavailable, or the dataset has no numeric columns.")
        if st.button("🔄 Retry KPI Detection"):
            _clear_kpi_cache()
            st.session_state.pop("_kpi_llm_error", None)
            st.session_state.pop("_kpi_used_fallback", None)
            st.rerun()
        nav_buttons(back=True, next_label="Next →", next_disabled=True)
        return
    else:
        # ── KPI type legend ──────────────────────────────────────────────── #
        st.markdown(
            '<div style="display:flex;gap:16px;margin-bottom:8px;flex-wrap:wrap">'
            '<span style="font-size:11px;background:#1E3A2F;color:#34D399;'
            'border-radius:4px;padding:2px 8px">◆ Direct</span>'
            '<span style="font-size:11px;background:#1E2A4A;color:#60A5FA;'
            'border-radius:4px;padding:2px 8px">⬡ Derived</span>'
            '<span style="font-size:11px;color:#6B7280;font-style:italic">'
            '— Direct = raw column &nbsp;|&nbsp; Derived = computed ratio or formula</span>'
            '</div>',
            unsafe_allow_html=True,
        )

        st.subheader(f"ARIA's AI-Generated KPIs  ({len(kpis)} found)")
        h1, h2, h3, h4, h5 = st.columns([0.4, 1.8, 0.8, 2.2, 3])
        h1.caption("On")
        h2.caption("KPI Name")
        h3.caption("Type")
        h4.caption("Formula")
        h5.caption("Description")

        updated: list = []
        for i, k in enumerate(kpis):
            c1, c2, c3, c4, c5 = st.columns([0.4, 1.8, 0.8, 2.2, 3])
            enabled   = c1.checkbox("", value=k.get("enabled", True), key=f"en_{i}",
                                    label_visibility="collapsed")
            user_name = c2.text_input("n", value=k["user_name"], key=f"nm_{i}",
                                      label_visibility="collapsed", disabled=not enabled)
            ktype = k.get("kpi_type", "direct")
            if ktype == "derived":
                c3.markdown(
                    '<span style="font-size:11px;background:#1E2A4A;color:#60A5FA;'
                    'border-radius:4px;padding:2px 6px">⬡ Derived</span>',
                    unsafe_allow_html=True,
                )
            else:
                c3.markdown(
                    '<span style="font-size:11px;background:#1E3A2F;color:#34D399;'
                    'border-radius:4px;padding:2px 6px">◆ Direct</span>',
                    unsafe_allow_html=True,
                )
            c4.caption(f"`{k['formula']}`")
            c5.caption(k["description"])
            updated.append({**k, "enabled": enabled, "user_name": user_name})
        st.session_state.kpis = updated

        # ── Keep role_kpi_map in sync with checkbox state ─────────────────── #
        # Every time the user ticks/unticks a KPI, remove disabled names from
        # the cached role assignment so Step 4 never shows unchecked KPIs.
        if "role_kpi_map" in st.session_state:
            _enabled_set = {k["user_name"] for k in updated if k.get("enabled")}
            st.session_state.role_kpi_map = {
                role: [kpi for kpi in kpi_list if kpi in _enabled_set]
                for role, kpi_list in st.session_state.role_kpi_map.items()
            }

        # ── Show auto-dedup notice ─────────────────────────────────────────── #
        _auto_dupes = [k for k in updated if k.get("_duplicate_of") and not k.get("enabled")]
        if _auto_dupes:
            with st.expander(
                f"ℹ️ {len(_auto_dupes)} duplicate KPI{'s' if len(_auto_dupes) > 1 else ''} "
                f"auto-disabled — click to review",
                expanded=False,
            ):
                st.caption(
                    "These KPIs measure the same thing as an already-active KPI. "
                    "Re-enable them above if you need both."
                )
                for dk in _auto_dupes:
                    st.caption(
                        f"• **{dk['user_name']}** → same as **{dk['_duplicate_of']}**"
                    )

    st.divider()
    with st.expander("➕ Add a custom KPI"):
        m1, m2, m3, m4, m5 = st.columns([2, 2, 1.5, 1.5, 1.5])
        m_col  = m1.selectbox("Source Column", df.columns.tolist(), key="m_col")
        m_name = m2.text_input("KPI Name", key="m_name", placeholder="e.g. Avg Handle Time")
        m_agg  = m3.selectbox("Aggregation", ["sum", "mean", "nunique", "median", "max"], key="m_agg")
        m_fmt  = m4.selectbox("Format", ["currency", "integer", "percent", "decimal", "minutes", "score"], key="m_fmt")
        m_desc = m5.text_input("Description", key="m_desc", placeholder="What this measures")
        if st.button("➕ Add KPI") and m_name:
            _fml_map = {"sum": f"SUM({m_col})", "mean": f"AVG({m_col})",
                        "nunique": f"COUNT DISTINCT({m_col})", "median": f"MEDIAN({m_col})",
                        "max": f"MAX({m_col})"}
            st.session_state.kpis.append({
                "column": m_col, "suggested_name": m_name, "user_name": m_name,
                "formula": _fml_map.get(m_agg, f"SUM({m_col})"), "agg": m_agg,
                "format": m_fmt, "description": m_desc or "Custom KPI",
                "kpi_type": "direct", "enabled": True,
            })
            st.rerun()

    active = sum(1 for k in st.session_state.get("kpis", []) if k.get("enabled"))
    st.caption(f"ℹ️ {active} KPI{'s' if active != 1 else ''} active — these will appear on your preview card.")
    nav_buttons(back=True,
                next_label=f"Next: Pick Role → ({active} KPI{'s' if active != 1 else ''} active)",
                next_disabled=active == 0)


# ════════════════════════════════════════════════════════════════════════════════
# STEP 4 — PICK ROLE
# ════════════════════════════════════════════════════════════════════════════════

def step_pick_role():
    st.header("👤 Pick Your Role")
    st.caption("Each role drives a different hero KPI, driver focus, accent colour, and narrative tone.")

    if "role_name" not in st.session_state:
        st.session_state.role_name = "CEO"

    # Role-specific KPI map from Gemini assignment (or fallback).
    # Always filter against the CURRENTLY ENABLED KPIs — the user may have
    # unchecked some in Step 3 after the map was first generated.
    _enabled_names = {k["user_name"] for k in st.session_state.get("kpis", [])
                      if k.get("enabled")}
    _raw_role_kpi_map = st.session_state.get("role_kpi_map", {})
    _role_kpi_map = {
        role: [kpi for kpi in kpi_list if kpi in _enabled_names]
        for role, kpi_list in _raw_role_kpi_map.items()
    }
    # Flat fallback list if map is empty or all filtered out
    _flat_kpis = list(_enabled_names) or ["—"]

    for group_name, roles in ROLE_GROUPS.items():
        st.subheader(group_name)
        role_items = list(roles.items())
        # Render in rows of 4 so groups with 5+ roles wrap cleanly
        _ROW_SIZE = 4
        for row_start in range(0, len(role_items), _ROW_SIZE):
            row_items = role_items[row_start: row_start + _ROW_SIZE]
            cols = st.columns(len(row_items))
            for col, (role_key, rd) in zip(cols, row_items):
                is_sel  = st.session_state.role_name == role_key
                border  = rd["accent_color"] if is_sel else "#374151"
                bg      = f"{rd['accent_color']}18" if is_sel else "transparent"
                col.markdown(
                    f'<div style="border:2px solid {border};background:{bg};border-radius:10px;'
                    f'padding:14px 16px;margin-bottom:10px;min-height:88px">'
                    f'<div style="font-weight:700;font-size:13px;margin-bottom:3px">{role_key}</div>'
                    f'<div style="font-size:11px;color:#9CA3AF;margin-bottom:6px">{rd["title"]}</div>'
                    f'<div style="font-size:10px;color:{rd["accent_color"]}">'
                    f'{"  ·  ".join((_role_kpi_map.get(role_key) or _flat_kpis)[:3])}</div></div>',
                    unsafe_allow_html=True,
                )
                if col.button("✓ Selected" if is_sel else "Select",
                              key=f"role_{role_key}", use_container_width=True,
                              type="primary" if is_sel else "secondary"):
                    st.session_state.role_name = role_key
                    _clear_preview()
                    st.rerun()

    st.divider()
    with st.expander("🎨 Build a Custom Role"):
        c1, c2 = st.columns(2)
        c_title  = c1.text_input("Role Title",  placeholder="e.g. Regional Sales Director", key="c_title")
        c_badge  = c2.text_input("Badge Text",  placeholder="e.g. REGIONAL  ·  SALES BRIEFING", key="c_badge")
        avail    = [k["user_name"] for k in st.session_state.get("kpis", []) if k.get("enabled")] \
                   or ["Sales", "Profit", "Orders", "Margin%"]
        c3, c4, c5 = st.columns([2, 2, 1])
        c_kpis   = c3.multiselect("KPIs", avail, default=avail[:4], key="c_kpis")
        c_prim   = c4.selectbox("Primary KPI", c_kpis or avail, key="c_primary")
        c_color  = c5.color_picker("Accent", value="#60A5FA", key="c_color")
        c_tone   = st.text_area("Narrative tone",
                                 placeholder="Tactical. Focus on territory performance.", key="c_tone", height=70)
        if st.button("Use this custom role") and c_title:
            ROLE_GROUPS.setdefault("🎨 Custom", {})["Custom"] = {
                "title": c_title,
                "badge": c_badge or f"{c_title.upper()}  ·  BRIEFING",
                "primary_kpi": c_prim, "kpis": c_kpis or avail[:4],
                "accent_color": c_color, "driver_focus": ["Category", "Region"],
                "tone": c_tone or "Executive and strategic.",
            }
            st.session_state.role_name = "Custom"
            _clear_preview()
            st.success(f"Custom role '{c_title}' created!")
            st.rerun()

    nav_buttons(back=True,
                next_label=f"Next: Choose AI → (Role: {st.session_state.get('role_name','—')})",
                next_disabled=not st.session_state.get("role_name"))


# ════════════════════════════════════════════════════════════════════════════════
# STEP 5 — CHOOSE AI
# ════════════════════════════════════════════════════════════════════════════════

def step_choose_ai():
    st.header("🤖 Choose Your AI Engine")
    st.caption("Gemini is highly recommended for the best narrative quality — it's free and takes 2 minutes to set up.")

    if "ai_provider" not in st.session_state:
        st.session_state.ai_provider = "gemini"

    col1, col2 = st.columns(2)
    for col, key, emoji, title, tagline, pros, con, ok_border, ok_bg in [
        (col1, "stub",   "⚙️", "Built-in Engine",  "No API key. Instant.",
         ["Zero setup", "Data-accurate", "Deterministic", "Offline"], "Less natural prose",
         "#10B981", "#10B98118"),
        (col2, "gemini", "✨", "Gemini 2.5 Flash",  "Free · 1 500/day · Works in India.",
         ["Best narrative quality", "Free — no credit card", "1 500 req/day", "Works in India"],
         "Requires API key setup",
         "#F59E0B", "#F59E0B18"),
    ]:
        is_sel = st.session_state.ai_provider == key
        border = ok_border if is_sel else "#374151"
        bg     = ok_bg     if is_sel else "transparent"
        pros_html = "".join(f"✅ {p}<br>" for p in pros)
        col.markdown(
            f'<div style="border:2px solid {border};background:{bg};border-radius:12px;'
            f'padding:20px;min-height:200px;margin-bottom:8px">'
            f'<div style="font-size:26px;margin-bottom:6px">{emoji}</div>'
            f'<div style="font-weight:800;font-size:15px;margin-bottom:3px">{title}</div>'
            f'<div style="font-size:11px;color:#9CA3AF;margin-bottom:10px">{tagline}</div>'
            f'<div style="font-size:12px;line-height:1.9">{pros_html}⚠️ {con}</div></div>',
            unsafe_allow_html=True,
        )
        if col.button(f"Use {title}", key=f"btn_{key}",
                      type="primary" if is_sel else "secondary", use_container_width=True):
            st.session_state.ai_provider = key
            st.rerun()

    if st.session_state.ai_provider == "gemini":
        ki1, ki2 = st.columns([11, 1], vertical_alignment="top")
        with ki1:
            st.markdown(
                '<div style="border:1px solid #374151;border-radius:6px;'
                'height:38px;box-sizing:border-box;display:flex;align-items:center;'
                'padding:0 14px;font-size:13px;white-space:nowrap;overflow:hidden">'
                '🔑&nbsp; Get your free Gemini key at '
                '<a href="https://aistudio.google.com/app/apikey" target="_blank" '
                'style="color:#F59E0B;margin-left:4px">aistudio.google.com/app/apikey</a></div>',
                unsafe_allow_html=True,
            )
        with ki2:
            help_tip("gemini_key")
        kv = st.text_input("Paste key here (stored as a GitHub Secret, not locally)",
                            type="password", placeholder="Paste your Gemini API key here", key="gemini_key_input")
        if kv:
            st.session_state.gemini_key = kv
        if not st.session_state.get("gemini_key") and not kv:
            st.caption("👆 Paste your Gemini key above to continue.")

    _gemini_selected = st.session_state.get("ai_provider") == "gemini"
    # Check both session state and the live widget value so Next enables immediately
    _gemini_key_ok   = bool(
        st.session_state.get("gemini_key") or
        st.session_state.get("gemini_key_input")
    )
    _next_disabled   = _gemini_selected and not _gemini_key_ok

    nav_buttons(back=True, next_label="Next: Discover KPIs →", next_disabled=_next_disabled)


# ════════════════════════════════════════════════════════════════════════════════
# STEP 6 — SET DELIVERY  (with inline help tips)
# ════════════════════════════════════════════════════════════════════════════════

def step_set_delivery():
    st.header("📬 Set Delivery")
    st.caption("Choose how ARIA sends your card each morning.")

    if "delivery" not in st.session_state:
        st.session_state.delivery = {"file": True, "slack": False, "teams": False}

    # ── File ──────────────────────────────────────────────────────────────── #
    st.session_state.delivery["file"] = st.toggle("📄 Download File",
                                                    value=st.session_state.delivery["file"])
    if st.session_state.delivery["file"]:
        st.caption("Generates a Markdown + Word .docx — downloadable from GitHub Actions after each run.")

    st.divider()

    # ── Slack ─────────────────────────────────────────────────────────────── #
    st.session_state.delivery["slack"] = st.toggle("💬 Slack",
                                                     value=st.session_state.delivery["slack"])
    if st.session_state.delivery["slack"]:
        tk1, tk2 = st.columns([11, 1])
        st.session_state.slack_bot_token_input = tk1.text_input(
            "Slack Bot Token", type="password", placeholder="xoxb-...",
            key="slack_token",
        )
        with tk2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("slack_bot_token")

        ch1, ch2 = st.columns([11, 1])
        _ch_entered = ch1.text_input(
            "Channel ID", placeholder="e.g. C012AB3CD45",
            key="_slack_channel_widget",
            value=st.session_state.get("slack_channel", ""),
        )
        if _ch_entered:
            st.session_state["slack_channel"] = _ch_entered
        with ch2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("slack_channel_id")

        st.caption("Multiple channels? ARIA posts one card per role to its own channel automatically.")

    st.divider()

    # ── Teams ─────────────────────────────────────────────────────────────── #
    st.session_state.delivery["teams"] = st.toggle("🟦 Microsoft Teams",
                                                     value=st.session_state.delivery["teams"])
    if st.session_state.delivery["teams"]:
        tw1, tw2 = st.columns([11, 1])
        st.session_state.teams_webhook_input = tw1.text_input(
            "Teams Webhook URL", type="password",
            placeholder="https://your-org.webhook.office.com/...",
            key="teams_webhook",
        )
        with tw2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("teams_webhook")

    st.divider()

    # ── Email ─────────────────────────────────────────────────────────────── #
    if "email" not in st.session_state.delivery:
        st.session_state.delivery["email"] = False
    st.session_state.delivery["email"] = st.toggle("📧 Email",
                                                     value=st.session_state.delivery["email"])
    if st.session_state.delivery["email"]:
        em1, em2 = st.columns([11, 1])
        st.session_state.email_recipient = em1.text_input(
            "Recipient email address", placeholder="you@example.com",
            key="email_to",
            value=st.session_state.get("email_recipient", ""),
        )
        with em2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("email_delivery")

        st.caption(
            "ARIA sends the card as an email attachment every morning. "
            "Sender is configured via Gmail App Password in your GitHub Secrets."
        )

    st.divider()

    # ── Schedule ──────────────────────────────────────────────────────────── #
    st.subheader("⏰ Delivery Schedule")

    if "schedule_type" not in st.session_state:
        st.session_state.schedule_type = "daily"

    sched_choice = st.radio(
        "When should ARIA run?",
        options=["📅 Daily recurring", "📄 One-time report"],
        index=0 if st.session_state.schedule_type == "daily" else 1,
        horizontal=True,
        key="sched_radio",
    )
    st.session_state.schedule_type = "daily" if sched_choice == "📅 Daily recurring" else "onetime"

    if st.session_state.schedule_type == "daily":
        tz_opts = ["Asia/Kolkata", "Asia/Singapore", "Asia/Tokyo",
                   "America/Toronto", "America/New_York", "America/Los_Angeles",
                   "Europe/London", "Europe/Paris", "Australia/Sydney"]
        c_tz, c_hr = st.columns(2)
        st.session_state.timezone = c_tz.selectbox("Timezone", tz_opts, index=0, key="tz_sel")
        delivery_hour = c_hr.selectbox("Deliver at",
                                        [f"{h:02d}:00" for h in range(6, 13)],
                                        index=3, key="del_hr")
        st.session_state.delivery_hour = int(delivery_hour.split(":")[0])
    else:
        st.session_state.timezone = st.session_state.get("timezone", "Asia/Kolkata")
        st.session_state.delivery_hour = st.session_state.get("delivery_hour", 9)
        st.info(
            "ARIA will set up a one-time job. After Go Live, trigger it manually from "
            "GitHub Actions → **Run workflow** whenever you need a fresh briefing.",
            icon="📄",
        )

    active = [k for k, v in st.session_state.delivery.items() if v]
    nav_buttons(back=True, next_label="Next: Go Live →", next_disabled=len(active) == 0)


# ════════════════════════════════════════════════════════════════════════════════
# STEP 7 — PREVIEW CARD
# ════════════════════════════════════════════════════════════════════════════════

def step_preview_card():
    st.header("🎨 Card Studio")
    st.caption(
        "Design your briefing card. Choose a style, pick a time window, select a "
        "template — then preview the **actual card** that gets posted to Slack."
    )

    # Use the SAME role_cfg builder that _build_configs uses to write
    # roles.yaml at launch. This guarantees the preview's KPIs, template,
    # style, and primary_kpi match what the delivered card will use.
    role_cfg = build_runtime_role_cfg()

    # ────────────────────────────────────────────────────────────────────────── #
    # 1. CHOOSE CARD STYLE — solid background swatches, uniform 4 boxes
    # ────────────────────────────────────────────────────────────────────────── #
    st.subheader("1. Choose Your Card Style")

    if "card_style" not in st.session_state:
        st.session_state.card_style = "dark"

    # Inject CSS so all 4 style buttons look uniform and render cleanly
    st.markdown("""
    <style>
    div[data-testid="stHorizontalBlock"] > div { flex: 1 1 0 !important; }
    </style>
    """, unsafe_allow_html=True)

    style_cols = st.columns(4, gap="small")
    for col, (sk, sd) in zip(style_cols, CARD_STYLES.items()):
        is_sel    = st.session_state.card_style == sk
        swatch    = sd["swatch"]
        _light_bg = sk in ("beige", "grey")
        txt_col   = "#111111" if _light_bg else "#E5E7EB"
        sub_col   = "#555"    if _light_bg else "#9CA3AF"
        col.markdown(
            f'<div style="border:2px solid #374151;border-radius:10px;'
            f'padding:10px 8px;text-align:center;height:90px;box-sizing:border-box;'
            f'background:{swatch};cursor:pointer;margin-bottom:6px;">'
            f'<div style="font-size:13px;font-weight:800;color:{txt_col};'
            f'line-height:1.2;margin-top:4px;">{sd["label"]}</div>'
            f'<div style="font-size:9px;color:{sub_col};margin-top:5px;line-height:1.4;">'
            f'{sd["tagline"]}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if col.button("✓ Selected" if is_sel else "Select",
                      key=f"sty_{sk}", use_container_width=True,
                      type="primary" if is_sel else "secondary"):
            st.session_state.card_style = sk
            _clear_preview()
            st.rerun()

    st.divider()

    # ────────────────────────────────────────────────────────────────────────── #
    # 2. TIME PERIOD — single chip row, date range caption, custom picker
    # ────────────────────────────────────────────────────────────────────────── #
    st.subheader("2. Time Period")

    role_now       = st.session_state.get("role_name", "")
    role_suggested = ROLE_TIMEFRAME_DEFAULTS.get(role_now, "")
    current_tf     = st.session_state.get("timeframe_key")

    _df_preview    = st.session_state.get("df")
    _date_col_p    = st.session_state.get("date_col")
    _today         = date.today()
    _min_data_dt   = None
    _max_data_dt   = None

    if _df_preview is not None and _date_col_p and _date_col_p in _df_preview.columns:
        try:
            _parsed      = pd.to_datetime(_df_preview[_date_col_p], errors="coerce").dt.date
            _min_data_dt = _parsed.min()
            _max_data_dt = _parsed.max()
        except Exception:
            _min_data_dt = _max_data_dt = None

    def _tf_available(tf: dict) -> bool:
        # Period-to-date and open-ended keys are always available
        if tf["key"] in ("alltime", "ytd", "wtd", "mtd", "qtd"):
            return True
        if _max_data_dt is None:
            return True
        days_needed = tf.get("days", 1)
        if days_needed is None:
            return True
        if _min_data_dt is None:
            return True
        return (_max_data_dt - _min_data_dt).days >= days_needed

    # Auto-select best valid timeframe on first load
    if not current_tf and _max_data_dt is not None:
        _candidates = ([role_suggested] if role_suggested else []) + \
                      [tf["key"] for tf in TIMEFRAME_OPTIONS]
        for _cand in _candidates:
            _cand_tf = _tf_by_key(_cand)
            if _tf_available(_cand_tf):
                st.session_state.timeframe_key = _cand
                current_tf = _cand
                break

    # Single chip row — one button per timeframe, NO duplicate markdown row above
    _use_custom = st.session_state.get("timeframe_key") == "custom"
    tf_cols = st.columns([1] * len(TIMEFRAME_OPTIONS) + [1.2])  # Custom gets slight extra width
    for col, tf in zip(tf_cols, TIMEFRAME_OPTIONS):
        is_sel = current_tf == tf["key"] and not _use_custom
        is_dis = not _tf_available(tf)
        label  = f"✓ {tf['short']}" if is_sel else tf["short"]
        if col.button(
            label,
            key=f"tf_{tf['key']}",
            use_container_width=True,
            disabled=is_dis,
            type="primary" if is_sel else "secondary",
            help=tf["desc"],
        ):
            st.session_state.timeframe_key = tf["key"]
            _clear_preview()
            st.rerun()
    # Custom date range chip
    _cust_col = tf_cols[-1]
    _cust_sel = _use_custom
    if _cust_col.button(
        "✓ Custom" if _cust_sel else "Custom",
        key="tf_custom",
        use_container_width=True,
        type="primary" if _cust_sel else "secondary",
        help="Pick an exact start and end date from your data",
    ):
        st.session_state.timeframe_key = "custom"
        _clear_preview()
        st.rerun()

    # Refresh current_tf after button interactions
    current_tf = st.session_state.get("timeframe_key")

    # Custom date range picker
    if current_tf == "custom":
        _c1, _c2 = st.columns(2)
        _cust_start_default = _min_data_dt if _min_data_dt else (_today - timedelta(days=180))
        _cust_end_default   = _max_data_dt if _max_data_dt else _today
        _cust_start = _c1.date_input("Start date", value=st.session_state.get("custom_start", _cust_start_default),
                                     min_value=_min_data_dt, max_value=_max_data_dt, key="custom_start_inp")
        _cust_end   = _c2.date_input("End date",   value=st.session_state.get("custom_end",   _cust_end_default),
                                     min_value=_min_data_dt, max_value=_max_data_dt, key="custom_end_inp")
        st.session_state.custom_start = _cust_start
        st.session_state.custom_end   = _cust_end
        st.caption(
            f"📅 Selected range: **{_cust_start.strftime('%b %d, %Y')}** → "
            f"**{_cust_end.strftime('%b %d, %Y')}** "
            f"({(_cust_end - _cust_start).days} days)"
        )
    elif current_tf and current_tf != "custom":
        # Show the actual date range for the selected chip
        active_tf = _tf_by_key(current_tf)
        _ref = _max_data_dt if _max_data_dt else _today
        if active_tf:
            if current_tf == "alltime" and _min_data_dt:
                _range_start, _range_end = _min_data_dt, _ref
            elif current_tf == "ytd":
                _range_start = date(_ref.year, 1, 1)
                _range_end   = _ref
            elif current_tf == "wtd":
                # Monday of the current week
                _range_start = _ref - timedelta(days=_ref.weekday())
                _range_end   = _ref
            elif current_tf == "mtd":
                _range_start = date(_ref.year, _ref.month, 1)
                _range_end   = _ref
            elif current_tf == "qtd":
                _q_start_month = ((_ref.month - 1) // 3) * 3 + 1
                _range_start = date(_ref.year, _q_start_month, 1)
                _range_end   = _ref
            elif active_tf.get("days"):
                _range_end   = _ref
                _range_start = _range_end - timedelta(days=active_tf["days"])
            else:
                _range_start = _range_end = None
            if _range_start and _range_end:
                st.caption(
                    f"📅 **{_range_start.strftime('%b %d, %Y')}** → "
                    f"**{_range_end.strftime('%b %d, %Y')}** · {active_tf['comparison']}"
                )
        if not _max_data_dt:
            st.caption("📅 Upload data to see actual date range.")
    else:
        st.info("👆 Select a time window above to generate your card.", icon="📅")
        nav_buttons(back=True, next_label="Next: Set Delivery →")
        return

    if not current_tf:
        nav_buttons(back=True, next_label="Next: Set Delivery →")
        return

    st.divider()

    # ────────────────────────────────────────────────────────────────────────── #
    # 3. CHOOSE CARD TEMPLATE — mini SVG wireframe thumbnails
    # ────────────────────────────────────────────────────────────────────────── #
    st.subheader("3. Choose Your Card Template")

    if "card_template" not in st.session_state:
        st.session_state.card_template = "editorial"

    # Mini SVG wireframes representing each template's visual layout (80×56 viewBox)
    _TMPL_WIREFRAMES = {
        "editorial": """<svg viewBox="0 0 80 56" xmlns="http://www.w3.org/2000/svg">
          <rect width="80" height="56" rx="3" fill="#0B1220"/>
          <rect x="3" y="3" width="74" height="8" rx="1" fill="#1E293B"/>
          <rect x="3" y="14" width="28" height="16" rx="1" fill="#1E3A5F"/>
          <rect x="3" y="14" width="28" height="4" rx="1" fill="#3B82F6" opacity="0.6"/>
          <rect x="34" y="14" width="43" height="2" rx="1" fill="#374151"/>
          <rect x="34" y="18" width="35" height="2" rx="1" fill="#374151"/>
          <rect x="34" y="22" width="40" height="2" rx="1" fill="#374151"/>
          <rect x="3" y="33" width="74" height="2" rx="1" fill="#1F2937"/>
          <rect x="3" y="37" width="22" height="5" rx="1" fill="#374151"/>
          <rect x="27" y="37" width="22" height="5" rx="1" fill="#374151"/>
          <rect x="51" y="37" width="26" height="5" rx="1" fill="#374151"/>
          <polyline points="3,52 12,48 21,50 30,45 39,47 48,43 57,46 66,41 74,44" fill="none" stroke="#3B82F6" stroke-width="1.5"/>
        </svg>""",
        "scorecard": """<svg viewBox="0 0 80 56" xmlns="http://www.w3.org/2000/svg">
          <rect width="80" height="56" rx="3" fill="#0B1220"/>
          <rect x="3" y="3" width="74" height="6" rx="1" fill="#1E293B"/>
          <rect x="3" y="12" width="17" height="13" rx="2" fill="#1E3A5F"/>
          <rect x="22" y="12" width="17" height="13" rx="2" fill="#1E3A5F"/>
          <rect x="41" y="12" width="17" height="13" rx="2" fill="#1E3A5F"/>
          <rect x="60" y="12" width="17" height="13" rx="2" fill="#1E3A5F"/>
          <circle cx="25" cy="39" r="12" fill="none" stroke="#1F2937" stroke-width="5"/>
          <circle cx="25" cy="39" r="12" fill="none" stroke="#3B82F6" stroke-width="5" stroke-dasharray="30 45" stroke-dashoffset="-6"/>
          <circle cx="25" cy="39" r="12" fill="none" stroke="#10B981" stroke-width="5" stroke-dasharray="18 57" stroke-dashoffset="-36"/>
          <rect x="44" y="29" width="33" height="2" rx="1" fill="#374151"/>
          <rect x="44" y="34" width="28" height="2" rx="1" fill="#374151"/>
          <rect x="44" y="39" width="33" height="2" rx="1" fill="#374151"/>
          <rect x="44" y="44" width="20" height="2" rx="1" fill="#374151"/>
        </svg>""",
        "story_arc": """<svg viewBox="0 0 80 56" xmlns="http://www.w3.org/2000/svg">
          <rect width="80" height="56" rx="3" fill="#0B1220"/>
          <rect x="3" y="3" width="74" height="5" rx="1" fill="#1E293B"/>
          <rect x="3" y="11" width="74" height="10" rx="1" fill="#1E3A5F" opacity="0.7"/>
          <rect x="5" y="13" width="50" height="2" rx="1" fill="#93C5FD"/>
          <rect x="5" y="17" width="38" height="2" rx="1" fill="#93C5FD" opacity="0.5"/>
          <rect x="3" y="24" width="36" height="14" rx="1" fill="#14532D" opacity="0.8"/>
          <rect x="41" y="24" width="18" height="14" rx="1" fill="#1E3A5F" opacity="0.8"/>
          <rect x="61" y="24" width="16" height="14" rx="1" fill="#374151" opacity="0.8"/>
          <rect x="41" y="39" width="36" height="6" rx="1" fill="#1A3D2B" opacity="0.7"/>
          <rect x="3" y="41" width="36" height="4" rx="1" fill="#374151" opacity="0.5"/>
        </svg>""",
        "ops_dashboard": """<svg viewBox="0 0 80 56" xmlns="http://www.w3.org/2000/svg">
          <rect width="80" height="56" rx="3" fill="#0B1220"/>
          <rect x="3" y="3" width="74" height="5" rx="1" fill="#1E293B"/>
          <rect x="3" y="11" width="17" height="10" rx="2" fill="#14532D"/>
          <circle cx="8" cy="16" r="3" fill="#10B981"/>
          <rect x="22" y="11" width="17" height="10" rx="2" fill="#7C2D12"/>
          <circle cx="27" cy="16" r="3" fill="#EF4444"/>
          <rect x="41" y="11" width="17" height="10" rx="2" fill="#1E3A5F"/>
          <circle cx="46" cy="16" r="3" fill="#3B82F6"/>
          <rect x="60" y="11" width="17" height="10" rx="2" fill="#78350F"/>
          <circle cx="65" cy="16" r="3" fill="#F59E0B"/>
          <rect x="3" y="25" width="74" height="2" rx="1" fill="#1F2937"/>
          <rect x="3" y="28" width="10" height="5" rx="1" fill="#14532D" opacity="0.9"/>
          <rect x="15" y="28" width="10" height="5" rx="1" fill="#14532D" opacity="0.7"/>
          <rect x="27" y="28" width="10" height="5" rx="1" fill="#7C2D12" opacity="0.9"/>
          <rect x="39" y="28" width="10" height="5" rx="1" fill="#1E3A5F" opacity="0.8"/>
          <rect x="51" y="28" width="10" height="5" rx="1" fill="#14532D" opacity="0.6"/>
          <rect x="63" y="28" width="14" height="5" rx="1" fill="#78350F" opacity="0.8"/>
          <rect x="3" y="35" width="10" height="5" rx="1" fill="#1E3A5F" opacity="0.6"/>
          <rect x="15" y="35" width="10" height="5" rx="1" fill="#14532D" opacity="0.8"/>
          <rect x="27" y="35" width="10" height="5" rx="1" fill="#14532D" opacity="0.5"/>
          <rect x="39" y="35" width="10" height="5" rx="1" fill="#7C2D12" opacity="0.6"/>
          <rect x="51" y="35" width="10" height="5" rx="1" fill="#1E3A5F" opacity="0.9"/>
          <rect x="63" y="35" width="14" height="5" rx="1" fill="#14532D" opacity="0.7"/>
          <rect x="3" y="43" width="74" height="4" rx="1" fill="#1E293B" opacity="0.6"/>
        </svg>""",
        "board_pack": """<svg viewBox="0 0 80 56" xmlns="http://www.w3.org/2000/svg">
          <rect width="80" height="56" rx="3" fill="#0B1220"/>
          <rect x="0" y="0" width="80" height="10" rx="3" fill="#1E293B"/>
          <rect x="3" y="2" width="40" height="3" rx="1" fill="#93C5FD" opacity="0.8"/>
          <rect x="3" y="6" width="25" height="2" rx="1" fill="#6B7280"/>
          <rect x="3" y="14" width="17" height="10" rx="2" fill="#1E3A5F"/>
          <rect x="22" y="14" width="17" height="10" rx="2" fill="#1E3A5F"/>
          <rect x="41" y="14" width="17" height="10" rx="2" fill="#1E3A5F"/>
          <rect x="60" y="14" width="17" height="10" rx="2" fill="#1E3A5F"/>
          <circle cx="18" cy="38" r="10" fill="none" stroke="#1F2937" stroke-width="4"/>
          <circle cx="18" cy="38" r="10" fill="none" stroke="#3B82F6" stroke-width="4" stroke-dasharray="32 31" stroke-dashoffset="-6"/>
          <circle cx="18" cy="38" r="10" fill="none" stroke="#10B981" stroke-width="4" stroke-dasharray="20 43" stroke-dashoffset="-38"/>
          <rect x="32" y="28" width="45" height="3" rx="1" fill="#374151"/>
          <rect x="32" y="33" width="38" height="3" rx="1" fill="#374151"/>
          <rect x="32" y="40" width="45" height="8" rx="2" fill="#1E3A5F" opacity="0.6"/>
          <rect x="34" y="42" width="40" height="2" rx="1" fill="#93C5FD" opacity="0.5"/>
          <rect x="34" y="45" width="30" height="2" rx="1" fill="#6B7280" opacity="0.5"/>
        </svg>""",
    }

    tmpl_cols = st.columns(5, gap="small")
    for col, (tk, td) in zip(tmpl_cols, CARD_TEMPLATES.items()):
        is_sel  = st.session_state.card_template == tk
        col.markdown(
            f'<div style="border:2px solid #374151;border-radius:10px;'
            f'padding:8px 6px;text-align:center;background:#0F172A;margin-bottom:6px;">'
            f'{_TMPL_WIREFRAMES[tk]}'
            f'<div style="font-size:10px;font-weight:800;color:#E5E7EB;'
            f'margin-top:6px;line-height:1.2;">{td["label"]}</div>'
            f'<div style="font-size:8px;color:#6B7280;margin-top:3px;line-height:1.3;">'
            f'{td["tagline"]}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if col.button("✓ Selected" if is_sel else "Select",
                      key=f"tmpl_{tk}", use_container_width=True,
                      type="primary" if is_sel else "secondary"):
            st.session_state.card_template = tk
            _clear_preview()
            st.rerun()

    st.divider()

    # ────────────────────────────────────────────────────────────────────────── #
    # 4. PREVIEW CARD
    # ────────────────────────────────────────────────────────────────────────── #
    st.subheader("4. Your Card Preview")

    df       = st.session_state.get("df")
    date_col = st.session_state.get("date_col")
    kpis_cfg = [k for k in st.session_state.get("kpis", []) if k.get("enabled")]
    tf_key   = current_tf

    if "narrative" not in st.session_state or "metrics" not in st.session_state:
        if df is not None and date_col and kpis_cfg:
            with st.spinner("🧠 ARIA is calling Gemini to write your card — this matches what will be delivered to Slack…"):
                try:
                    metrics = compute_preview_metrics_v2(df, kpis_cfg, date_col, tf_key)
                    # Use live Gemini (with stub fallback) so the preview's
                    # headline, exec summary, recommended action, and speaker
                    # notes are EXACTLY what the scheduled agent will post.
                    narr_dict, narr_obj            = build_live_narrative(metrics, role_cfg)
                    st.session_state.metrics       = metrics
                    st.session_state.narrative     = narr_dict
                    st.session_state.narrative_obj = narr_obj
                    _u = st.session_state.get("user", {})
                    log_activity(_u.get("email",""), _u.get("name",""),
                                 "preview_card",
                                 f"Role={st.session_state.get('role_name','?')} "
                                 f"Template={st.session_state.get('card_template','?')}")
                except Exception as e:
                    st.error(f"Preview generation failed: {e}")
        else:
            st.info("No data loaded — showing a sample card. Go back to Step 2 to upload your file.", icon="ℹ️")
            _kpi_names = [k["user_name"] for k in kpis_cfg] if kpis_cfg else ["Revenue", "Profit", "Orders", "Margin"]
            _sample_kpis = {}
            _sample_vals = [12847, 3341, 47, 26.0]
            _sample_fmts = ["$12,847", "$3,341", "47", "26.0%"]
            for i, n in enumerate(_kpi_names[:4]):
                _sample_kpis[n] = {
                    "value": _sample_vals[i], "value_fmt": _sample_fmts[i],
                    "mom_pct": 0.078, "yoy_pct": 0.114,
                }
            sample_metrics = {
                "reference_date": str(date.today() - timedelta(1)),
                "kpis": _sample_kpis,
                "trend_series": [
                    8200,9100,7800,10200,9600,11400,8900,10800,12100,9700,
                    11200,10400,13100,11800,12400,10900,14200,13500,11700,12900,
                    10600,13800,12300,14500,11400,13200,12600,14800,13100,12847,
                ],
                "drivers": [
                    {"dimension":"Region","member":"Asia Pacific","delta":3240,"delta_pct":0.18},
                    {"dimension":"Region","member":"Europe",      "delta":1420,"delta_pct":0.09},
                    {"dimension":"Region","member":"N America",   "delta":-1180,"delta_pct":-0.11},
                ],
            }
            narr_dict, narr_obj = build_stub_narrative(sample_metrics, role_cfg)
            st.session_state.metrics       = sample_metrics
            st.session_state.narrative     = narr_dict
            st.session_state.narrative_obj = narr_obj

    if "narrative_obj" in st.session_state and "metrics" in st.session_state:
        import streamlit.components.v1 as components

        # ── PARITY BANNER ────────────────────────────────────────────────── #
        # Tells the user honestly whether this preview is exact-match-with-
        # delivery (live Gemini ran) or approximate (stub fallback fired).
        _narr_model = (st.session_state.narrative or {}).get("model", "")
        _is_live    = _narr_model and not _narr_model.startswith(("stub", "fallback"))
        if _is_live:
            st.markdown(
                '<div style="background:#10B98118;border:1px solid #10B98140;'
                'border-radius:10px;padding:10px 16px;margin-bottom:14px;'
                'font-size:13px;color:#A7F3D0;">'
                '✅ <b>Exact match with delivery</b> — same KPIs, same Gemini '
                'narrative, same template. This is what your Slack channel will receive.'
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="background:#F59E0B18;border:1px solid #F59E0B40;'
                'border-radius:10px;padding:10px 16px;margin-bottom:14px;'
                'font-size:13px;color:#FCD34D;">'
                '⚠️ <b>Approximate preview</b> — layout, KPIs and template match '
                'delivery, but narrative copy used a stub (Gemini key missing or '
                'call failed). The scheduled agent will generate real Gemini copy.'
                '</div>',
                unsafe_allow_html=True,
            )

        svg = generate_svg_preview(
            st.session_state.narrative_obj,
            st.session_state.metrics,
            role_cfg,
            st.session_state.card_style,
            st.session_state.card_template,
        )

        if svg:
            # svg is now a full self-contained HTML string from html_generator.
            # Initial height is 700; the postMessage resize script inside the
            # HTML adjusts it to the card's actual rendered height automatically.
            components.html(svg, height=700, scrolling=False)
        else:
            st.warning(
                "Card preview unavailable — check agent/html_generator.py is present.",
                icon="⚠️",
            )

        narr = st.session_state.narrative
        m    = st.session_state.metrics
        style_label    = CARD_STYLES[st.session_state.card_style]["label"]
        template_label = CARD_TEMPLATES[st.session_state.card_template]["label"]
        _atf = _tf_by_key(st.session_state.get("timeframe_key", "30d"))
        st.caption(
            f"Style: **{style_label}** · Template: **{template_label}** · "
            f"Role: **{role_cfg.get('title','—')}** · "
            f"Window: **{_atf['label'] if _atf else 'Custom'}** · "
            f"Engine: **{narr.get('model','stub')}** · "
            f"Drivers: **{len(m.get('drivers',[]))}**"
        )
        _r1, _r2 = st.columns([1, 5])
        if _r1.button("🔄 Refresh", use_container_width=True):
            _clear_preview()
            st.rerun()

    nav_buttons(back=True, next_label="Next: Set Delivery →")


# ════════════════════════════════════════════════════════════════════════════════
# STEP 8 — GO LIVE  (single form → one button → done)
# ════════════════════════════════════════════════════════════════════════════════

def _merge_roles_yaml(owner: str, repo: str, pat: str, new_roles_yaml: str) -> str:
    """
    Returns the new roles.yaml as-is, replacing whatever was in the repo.
    Each wizard run configures exactly one role — old accumulated roles are cleared
    so the agent never posts to stale channels from previous runs.
    """
    return new_roles_yaml


def build_runtime_role_cfg(role_name: str | None = None) -> dict:
    """Single source of truth for the role config used by BOTH the wizard
    preview (Step 6) AND the YAML written to roles.yaml at launch (Step 8).

    Pulls from session_state:
      - role_name        (Step 5 selection)
      - role_kpi_map     (Step 5 LLM-assigned KPI list per role)
      - kpis             (Step 4 enabled KPI list)
      - card_template    (Step 6 selection)
      - card_style       (Step 6 selection)
      - slack_channel    (Step 7 input)

    Merges them onto the static ROLE_GROUPS entry so callers always get a
    COMPLETE role_cfg in one shape: title, badge, primary_kpi, kpis,
    accent_color, driver_focus, tone, card_template, card_style,
    slack_channel.

    This function is what guarantees the preview and the delivered card
    use identical inputs — they both call it.
    """
    role_name = role_name or st.session_state.get("role_name", "CEO")

    # 1. Static base from ROLE_GROUPS (title, badge, accent, driver_focus, tone)
    base_cfg = None
    for grp in ROLE_GROUPS.values():
        if role_name in grp:
            base_cfg = grp[role_name]
            break
    if base_cfg is None:
        base_cfg = ROLE_GROUPS["🏛️ C-Suite"]["CEO"]

    # 2. Apply Step 5 LLM KPI assignment — overrides the hardcoded role.kpis
    _role_map     = st.session_state.get("role_kpi_map", {}) or {}
    _kpis_cfg     = [k for k in st.session_state.get("kpis", []) if k.get("enabled")]
    llm_kpi_names = [k["user_name"] for k in _kpis_cfg] if _kpis_cfg else []
    role_kpis     = (_role_map.get(role_name)
                     or llm_kpi_names[:6]
                     or base_cfg.get("kpis", ["Sales"]))
    primary_kpi   = role_kpis[0] if role_kpis else base_cfg.get("primary_kpi", "Sales")

    return {
        "title":         base_cfg.get("title", role_name),
        "badge":         base_cfg.get("badge", f"{role_name.upper()}  ·  BRIEFING"),
        "slack_channel": st.session_state.get("slack_channel", ""),
        "primary_kpi":   primary_kpi,
        "kpis":          role_kpis,
        "accent_color":  base_cfg.get("accent_color", "#F59E0B"),
        "driver_focus":  base_cfg.get("driver_focus", ["Category", "Region"]),
        "tone":          base_cfg.get("tone", "Executive and strategic."),
        "card_template": st.session_state.get("card_template", "editorial"),
        "card_style":    st.session_state.get("card_style", "dark"),
    }


def _build_configs() -> tuple[str, str]:
    """Build config.yaml and roles.yaml strings from session state."""
    channels      = [k for k, v in st.session_state.get("delivery", {}).items() if v]
    kpis_cfg      = [k for k in st.session_state.get("kpis", []) if k.get("enabled")]
    provider      = st.session_state.get("ai_provider", "stub")
    tz            = st.session_state.get("timezone", "Asia/Kolkata")
    gs_url        = st.session_state.get("google_sheet_url", "")
    data_source   = st.session_state.get("data_source", "file")
    data_name     = st.session_state.get("data_name", "Superstore.xls")
    df            = st.session_state.get("df")
    role_cfg      = _resolve_role()
    role_name     = st.session_state.get("role_name", "CEO")
    slack_channel = st.session_state.get("slack_channel", "")
    style_key     = st.session_state.get("card_style", "dark")
    template_key  = st.session_state.get("card_template", "editorial")
    eff_accent    = role_cfg.get("accent_color", "#F59E0B")

    # ── Resolve data source ──────────────────────────────────────────────── #
    # Priority order:
    #   1. uploaded_file_bytes present → ALWAYS use file, never use Google Sheet URL
    #   2. data_source == "google_sheets" and no uploaded bytes → use Google Sheet
    #   3. Fallback — empty config
    date_col        = st.session_state.get("date_col", "Order Date")
    _uploaded_bytes = st.session_state.get("uploaded_file_bytes")

    if _uploaded_bytes and data_name:
        # File was uploaded — this overrides data_source regardless of what it says.
        # Covers the case where the user clicked "Use Superstore sample" first and
        # then uploaded their own file in the same session.
        safe_name  = re.sub(r"[^\w.\-]", "_", data_name)
        cfg_gs_url = ""
        cfg_excel  = f"data/{safe_name}"
        cfg_sheet  = st.session_state.get("excel_sheet_name", "Sheet1")

    elif data_source == "google_sheets" and gs_url and not _uploaded_bytes:
        # Pure Google Sheets path — no local file uploaded
        cfg_gs_url = gs_url
        cfg_excel  = ""
        cfg_sheet  = "Sheet1"

    else:
        # Fallback — safe empty config
        cfg_gs_url = ""
        _safe_fb   = re.sub(r"[^\w.\-]", "_", data_name) if data_name else ""
        cfg_excel  = f"data/{_safe_fb}" if _safe_fb else ""
        cfg_sheet  = st.session_state.get("excel_sheet_name", "Sheet1")

    # ── Auto-detect dimension columns from the actual dataframe ─────────── #
    if df is not None:
        kpi_cols = {k.get("column") for k in kpis_cfg}
        kpi_cols |= {k.get("num_col") for k in kpis_cfg if k.get("num_col")}
        kpi_cols |= {k.get("den_col") for k in kpis_cfg if k.get("den_col")}
        date_like = {"date", "time", "year", "month", "quarter", "week", "day"}
        dim_cols  = [
            c for c in df.columns
            if c not in kpi_cols
            and c != date_col
            and df[c].dtype == object
            and not any(d in c.lower() for d in date_like)
            and df[c].nunique() < 50
        ][:6]  # cap at 6 dimensions
    else:
        dim_cols = ["Category", "Region", "Segment"]

    config = {
        "data": {
            "google_sheet_url": cfg_gs_url,
            "excel_path":       cfg_excel,
            "sheet_name":       cfg_sheet,
            "date_column":      date_col,
            "timezone":         tz,
            "fallback_to_max_date_if_missing": True,
        },
        "metrics": {
            "timeframe": (
                st.session_state.get("timeframe_key", "1d")
                if st.session_state.get("timeframe_key") not in (None, "custom")
                else "alltime"   # custom date range → treat as full history
            ),
            "kpis": [
                {"name": k["user_name"], "column": k.get("column", k["user_name"]),
                 "agg": k["agg"], "format": k["format"]}
                for k in kpis_cfg
            ],
            "derived": [],
            "anomaly_zscore_threshold": 2.0,
            "anomaly_lookback_days":    90,
        },
        "drivers": {
            "dimensions": dim_cols,
            "top_n": 3,
        },
        "llm": {
            "provider": provider, "model": "gemini-2.5-flash",
            "temperature": 0.4, "max_output_tokens": 8192,
        },
        "delivery": {
            "channels": channels,
            "slack":  {"webhook_env_var": "SLACK_WEBHOOK_URL"},
            "teams":  {"webhook_env_var": "TEAMS_WEBHOOK_URL",
                       "title_prefix": "Daily Performance Briefing"},
            "email":  {"recipient_env_var": "EMAIL_RECIPIENT",
                       "sender_env_var": "EMAIL_SENDER",
                       "app_password_env_var": "EMAIL_APP_PASSWORD",
                       "subject_prefix": "ARIA Daily Briefing"},
            "file":   {"output_dir": "output", "formats": ["markdown", "docx"]},
            "schedule_type": st.session_state.get("schedule_type", "daily"),
        },
    }

    # Build the role entry via the shared helper so this matches what the
    # preview renders. Single source of truth — no duplicated logic.
    roles = {"roles": {role_name: build_runtime_role_cfg(role_name)}}

    return (
        yaml.dump(config, default_flow_style=False, allow_unicode=True, sort_keys=False),
        yaml.dump(roles,  default_flow_style=False, allow_unicode=True, sort_keys=False),
    )


def _render_key_inputs(provider: str, channels: list):
    """Render API key input fields. Called from Step 8 for missing or editable keys."""
    if provider == "gemini":
        g1, g2 = st.columns([5, 1])
        gemini_key = g1.text_input(
            "Gemini API key",
            type="password",
            placeholder="Paste your Gemini API key here",
            key="gemini_key_launch",
            value=st.session_state.get("gemini_key", ""),
        )
        with g2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("gemini_key", "ℹ️")
        if gemini_key:
            st.session_state.gemini_key = gemini_key

    if "slack" in channels:
        s1, s2 = st.columns([5, 1])
        slack_tok = s1.text_input(
            "Slack Bot Token",
            type="password",
            placeholder="xoxb-...",
            key="slack_tok_launch",
            value=st.session_state.get("slack_token", ""),
        )
        with s2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("slack_bot_token", "ℹ️")
        c1, c2 = st.columns([5, 1])
        slack_ch = c1.text_input(
            "Slack Channel ID",
            placeholder="e.g. C012AB3CD45",
            key="slack_ch_launch",
            value=st.session_state.get("slack_channel", ""),
        )
        with c2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("slack_channel_id", "ℹ️")
        if slack_tok:
            st.session_state.slack_token   = slack_tok
        if slack_ch:
            st.session_state["slack_channel"] = slack_ch

    if "teams" in channels:
        t1, t2 = st.columns([5, 1])
        teams_wh = t1.text_input(
            "Microsoft Teams Webhook URL",
            type="password",
            placeholder="https://your-org.webhook.office.com/...",
            key="teams_wh_launch",
            value=st.session_state.get("teams_webhook", ""),
        )
        with t2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("teams_webhook", "ℹ️")
        if teams_wh:
            st.session_state.teams_webhook = teams_wh

    if "email" in channels:
        e1, e2 = st.columns([5, 1])
        email_sender = e1.text_input(
            "Gmail sender address (e.g. aria.briefings@gmail.com)",
            placeholder="aria.briefings@gmail.com",
            key="email_sender_launch",
            value=st.session_state.get("email_sender", ""),
        )
        with e2:
            st.markdown("<br>", unsafe_allow_html=True)
            help_tip("email_delivery", "ℹ️")
        email_pw = st.text_input(
            "Gmail App Password (16-char, spaces optional)",
            type="password",
            placeholder="xxxx xxxx xxxx xxxx",
            key="email_pw_launch",
            value=st.session_state.get("email_app_password", ""),
        )
        if email_sender:
            st.session_state.email_sender = email_sender
        if email_pw:
            st.session_state.email_app_password = email_pw


def step_export_go():
    st.header("🚀 Go Live")

    config_yaml, roles_yaml = _build_configs()
    channels  = [k for k, v in st.session_state.get("delivery", {}).items() if v]
    provider  = st.session_state.get("ai_provider", "stub")
    tz        = st.session_state.get("timezone", "Asia/Kolkata")
    del_hour  = st.session_state.get("delivery_hour", 9)
    role_name = st.session_state.get("role_name", "CEO")
    role_cfg  = _resolve_role()
    kpis_cfg  = [k for k in st.session_state.get("kpis", []) if k.get("enabled")]
    style_key    = st.session_state.get("card_style", "dark")
    template_key = st.session_state.get("card_template", "editorial")
    eff_accent   = role_cfg.get("accent_color", "#F59E0B")

    # ═══════════════════════════════════════════════════════════════════════════
    # SCREEN A — LAUNCH PAD  (the only screen the user needs to fill in)
    # ═══════════════════════════════════════════════════════════════════════════
    if "launch_done" not in st.session_state:

        # ── Config summary ────────────────────────────────────────────────── #
        st.markdown(
            f"Your card is ready. One last thing — paste your tokens below and "
            f"ARIA will handle **everything else** automatically."
        )
        style_lbl    = CARD_STYLES.get(style_key, {}).get("label", style_key)
        template_lbl = CARD_TEMPLATES.get(template_key, {}).get("label", template_key)
        st.markdown(
            f'<div style="background:#F59E0B18;border:1px solid #F59E0B40;border-radius:10px;'
            f'padding:14px 20px;margin-bottom:16px;font-size:13px;line-height:2">'
            f'📊 <b>{len(kpis_cfg)} KPIs</b> &nbsp;·&nbsp; '
            f'👤 <b>{role_name}</b> &nbsp;·&nbsp; '
            f'🎨 <b>{style_lbl} · {template_lbl}</b> &nbsp;·&nbsp; '
            f'⏰ <b>{del_hour}:00 {tz}</b> &nbsp;·&nbsp; '
            f'📬 <b>{", ".join(channels) or "File"}</b>'
            f'</div>',
            unsafe_allow_html=True,
        )

        st.divider()

        # ── SECTION 1: GitHub PAT ─────────────────────────────────────────── #
        st.markdown("### 1️⃣ &nbsp; GitHub — where ARIA runs every morning")

        st.markdown(
            "GitHub Actions is a **free** cloud scheduler. ARIA will create your "
            "repo automatically — you just need to give it a token once.",
            help="GitHub Actions gives you 2,000 free minutes/month. ARIA uses ~2 min/day.",
        )

        # No account yet?
        no_gh = st.checkbox("I don't have a GitHub account yet", key="no_github")
        if no_gh:
            st.info(
                "**Create your free account (2 min):**\n\n"
                "👉 [github.com/signup](https://github.com/signup) — enter email, "
                "pick a username, verify email. Then come back here.",
                icon="👤",
            )

        st.markdown(
            "**Personal Access Token** &nbsp;"
            "<span style='font-size:12px;color:#9CA3AF'>— like a one-time password for ARIA</span>",
            unsafe_allow_html=True,
        )

        # Inline guide — collapsed by default so it's not overwhelming
        with st.expander("📋 How to get your token in 60 seconds"):
            st.markdown(
                "1. Click this link → **[github.com/settings/tokens/new](https://github.com/settings/tokens/new)**\n"
                "   *(opens GitHub's token page directly)*\n\n"
                "2. Give it any name, e.g. **ARIA**\n\n"
                "3. Tick **`repo`** ✅ and **`workflow`** ✅\n\n"
                "4. Scroll down → click **Generate token**\n\n"
                "5. Copy the token (starts with `ghp_`) and paste below\n\n"
                "⚠️ The token is shown **only once** — copy it before closing."
            )
            st.warning(
                "Open the link above in a **new tab**, keep this wizard open.",
                icon="💡",
            )

        pat_input = st.text_input(
            "Paste token here",
            type="password",
            placeholder="ghp_xxxxxxxxxxxxxxxxxxxx",
            key="gh_pat_input",
            label_visibility="collapsed",
        )

        # Validate token live — username auto-detected from the token itself
        pat_ok = False
        if pat_input:
            r = _gh_api("get", "/user", pat_input)
            if r.status_code == 200:
                actual_user = r.json().get("login", "")
                st.success(f"✅ Token valid — connected as **{actual_user}**")
                st.session_state.gh_pat_valid = pat_input
                st.session_state.gh_owner     = actual_user
                st.session_state.gh_repo      = "aria-daily-briefing"
                pat_ok = True
            elif r.status_code == 401:
                st.error("Token not recognised — double-check you copied it fully.")
                st.session_state.pop("gh_pat_valid", None)
                st.session_state.pop("gh_owner", None)
            else:
                st.caption(f"Couldn't reach GitHub right now (HTTP {r.status_code}). Try again in a moment.")

        st.divider()

        # ── LAUNCH button ─────────────────────────────────────────────────── #
        ready = bool(st.session_state.get("gh_pat_valid"))
        st.button(
            "🚀  Launch ARIA — set everything up automatically",
            type="primary",
            use_container_width=True,
            disabled=not ready,
            key="launch_btn",
            on_click=lambda: st.session_state.update({"launch_triggered": True}),
        )
        if not ready:
            st.caption("👆 Paste your GitHub token above to unlock.")

        # Manual fallback (collapsed, out of the way)
        with st.expander("⬇️ Prefer to set up manually? Download config files"):
            dl1, dl2 = st.columns(2)
            dl1.download_button("config.yaml", config_yaml, "config.yaml", "text/plain", use_container_width=True)
            dl2.download_button("roles.yaml",  roles_yaml,  "roles.yaml",  "text/plain", use_container_width=True)

        # ── Bottom nav — Back (left) + Reset All (right) ─────────────────── #
        st.divider()
        _nav_a_l, _nav_a_mid, _nav_a_r = st.columns([1, 2, 2])
        if _nav_a_l.button("← Back", use_container_width=True, key="s8a_back"):
            st.session_state.step -= 1
            st.rerun()
        if _nav_a_r.button("🔄 Reset All Selections", use_container_width=True,
                            type="secondary", key="s8a_reset"):
            _u = st.session_state.get("user")
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            if _u:
                st.session_state.user = _u
            st.rerun()

        # ── Execute launch when button clicked ───────────────────────────── #
        if st.session_state.get("launch_triggered"):
            st.session_state.pop("launch_triggered", None)
            owner  = st.session_state.get("gh_owner", "")
            repo   = st.session_state.get("gh_repo",  "aria-daily-briefing")
            pat    = st.session_state.get("gh_pat_valid", "")
            results: dict[str, tuple[bool, str]] = {}

            progress = st.progress(0, text="Starting…")

            # 1. Create repo if it doesn't exist
            progress.progress(10, text="Creating your ARIA repo…")
            r_check = _gh_api("get", f"/repos/{owner}/{repo}", pat)
            if r_check.status_code == 404:
                r_create = _gh_api("post", "/user/repos", pat, data={
                    "name":        repo,
                    "description": "ARIA — Autonomous Report & Insight AI Agent",
                    "private":     True,
                    "auto_init":   True,
                })
                results["Create repo"] = (
                    r_create.status_code in (200, 201),
                    "created" if r_create.status_code in (200, 201) else f"HTTP {r_create.status_code}",
                )
                import time as _time; _time.sleep(2)  # let GitHub finish init
            else:
                results["Create repo"] = (True, "already exists")

            # 2. Push config files
            progress.progress(20, text="Pushing config files…")
            # Push config.yaml as-is (one config per repo is fine)
            ok, m = _gh_push_file(owner, repo, pat, "config.yaml", config_yaml,
                                   "chore: ARIA wizard config")
            results["config.yaml"] = (ok, m)

            # Replace roles.yaml entirely — only generate a card for the role
            # the user selected in Step 4. Merging old roles caused all previously
            # configured roles to keep running even if the user didn't select them.
            ok, m = _gh_push_file(owner, repo, pat, "roles.yaml", roles_yaml,
                                   f"chore: ARIA wizard — set role to {role_name}")
            results["roles.yaml"] = (ok, m)

            # 2b. Push uploaded data file — trigger on uploaded_file_bytes existing,
            # NOT on data_source, because the user may have started with "Superstore"
            # and then uploaded their own file (data_source stays "google_sheets").
            _data_name   = st.session_state.get("data_name", "")
            _uploaded_file_obj = st.session_state.get("uploaded_file_bytes")
            if _data_name and _uploaded_file_obj:
                progress.progress(28, text=f"Uploading your data file ({_data_name})…")
                _safe_name    = re.sub(r"[^\w.\-]", "_", _data_name)
                _repo_path    = f"data/{_safe_name}"
                # _uploaded_file_bytes is stored as bytes in session state
                _file_b64     = base64.b64encode(_uploaded_file_obj).decode("utf-8")
                # Use the GitHub Contents API directly (binary-safe base64)
                _hdr = {
                    "Authorization": f"token {pat}",
                    "Accept": "application/vnd.github.v3+json",
                    "X-GitHub-Api-Version": "2022-11-28",
                }
                _chk = _req.get(
                    f"https://api.github.com/repos/{owner}/{repo}/contents/{_repo_path}",
                    headers=_hdr, timeout=20,
                )
                _body = {
                    "message": f"data: add {_safe_name}",
                    "content": _file_b64,
                }
                if _chk.status_code == 200:
                    _body["sha"] = _chk.json().get("sha", "")
                _push_r = _req.put(
                    f"https://api.github.com/repos/{owner}/{repo}/contents/{_repo_path}",
                    headers=_hdr, json=_body, timeout=60,
                )
                results[f"📁 data/{_safe_name}"] = (
                    _push_r.status_code in (200, 201),
                    "uploaded" if _push_r.status_code in (200, 201)
                    else f"HTTP {_push_r.status_code}: {_push_r.text[:150]}",
                )

            # 2c. Push requirements.txt and all agent code
            progress.progress(35, text="Uploading agent code…")
            _agent_files = [
                "requirements.txt",
                "agent/__init__.py",
                "agent/main.py",
                "agent/narrative_generator.py",
                "agent/svg_generator.py",
                "agent/metrics_engine.py",
                "agent/data_loader.py",
                "agent/driver_analysis.py",
                "agent/slack_publisher.py",
                "agent/report_writer.py",
                "agent/teams_publisher.py",
            ]
            for _rel in _agent_files:
                _local = _ROOT / _rel
                if _local.exists():
                    try:
                        _content = _local.read_text(encoding="utf-8")
                    except Exception:
                        _content = _local.read_bytes().decode("latin-1")
                    ok, m = _gh_push_file(owner, repo, pat, _rel, _content,
                                          f"chore: ARIA agent — {_rel}")
                    results[f"📄 {_rel}"] = (ok, m)
                else:
                    results[f"📄 {_rel}"] = (False, "not found locally")

            # 3. Push workflow
            progress.progress(60, text="Setting up daily schedule…")
            _one_time = st.session_state.get("schedule_type", "daily") == "onetime"
            workflow_yaml = _generate_workflow_yaml(del_hour, tz, provider, channels, one_time=_one_time)
            ok, m = _gh_push_file(owner, repo, pat,
                                   ".github/workflows/aria-daily.yml",
                                   workflow_yaml, "chore: ARIA daily workflow")
            results["workflow"] = (ok, m)

            # 4. Set secrets
            progress.progress(80, text="Encrypting and storing your API keys…")
            secrets_to_set: dict[str, str] = {}

            # Google credentials — try st.secrets first (Streamlit Cloud),
            # then fall back to local file (local dev). Not needed for public sheets
            # (data_loader uses CSV export), but set it if available for future use.
            import json as _json

            def _to_str(v) -> str:
                """Ensure a secrets value is a plain string.
                st.secrets.get() returns AttrDict for TOML-section secrets, not str."""
                if v is None:
                    return ""
                if isinstance(v, str):
                    return v
                if hasattr(v, "keys"):
                    # TOML-section AttrDict → serialize back to JSON string
                    return _json.dumps({k: str(val) for k, val in v.items()})
                return str(v)

            _gcreds_val = None
            try:
                _raw = st.secrets.get("GOOGLE_CREDS_JSON")
                if _raw is not None:
                    _gcreds_val = _to_str(_raw)
            except Exception:
                pass
            if not _gcreds_val:
                _gcreds_path = _ROOT / "google_creds.json"
                if _gcreds_path.exists():
                    try:
                        _gcreds_val = _gcreds_path.read_text(encoding="utf-8")
                    except Exception:
                        pass
            if _gcreds_val:
                secrets_to_set["GOOGLE_CREDS_JSON"] = _gcreds_val

            if provider == "gemini" and st.session_state.get("gemini_key"):
                secrets_to_set["GEMINI_API_KEY"] = _to_str(st.session_state.gemini_key)
            if "slack" in channels:
                if st.session_state.get("slack_token"):
                    secrets_to_set["SLACK_BOT_TOKEN"] = _to_str(st.session_state.slack_token)
                if st.session_state.get("slack_channel"):
                    secrets_to_set["SLACK_CHANNEL_ID"] = _to_str(st.session_state.slack_channel)
            if "teams" in channels and st.session_state.get("teams_webhook"):
                secrets_to_set["TEAMS_WEBHOOK_URL"] = _to_str(st.session_state.teams_webhook)
            if "email" in channels:
                if st.session_state.get("email_sender"):
                    secrets_to_set["EMAIL_SENDER"] = _to_str(st.session_state.email_sender)
                if st.session_state.get("email_app_password"):
                    secrets_to_set["EMAIL_APP_PASSWORD"] = _to_str(st.session_state.email_app_password)
                if st.session_state.get("email_recipient"):
                    secrets_to_set["EMAIL_RECIPIENT"] = _to_str(st.session_state.email_recipient)

            for sname, sval in secrets_to_set.items():
                ok, m = _gh_set_secret(owner, repo, pat, sname, sval)
                results[f"🔑 {sname}"] = (ok, m)

            progress.progress(100, text="Done!")

            st.session_state.setup_results = results
            st.session_state.launch_done   = True
            st.session_state.gh_repo_final = repo

            # ── Log to ARIA_Activity + ARIA_Schedules ──────────────────── #
            _u    = st.session_state.get("user", {})
            _ueml = _u.get("email", "anonymous")
            _unm  = _u.get("name", "Anonymous")
            _sched_type = st.session_state.get("schedule_type", "daily")
            _chs  = ", ".join(channels) or "file"
            _rurl = f"https://github.com/{owner}/{repo}"
            log_activity(_ueml, _unm, "wizard_completed",
                         f"Role={role_name} | Channels={_chs} | Repo={_rurl}")
            if _sched_type == "onetime":
                log_activity(_ueml, _unm, "one_time_card",
                             f"One-time card for {role_name}")
            log_schedule(_ueml, _unm, _rurl, owner,
                         role_name, _sched_type, _chs, tz, del_hour)

            st.rerun()

    # ═══════════════════════════════════════════════════════════════════════════
    # SCREEN B — RESULT  (auto-shown after launch)
    # ═══════════════════════════════════════════════════════════════════════════
    else:
        owner    = st.session_state.get("gh_owner", "—")
        repo     = st.session_state.get("gh_repo_final", "aria-daily-briefing")
        results  = st.session_state.get("setup_results", {})
        all_ok   = all(v[0] for v in results.values())

        if all_ok:
            st.markdown("""
            <div style="text-align:center;padding:24px 0 12px">
              <div style="font-size:52px">🎉</div>
              <h2 style="margin:8px 0 4px">ARIA is live!</h2>
              <p style="color:#9CA3AF;margin:0">Everything was set up automatically.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.warning("Setup completed with some issues — see details below.", icon="⚠️")

        # Result detail (collapsed by default if all OK)
        with st.expander("Setup log", expanded=not all_ok):
            for item, (ok, msg) in results.items():
                st.markdown(f"{'✅' if ok else '❌'} **{item}** — {msg if not ok else 'done'}")

        if not all_ok:
            failed_secrets = [
                k.replace("🔑 ", "") for k, (ok, _) in results.items()
                if not ok and "🔑" in k
            ]
            if failed_secrets:
                st.info(
                    "The config files were pushed but some secrets need to be added manually.\n\n"
                    "Go to: **your repo → Settings → Secrets and variables → Actions → New repository secret**\n\n"
                    + "\n".join(f"- `{s}`" for s in failed_secrets),
                    icon="🔑",
                )

        st.divider()

        repo_url = f"https://github.com/{owner}/{repo}"

        # Determine schedule label
        _is_onetime = st.session_state.get("schedule_type", "daily") == "onetime"
        if _is_onetime:
            _delivery_label = "One-time report · trigger manually from GitHub Actions"
        else:
            try:
                import pytz as _pytz, datetime as _dtt
                _now_local = _dtt.datetime.now(_pytz.timezone(tz))
                _delivery_day = "Today" if _now_local.hour < del_hour else "Tomorrow"
            except Exception:
                _delivery_day = "Tomorrow"
            _delivery_label = f"{_delivery_day} at **{del_hour}:00 {tz}**"

        st.markdown(f"""
| | |
|---|---|
| 📦 Repository | [{owner}/{repo}]({repo_url}) |
| 👤 Role | {role_name} — {role_cfg['title']} |
| ⏰ Schedule | {_delivery_label} |
| 📬 Delivery | {', '.join(channels) or 'File only'} |
        """)

        # ── Test now block ────────────────────────────────────────────────── #
        st.markdown("#### ▶️ Test it right now")
        tw_col, tip_col = st.columns([8, 1])
        tw_col.info(
            f"Go to **[{owner}/{repo} → Actions]({repo_url}/actions)** and trigger "
            f"the **ARIA Daily Briefing** workflow manually.",
            icon="▶️",
        )
        with tip_col:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.popover("ℹ️", use_container_width=False):
                st.markdown("**How to run the workflow — step by step**")
                st.markdown(
                    f"1. Click this link → **[{owner}/{repo} → Actions]({repo_url}/actions)**\n\n"
                    "2. In the left sidebar, click **ARIA Daily Briefing**\n\n"
                    "3. On the right side you'll see a grey **Run workflow** button — click it\n\n"
                    "4. A small dropdown appears — leave the branch as `main` and click the "
                    "green **Run workflow** button inside the dropdown\n\n"
                    "5. Refresh the page after a few seconds — you'll see a new run appear "
                    "with a spinning yellow circle ⏳ (in progress) or green ✓ (done)\n\n"
                    "6. Click the run row to open it and watch each step complete in real time"
                )

        # ── Queue delay notice ────────────────────────────────────────────── #
        with st.expander("⏱️ When will my card arrive? — delivery timing & how to check status"):
            st.markdown(
                "**GitHub Actions free tier runs jobs from a shared queue.** "
                "There is usually a short wait before your job starts.\n\n"
            )
            dc1, dc2 = st.columns(2)
            dc1.markdown(
                "**📥 Fastest: Download file (Artifacts)**\n\n"
                "Go to your Actions run → click on **Artifacts** available at the top of the "
                "summary page right next to Status & Total duration → at the bottom of the page "
                "click **aria-reports** under Artifacts to download the Markdown + Word report. "
                "No queue delay for reading — the file is ready the moment the green ✓ appears."
            )
            dc2.markdown(
                "**💬 Slack / 🟦 Teams / 📧 Email**\n\n"
                "The card posts/sends at the end of the run. "
                "Expect **2 – 5 minutes** total (queue wait + job runtime). "
                "If nothing arrives after 10 minutes, check the Actions log for errors "
                "(red ✗ next to the run row)."
            )
            st.divider()
            st.markdown("**🔍 How to tell if the run succeeded or failed**")
            st.markdown(
                f"1. Go to **[{owner}/{repo} → Actions]({repo_url}/actions)**\n\n"
                "2. Look at the latest run row:\n"
                "   - 🟡 **Yellow spinning circle** — job is still running, wait a moment\n"
                "   - ✅ **Green check** — run completed successfully, check your channel/inbox\n"
                "   - ❌ **Red X** — something went wrong\n\n"
                "3. If you see a **red ✗**, click the run → click the failed step (shown in red) "
                "→ expand the log to read the error message. Common causes:\n"
                "   - Missing GitHub Secret (e.g. `SLACK_BOT_TOKEN` not set)\n"
                "   - Google Sheet not shared publicly (needs *Anyone with link → Viewer*)\n"
                "   - Gemini API key quota exceeded (switch to Built-in Engine to test)\n\n"
                "4. After fixing, click **Re-run all jobs** (top-right of the run page) to retry."
            )

        # ── Bottom nav — Back (left) + Reset All (right) ─────────────────── #
        st.divider()
        _nav_b_l, _nav_b_mid, _nav_b_r = st.columns([1, 2, 2])
        if _nav_b_l.button("← Back", use_container_width=True, key="s8b_back"):
            # Clear launch state so user returns to Screen A (token / launch pad)
            _launch_keys = ["launch_done", "setup_results", "launch_triggered",
                            "gh_pat_valid", "gh_owner", "gh_repo", "gh_repo_final",
                            "gh_username_input", "gh_pat_input"]
            for k in _launch_keys:
                st.session_state.pop(k, None)
            st.rerun()
        if _nav_b_r.button("🔄 Reset All Selections", use_container_width=True,
                            type="secondary", key="s8b_reset"):
            _u = st.session_state.get("user")
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            if _u:
                st.session_state.user = _u
            st.rerun()


# ════════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════════

def main():
    st.markdown("""
    <style>
      section.main > div { max-width: 860px; margin: 0 auto; }
      .stButton > button  { border-radius: 8px; font-weight: 600; }

      /* ── Green primary buttons (Next, Launch, etc.) ── */
      .stButton > button[kind="primary"] {
          background-color: #10B981 !important;
          border-color:     #059669 !important;
          color:            #ffffff !important;
      }
      .stButton > button[kind="primary"]:hover {
          background-color: #059669 !important;
          border-color:     #047857 !important;
      }
      .stButton > button[kind="primary"]:disabled {
          background-color: #374151 !important;
          border-color:     #374151 !important;
          color:            #6B7280 !important;
      }

      /* ── Sleek info + help popover buttons ── */
      [data-testid="stPopover"] > button {
          width: 28px !important;
          height: 28px !important;
          min-width: 28px !important;
          padding: 0 !important;
          border-radius: 50% !important;
          border: 1.5px solid #374151 !important;
          background: transparent !important;
          color: #9CA3AF !important;
          font-size: 13px !important;
          display: flex !important;
          align-items: center !important;
          justify-content: center !important;
      }
      [data-testid="stPopover"] > button:hover {
          border-color: #10B981 !important;
          color: #10B981 !important;
          background: #10B98112 !important;
      }
      [data-testid="stPopover"] > button svg { display: none !important; }

      /* ── Sidebar account panel — wider to avoid text truncation ── */
      [data-testid="stSidebar"] { min-width: 270px !important; max-width: 320px !important; }

      /* ── Hide sidebar scrollbar entirely (scroll still works on drag/touch) ── */
      [data-testid="stSidebar"] > div:first-child {
          overflow-x: hidden !important;
          scrollbar-width: none !important;       /* Firefox */
          -ms-overflow-style: none !important;    /* IE/Edge */
      }
      [data-testid="stSidebar"] > div:first-child::-webkit-scrollbar {
          display: none !important;               /* Chrome/Safari */
      }

      /* ── Tighten sidebar element gaps to fit help expander without scroll ── */
      [data-testid="stSidebar"] .stButton > button {
          padding-top: 4px !important;
          padding-bottom: 4px !important;
          min-height: 34px !important;
      }
      [data-testid="stSidebar"] hr {
          margin: 6px 0 !important;
      }
      [data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {
          gap: 4px !important;
      }

      /* ── Time period chips — prevent text wrapping ── */
      button[data-testid="baseButton-secondary"] p,
      button[data-testid="baseButton-primary"] p {
          white-space: nowrap !important;
          overflow: visible !important;
      }
      button[data-testid="baseButton-secondary"],
      button[data-testid="baseButton-primary"] {
          overflow: visible !important;
      }
    </style>
    """, unsafe_allow_html=True)

    # ── Auth gate ── show login screen if not logged in ──────────────────────
    if "user" not in st.session_state:
        screen_auth()
        return

    # ── Logged in — render account sidebar (left) ────────────────────────────
    render_account_sidebar()

    # ── Bootstrap API keys from Streamlit Cloud secrets ──────────────────────
    # NOTE: slack_channel is intentionally NOT bootstrapped from secrets.
    # It is a per-role setting that must always be entered explicitly in Step 7.
    # Bootstrapping it would silently pre-fill an old/wrong channel.
    try:
        _sec = st.secrets
        _sec_map = {
            "gemini_key":    _sec.get("GEMINI_API_KEY"),
            "slack_token":   _sec.get("SLACK_BOT_TOKEN"),
            "teams_webhook": _sec.get("TEAMS_WEBHOOK_URL"),
        }
        for _k, _v in _sec_map.items():
            if _v and not st.session_state.get(_k):
                st.session_state[_k] = _v
    except Exception:
        pass

    # ── Screen router ────────────────────────────────────────────────────────
    screen = st.session_state.get("aria_screen", "wizard")

    if screen == "dashboard":
        screen_dashboard()
        return

    if screen == "admin":
        user = st.session_state.get("user", {})
        if user.get("is_admin"):
            screen_admin()
        else:
            st.error("Admin access only.")
        return

    # ── Wizard flow ──────────────────────────────────────────────────────────
    if "step" not in st.session_state:
        st.session_state.step = 1

    # Default to wizard screen
    if screen != "wizard":
        st.session_state.aria_screen = "wizard"

    render_progress()

    # Wrap every step in a single st.empty() placeholder so Streamlit atomically
    # replaces the entire step area when navigating — prevents ghost widgets from
    # the previous step briefly appearing during the next step's first render.
    _step_ph = st.empty()
    with _step_ph.container():
        step = st.session_state.step
        if   step == 1: step_welcome()
        elif step == 2: step_upload_data()
        elif step == 3: step_choose_ai()
        elif step == 4: step_discover_kpis()
        elif step == 5: step_pick_role()
        elif step == 6: step_preview_card()
        elif step == 7: step_set_delivery()
        elif step == 8: step_export_go()
        else:
            st.session_state.step = 1
            st.rerun()

    # ── Subtle attribution ───────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(
            '<div style="position:fixed;bottom:12px;left:0;width:270px;'
            'text-align:center;pointer-events:none;">'
            '<span style="font-size:12px;letter-spacing:1px;'
            'font-family:monospace;color:#9CA3AF">© Mona Rath · ARIA v1.0</span>'
            '</div>',
            unsafe_allow_html=True,
        )


if __name__ == "__main__":
    main()
